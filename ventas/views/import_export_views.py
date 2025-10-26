import xlwt
from openpyxl import load_workbook
from datetime import datetime
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from ..models import Cliente, Company, Contact  # Relative import

# Helper function to check if the user is an administrator
def es_administrador(user):
    return user.is_staff or user.is_superuser

@login_required
def exportar_clientes_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Clientes_{}.xls"'.format(
        datetime.now().strftime('%Y%m%d_%H%M%S')
    )

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Clientes')

    # Estilos
    header_style = xlwt.easyxf('font: bold on; pattern: pattern solid, fore_colour gray25;')

    # Headers
    headers = ['ID', 'Nombre', 'Teléfono', 'Email', 'Documento Identidad', 'Ciudad'] # Added Documento and Ciudad
    for col, header in enumerate(headers):
        ws.write(0, col, header, header_style)
        ws.col(col).width = 256 * 20

    # Obtener todos los clientes
    clientes = Cliente.objects.all().order_by('nombre')

    # Datos
    for row, cliente in enumerate(clientes, 1):
        ws.write(row, 0, cliente.id)
        ws.write(row, 1, cliente.nombre)
        ws.write(row, 2, cliente.telefono or '')
        ws.write(row, 3, cliente.email or '')
        ws.write(row, 4, cliente.documento_identidad or '') # Added
        ws.write(row, 5, cliente.ciudad or '') # Added

    wb.save(response)
    return response

@login_required
@user_passes_test(es_administrador)  # Solo administradores pueden importar
def importar_clientes_excel(request):
    BATCH_SIZE = 500

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        try:
            archivo = request.FILES['archivo_excel']
            # Validate file type (optional but recommended)
            if not archivo.name.endswith(('.xlsx', '.xls')):
                 messages.error(request, 'Formato de archivo inválido. Por favor, suba un archivo Excel (.xlsx o .xls).')
                 return render(request, 'ventas/importar_clientes.html')

            wb = load_workbook(archivo, data_only=True) # data_only=True to get values, not formulas
            ws = wb.active

            clientes_nuevos = []
            clientes_actualizados = []
            errores = []
            total_procesados = 0
            batch_data = []

            # Simple header check (optional)
            expected_headers = ['documento_identidad', 'nombre', 'telefono', 'email', 'ciudad']
            actual_headers = [cell.value.lower().strip().replace(' ', '_') if cell.value else '' for cell in ws[1]]
            # if actual_headers[:len(expected_headers)] != expected_headers:
            #      messages.error(request, f"Las cabeceras del archivo no coinciden. Se esperaban: {', '.join(expected_headers)}")
            #      return render(request, 'ventas/importar_clientes.html')


            for row in ws.iter_rows(min_row=2): # Start from row 2
                # Check if row seems empty
                if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                    continue

                try:
                    # Get values by position (adjust indices if needed)
                    documento_identidad = row[0].value if row[0].value is not None else ''
                    nombre = row[1].value if row[1].value is not None else ''
                    telefono = row[2].value if row[2].value is not None else ''
                    email = row[3].value if row[3].value is not None else ''
                    # Check if row has enough columns before accessing index 4
                    ciudad = row[4].value if len(row) > 4 and row[4].value is not None else ''

                    # Basic validation: Name is required
                    if not str(nombre).strip():
                        errores.append(f"Error en fila {row[0].row}: El nombre es requerido.")
                        continue

                    # Add data to batch
                    batch_data.append({
                        'row': row[0].row, # Use cell's row attribute for accurate reporting
                        'documento_identidad': documento_identidad,
                        'nombre': nombre,
                        'telefono': telefono,
                        'email': email,
                        'ciudad': ciudad
                    })

                    # Process batch if size reached
                    if len(batch_data) >= BATCH_SIZE:
                        process_batch(batch_data, clientes_nuevos, clientes_actualizados, errores)
                        total_procesados += len(batch_data)
                        messages.info(request, f'Procesados {total_procesados} registros...')
                        batch_data = [] # Reset batch

                except Exception as e:
                    # Catch errors during row processing
                    errores.append(f"Error procesando fila {row[0].row if row and row[0] else 'desconocida'}: {str(e)}")

            # Process the final batch
            if batch_data:
                process_batch(batch_data, clientes_nuevos, clientes_actualizados, errores)
                total_procesados += len(batch_data)

            # Display results
            messages.success(request, f'Importación completada. Total procesados: {total_procesados}.')
            if clientes_nuevos:
                messages.info(request, f'Se crearon {len(clientes_nuevos)} nuevos clientes.')
            if clientes_actualizados:
                messages.info(request, f'Se actualizaron {len(clientes_actualizados)} clientes existentes.')
            if errores:
                messages.warning(request, f'Se encontraron {len(errores)} errores durante la importación.')
                # Show first few errors
                for error in errores[:10]:
                    messages.error(request, error)
                if len(errores) > 10:
                    messages.error(request, f'... y {len(errores) - 10} errores más.')

        except Exception as e:
            messages.error(request, f'Error general al procesar el archivo: {str(e)}')
            import traceback
            traceback.print_exc() # Log detailed error

    return render(request, 'ventas/importar_clientes.html')


# --- Helper Functions for Import ---

def limpiar_telefono(telefono):
    """Normaliza a formato E.164 de Chile: +569XXXXXXXX (9 dígitos locales)."""
    try:
        if not telefono:
            return ''
        # Mantener solo dígitos
        digits = ''.join(filter(str.isdigit, str(telefono)))
        # Quitar prefijo país si viene (56...)
        if digits.startswith('56') and len(digits) >= 11:
            # Tomar los últimos 9 dígitos locales (comienzan con 9)
            digits = digits[-9:]
        # Si viene con 8 dígitos, asumir móvil y anteponer 9
        if len(digits) == 8:
            digits = '9' + digits
        # Validar 9 dígitos locales iniciando en 9
        if len(digits) == 9 and digits.startswith('9'):
            return f'+56{digits}'
        return ''
    except Exception:
        return ''

def limpiar_email(email):
    """Limpia y valida el email."""
    if not email: return ''
    email_str = str(email).strip().lower()
    # Basic check for '@' and '.'
    return email_str if '@' in email_str and '.' in email_str.split('@')[-1] else ''

def process_batch(batch_data, clientes_nuevos, clientes_actualizados, errores):
    """Procesa un lote de datos de clientes para crear o actualizar."""
    documentos_en_lote = {str(data['documento_identidad']).strip() for data in batch_data if str(data.get('documento_identidad', '')).strip()}
    telefonos_en_lote = {limpiar_telefono(data['telefono']) for data in batch_data if limpiar_telefono(data.get('telefono', ''))}
    emails_en_lote = {limpiar_email(data['email']) for data in batch_data if limpiar_email(data.get('email', ''))}

    # Find existing clients based on unique fields in the batch
    existing_q = Q()
    if documentos_en_lote: existing_q |= Q(documento_identidad__in=documentos_en_lote)
    if telefonos_en_lote: existing_q |= Q(telefono__in=telefonos_en_lote)
    if emails_en_lote: existing_q |= Q(email__in=emails_en_lote)

    existing_clients_map = {}
    if existing_q:
        for client in Cliente.objects.filter(existing_q):
            if client.documento_identidad: existing_clients_map[f"doc_{client.documento_identidad}"] = client
            if client.telefono: existing_clients_map[f"tel_{client.telefono}"] = client
            if client.email: existing_clients_map[f"email_{client.email}"] = client

    clients_to_create = []
    clients_to_update = []
    processed_in_batch = set() # Track clients processed within this batch to avoid duplicates

    with transaction.atomic():
        for data in batch_data:
            try:
                row_num = data['row']
                documento = str(data.get('documento_identidad', '')).strip()
                nombre = str(data.get('nombre', '')).strip()
                telefono_limpio = limpiar_telefono(data.get('telefono', ''))
                email_limpio = limpiar_email(data.get('email', ''))
                ciudad = str(data.get('ciudad', '')).strip()

                # Skip if essential data missing (already checked name, but double-check)
                if not nombre: continue

                # --- Find existing client ---
                found_client = None
                lookup_keys = []
                if documento: lookup_keys.append(f"doc_{documento}")
                if telefono_limpio: lookup_keys.append(f"tel_{telefono_limpio}")
                if email_limpio: lookup_keys.append(f"email_{email_limpio}")

                for key in lookup_keys:
                    if key in existing_clients_map:
                        found_client = existing_clients_map[key]
                        break # Found a match

                # Avoid processing the same DB client multiple times within the batch
                if found_client and found_client.id in processed_in_batch:
                    continue

                if found_client:
                    # --- Update existing client ---
                    update_fields = []
                    if nombre and found_client.nombre != nombre:
                        found_client.nombre = nombre
                        update_fields.append('nombre')
                    # Update only if new data is provided and different
                    if documento and found_client.documento_identidad != documento:
                        found_client.documento_identidad = documento
                        update_fields.append('documento_identidad')
                    if telefono_limpio and found_client.telefono != telefono_limpio:
                        found_client.telefono = telefono_limpio
                        update_fields.append('telefono')
                    if email_limpio and found_client.email != email_limpio:
                        found_client.email = email_limpio
                        update_fields.append('email')
                    if ciudad and found_client.ciudad != ciudad:
                        found_client.ciudad = ciudad
                        update_fields.append('ciudad')

                    if update_fields:
                        # found_client.save(update_fields=update_fields) # Use update_fields for efficiency
                        clients_to_update.append(found_client) # Add to bulk update list
                        clientes_actualizados.append(f"{nombre} (fila {row_num})")

                    processed_in_batch.add(found_client.id)

                else:
                    # --- Prepare to create new client ---
                    # Check for potential duplicates based on data *within the batch* before adding
                    if (documento and f"doc_{documento}" in existing_clients_map) or \
                       (telefono_limpio and f"tel_{telefono_limpio}" in existing_clients_map) or \
                       (email_limpio and f"email_{email_limpio}" in existing_clients_map):
                        # This data matches a client already processed/found in this batch run, skip creation
                        continue

                    new_client_obj = Cliente(
                        documento_identidad=documento or None, # Use None if empty for unique constraints
                        nombre=nombre,
                        telefono=telefono_limpio or None,
                        email=email_limpio or None,
                        ciudad=ciudad or None
                    )
                    clients_to_create.append(new_client_obj)
                    clientes_nuevos.append(f"{nombre} (fila {row_num})")
                    # Add keys to map to prevent duplicates within the batch creation step
                    if documento: existing_clients_map[f"doc_{documento}"] = new_client_obj # Placeholder
                    if telefono_limpio: existing_clients_map[f"tel_{telefono_limpio}"] = new_client_obj
                    if email_limpio: existing_clients_map[f"email_{email_limpio}"] = new_client_obj


            except Exception as e:
                errores.append(f"Error en fila {data.get('row', 'desconocida')}: {str(e)}")

        # Bulk operations (more efficient for large batches)
        if clients_to_create:
            try:
                Cliente.objects.bulk_create(clients_to_create, ignore_conflicts=True) # ignore_conflicts might hide issues
            except Exception as e:
                 errores.append(f"Error en bulk_create: {str(e)}")
        if clients_to_update:
            try:
                 # Determine all fields that might have been updated across the batch
                 all_update_fields = {'nombre', 'documento_identidad', 'telefono', 'email', 'ciudad'}
                 Cliente.objects.bulk_update(clients_to_update, list(all_update_fields))
            except Exception as e:
                 errores.append(f"Error en bulk_update: {str(e)}")


# ========================
# Importadores CSV (Company & Contact)
# ========================

@login_required
@user_passes_test(es_administrador)
def importar_companies_csv(request):
    """Importa empresas desde un CSV con columnas: name,industry,city,website"""
    BATCH_SIZE = 500
    if request.method == 'POST' and request.FILES.get('archivo_csv'):
        import csv, io
        archivo = request.FILES['archivo_csv']
        if not archivo.name.endswith('.csv'):
            messages.error(request, 'Formato inválido. Suba un archivo .csv')
            return render(request, 'ventas/importar_companies.html')

        try:
            data = archivo.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(data))

            required_cols = {'name'}
            if not required_cols.issubset(set(h.strip() for h in reader.fieldnames or [])):
                messages.error(request, 'El CSV debe incluir al menos la columna "name".')
                return render(request, 'ventas/importar_companies.html')

            to_create = []
            to_update = []
            created, updated = 0, 0
            cache_names = {c['name'] for c in reader if c.get('name')}
            existing = {c.name.lower(): c for c in Company.objects.filter(name__in=list(cache_names))}

            # Reiterar desde el inicio (reader fue consumido para cache_names)
            data_io = io.StringIO(data)
            reader = csv.DictReader(data_io)
            batch = []
            for row in reader:
                name = (row.get('name') or '').strip()
                if not name:
                    continue
                industry = (row.get('industry') or '').strip() or None
                city = (row.get('city') or '').strip() or None
                website = (row.get('website') or '').strip() or None

                existing_obj = existing.get(name.lower())
                if existing_obj:
                    changed = False
                    if industry and getattr(existing_obj, 'industry', None) != industry:
                        setattr(existing_obj, 'industry', industry)
                        changed = True
                    if city and getattr(existing_obj, 'city', None) != city:
                        setattr(existing_obj, 'city', city)
                        changed = True
                    if website and getattr(existing_obj, 'website', None) != website:
                        setattr(existing_obj, 'website', website)
                        changed = True
                    if changed:
                        to_update.append(existing_obj)
                else:
                    obj = Company(name=name)
                    if hasattr(obj, 'industry'): obj.industry = industry
                    if hasattr(obj, 'city'): obj.city = city
                    if hasattr(obj, 'website'): obj.website = website
                    to_create.append(obj)

                if len(to_create) + len(to_update) >= BATCH_SIZE:
                    if to_create:
                        Company.objects.bulk_create(to_create, ignore_conflicts=True)
                        created += len(to_create)
                        to_create = []
                    if to_update:
                        # Actualiza campos comunes si existen en el modelo
                        fields = [f for f in ['industry', 'city', 'website'] if hasattr(Company, f)]
                        if fields:
                            Company.objects.bulk_update(to_update, fields)
                        updated += len(to_update)
                        to_update = []

            if to_create:
                Company.objects.bulk_create(to_create, ignore_conflicts=True)
                created += len(to_create)
            if to_update:
                fields = [f for f in ['industry', 'city', 'website'] if hasattr(Company, f)]
                if fields:
                    Company.objects.bulk_update(to_update, fields)
                updated += len(to_update)

            messages.success(request, f'Empresas importadas. Nuevas: {created}, Actualizadas: {updated}.')
        except Exception as e:
            messages.error(request, f'Error importando CSV: {e}')

    return render(request, 'ventas/importar_companies.html')


@login_required
@user_passes_test(es_administrador)
def importar_contacts_csv(request):
    """Importa contactos desde CSV. Acepta encabezados en inglés o español.
    Columnas soportadas (sin ordenar):
      - first_name | nombre
      - last_name  | apellido
      - email (obligatoria)
      - phone | celular
      - job_title | cargo (opcional)
      - company_name | empresa (opcional)
      - city | ciudad (opcional, se guarda en notas si no hay empresa)
    """
    BATCH_SIZE = 500
    if request.method == 'POST' and request.FILES.get('archivo_csv'):
        import csv, io
        archivo = request.FILES['archivo_csv']
        if not archivo.name.endswith('.csv'):
            messages.error(request, 'Formato inválido. Suba un archivo .csv')
            return render(request, 'ventas/importar_contacts.html')

        try:
            data = archivo.read().decode('utf-8')
            data_io = io.StringIO(data)
            reader = csv.DictReader(data_io)

            headers = [h.strip().lower() for h in (reader.fieldnames or [])]
            if 'email' not in headers:
                messages.error(request, 'El CSV debe incluir la columna "email".')
                return render(request, 'ventas/importar_contacts.html')

            # Utilidades de lectura tolerantes a español/inglés
            def val(row, *keys):
                for k in keys:
                    if k in row and row.get(k) not in (None, ''):
                        return str(row.get(k)).strip()
                return ''

            # Pre-cargar companies existentes
            data_io.seek(0); reader = csv.DictReader(data_io)
            company_names = { val(r, 'company_name', 'empresa') for r in reader if val(r, 'company_name', 'empresa') }
            existing_companies = {c.name.lower(): c for c in Company.objects.filter(name__in=list(company_names))}

            # Pre-cargar contactos por email
            data_io.seek(0); reader = csv.DictReader(data_io)
            emails = { (val(r, 'email') or '').lower() for r in reader if val(r, 'email') }
            existing_contacts = {c.email.lower(): c for c in Contact.objects.filter(email__in=list(emails))}

            data_io.seek(0); reader = csv.DictReader(data_io)

            to_create_c = []
            to_update_c = []
            created, updated = 0, 0

            for row in reader:
                # Campos normalizados
                email = limpiar_email(val(row, 'email'))
                if not email:
                    continue
                first_name = val(row, 'first_name', 'nombre')
                last_name = val(row, 'last_name', 'apellido')
                # Si solo viene "nombre", separarlo en nombre/apellido
                if first_name and not last_name and ' ' in first_name:
                    parts = first_name.split()
                    first_name, last_name = parts[0], ' '.join(parts[1:])
                phone = limpiar_telefono(val(row, 'phone', 'celular')) or None
                job_title = val(row, 'job_title', 'cargo') or None
                industry_val = val(row, 'industry', 'rubro')
                company_name = val(row, 'company_name', 'empresa')
                city_val = val(row, 'city', 'ciudad')

                company_obj = None
                if company_name:
                    company_obj = existing_companies.get(company_name.lower())
                    if not company_obj:
                        company_obj = Company.objects.create(name=company_name)
                        existing_companies[company_name.lower()] = company_obj
                    # Actualizar rubro/industry si viene
                    if industry_val and hasattr(company_obj, 'industry') and getattr(company_obj, 'industry', None) != industry_val:
                        company_obj.industry = industry_val
                        company_obj.save(update_fields=['industry'])

                existing_contact = existing_contacts.get(email)
                if existing_contact:
                    changed = False
                    if first_name and existing_contact.first_name != first_name:
                        existing_contact.first_name = first_name; changed = True
                    if last_name and existing_contact.last_name != last_name:
                        existing_contact.last_name = last_name; changed = True
                    if phone and existing_contact.phone != phone:
                        existing_contact.phone = phone; changed = True
                    if job_title and getattr(existing_contact, 'job_title', None) != job_title:
                        existing_contact.job_title = job_title; changed = True
                    if company_obj and existing_contact.company_id != company_obj.id:
                        existing_contact.company = company_obj; changed = True
                    if changed:
                        to_update_c.append(existing_contact)
                else:
                    notes = ''
                    parts = []
                    if city_val:
                        parts.append(f"Ciudad: {city_val}")
                    if industry_val:
                        parts.append(f"Rubro: {industry_val}")
                    if parts:
                        notes = '; '.join(parts)
                    c = Contact(
                        first_name=first_name or '-',
                        last_name=last_name or '',
                        email=email,
                        phone=phone,
                        job_title=job_title if hasattr(Contact, 'job_title') else None,
                        company=company_obj,
                        notes=notes if hasattr(Contact, 'notes') else None
                    )
                    to_create_c.append(c)

                if len(to_create_c) + len(to_update_c) >= BATCH_SIZE:
                    if to_create_c:
                        Contact.objects.bulk_create(to_create_c, ignore_conflicts=True)
                        created += len(to_create_c)
                        to_create_c = []
                    if to_update_c:
                        fields = [f for f in ['first_name', 'last_name', 'phone', 'job_title', 'company', 'notes'] if hasattr(Contact, f)]
                        Contact.objects.bulk_update(to_update_c, fields)
                        updated += len(to_update_c)
                        to_update_c = []

            if to_create_c:
                Contact.objects.bulk_create(to_create_c, ignore_conflicts=True)
                created += len(to_create_c)
            if to_update_c:
                fields = [f for f in ['first_name', 'last_name', 'phone', 'job_title', 'company', 'notes'] if hasattr(Contact, f)]
                Contact.objects.bulk_update(to_update_c, fields)
                updated += len(to_update_c)

            messages.success(request, f'Contactos importados. Nuevos: {created}, Actualizados: {updated}.')
        except Exception as e:
            messages.error(request, f'Error importando CSV: {e}')

    return render(request, 'ventas/importar_contacts.html')
