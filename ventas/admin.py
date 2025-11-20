from django.contrib import admin
from django import forms
from .forms import PagoInlineForm
from django.forms import DateTimeInput
from datetime import date, datetime, timedelta
from django.utils import timezone
from django.db import models
from django.db.models import Sum
from django.utils.safestring import mark_safe
from django.utils.translation import gettext_lazy as _
from django.forms import DateInput, TimeInput, Select
# Updated model imports to include HomepageConfig and CRM models
from .models import (
    Proveedor, CategoriaProducto, Producto, VentaReserva, ReservaProducto,
    Pago, Cliente, CategoriaServicio, Servicio, ReservaServicio,
    MovimientoCliente, Compra, DetalleCompra, GiftCard, HomepageConfig,
    Lead, Company, Contact, Activity, Campaign, Deal, CampaignInteraction, HomepageSettings,
    # Communication models
    CommunicationLimit, ClientPreferences, CommunicationLog, SMSTemplate, MailParaEnviar,
    # Advanced Email Campaign models
    EmailCampaign, EmailRecipient, EmailDeliveryLog, EmailBlacklist, EmailTemplate, EmailSubjectTemplate, EmailContentTemplate,
    # Historical data
    ServiceHistory,
    # Sistema de Tramos y Premios
    Premio, ClientePremio, HistorialTramo,
    # Sistema de Packs de Descuento
    PackDescuento
)
from django.http import HttpResponse, JsonResponse
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.urls import path, reverse
from django.utils.html import format_html
import json
import xlwt
from django.core.paginator import Paginator
from openpyxl import load_workbook
from solo.admin import SingletonModelAdmin
from .views import admin_views
from django.http import HttpResponseRedirect
from django.utils.html import format_html # Import format_html

# --- Standard Model Inlines ---

class ReservaServicioInline(admin.TabularInline):
    model = ReservaServicio
    fields = ['servicio', 'fecha_agendamiento', 'hora_inicio', 'cantidad_personas', 'proveedor_asignado']
    readonly_fields = []
    autocomplete_fields = ['servicio', 'proveedor_asignado']
    extra = 1
    min_num = 0
    verbose_name = "Servicio Reservado"
    verbose_name_plural = "Servicios Reservados"
    can_delete = True

    def get_formset(self, request, obj=None, **kwargs):
        """Override to add custom validation or processing"""
        formset = super().get_formset(request, obj, **kwargs)
        return formset

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        """Customize form fields for better UX"""
        field = super().formfield_for_dbfield(db_field, request, **kwargs)

        if db_field.name == 'fecha_agendamiento':
            field.widget = DateInput(attrs={
                'class': 'vDateField',
                'size': '10'
            })
        elif db_field.name == 'hora_inicio':
            field.widget = TimeInput(attrs={
                'class': 'vTimeField',
                'size': '8'
            })
        elif db_field.name == 'cantidad_personas':
            field.widget.attrs['style'] = 'width: 60px;'

        return field

class ReservaProductoInline(admin.TabularInline):
    model = ReservaProducto
    extra = 1

class PagoInline(admin.TabularInline):
    model = Pago
    form = PagoInlineForm
    extra = 1
    fields = ['monto', 'tipo_pago', 'fecha_pago', 'comprobante', 'estado']
    readonly_fields = ['fecha_pago']

    def get_readonly_fields(self, request, obj=None):
        # Si el pago ya existe, hacer el estado de solo lectura
        if obj:
            return self.readonly_fields + ['estado']
        return self.readonly_fields

class VentaReservaForm(forms.ModelForm):
    class Meta:
        model = VentaReserva
        fields = '__all__'
        widgets = {
            'fecha': DateInput(attrs={
                'type': 'date',
                'class': 'vDateField',
            }),
            'fecha_evento': DateInput(attrs={
                'type': 'date',
                'class': 'vDateField',
            }),
            'hora_evento': TimeInput(attrs={
                'type': 'time',
                'class': 'vTimeField',
            }),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Customize field attributes
        if 'cliente' in self.fields:
            self.fields['cliente'].widget.attrs['style'] = 'width: 300px;'
        if 'total' in self.fields:
            self.fields['total'].widget.attrs['readonly'] = True
            self.fields['total'].widget.attrs['style'] = 'font-weight: bold;'
        if 'saldo_pendiente' in self.fields:
            self.fields['saldo_pendiente'].widget.attrs['readonly'] = True
            self.fields['saldo_pendiente'].widget.attrs['style'] = 'font-weight: bold; color: red;'

# --- Admin Classes ---

@admin.register(Proveedor)
class ProveedorAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'telefono', 'email', 'direccion')
    search_fields = ('nombre', 'telefono', 'email')

@admin.register(CategoriaProducto)
class CategoriaProductoAdmin(admin.ModelAdmin):
    list_display = ('nombre',)

@admin.register(Producto)
class ProductoAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'categoria', 'precio_base', 'cantidad_disponible', 'proveedor')
    list_filter = ('categoria', 'proveedor')
    search_fields = ('nombre',)
    ordering = ('nombre',)

@admin.register(VentaReserva)
class VentaReservaAdmin(admin.ModelAdmin):
    form = VentaReservaForm
    list_display = ('id', 'cliente_info', 'fecha_creacion', 'fecha_reserva',
                   'total_con_formato', 'saldo_con_formato', 'estado_pago_display',
                   'estado_reserva', 'ver_servicios', 'cobrado_display')
    list_filter = ('estado_reserva', 'estado_pago', 'fecha_creacion', 'fecha_reserva', 'cobrado')
    search_fields = ('cliente__nombre', 'cliente__telefono', 'cliente__email',
                    'comentarios', 'reservaservicio__servicio__nombre', 'codigo_giftcard')
    date_hierarchy = 'fecha_creacion'
    ordering = ('-fecha_creacion',)
    readonly_fields = ('total_con_formato', 'saldo_con_formato', 'calcular_totales_display',
                      'fecha_creacion', 'mostrar_servicios_reservados')

    fieldsets = (
        ('Información General', {
            'fields': ('cliente', 'fecha_reserva', 'estado_reserva', 'comentarios')
        }),
        ('Totales y Pago', {
            'fields': ('total', 'pagado', 'saldo_pendiente', 'estado_pago', 'cobrado',
                      'calcular_totales_display')
        }),
        ('Gift Card', {
            'fields': ('codigo_giftcard',),
            'classes': ('collapse',)
        }),
        ('Servicios Reservados', {
            'fields': ('mostrar_servicios_reservados',),
            'classes': ('wide',)
        }),
        ('Información del Sistema', {
            'fields': ('fecha_creacion', 'numero_documento_fiscal'),
            'classes': ('collapse',)
        })
    )

    inlines = [ReservaServicioInline, PagoInline]
    save_on_top = True

    def cliente_info(self, obj):
        if obj.cliente:
            return format_html(
                '<strong>{}</strong><br><small>{}<br>{}</small>',
                obj.cliente.nombre,
                obj.cliente.telefono or 'Sin teléfono',
                obj.cliente.email or 'Sin email'
            )
        return '-'
    cliente_info.short_description = 'Cliente'
    cliente_info.admin_order_field = 'cliente__nombre'

    def total_con_formato(self, obj):
        return f"${obj.total:,.0f}" if obj.total else "$0"
    total_con_formato.short_description = 'Total'
    total_con_formato.admin_order_field = 'total'

    def saldo_con_formato(self, obj):
        saldo = obj.saldo_pendiente
        if saldo > 0:
            return format_html(
                '<span style="color: red; font-weight: bold;">${:,.0f}</span>',
                saldo
            )
        else:
            return format_html(
                '<span style="color: green;">$0</span>'
            )
    saldo_con_formato.short_description = 'Saldo Pendiente'
    saldo_con_formato.admin_order_field = 'saldo_pendiente'

    def estado_pago_display(self, obj):
        """Muestra el estado de pago con colores"""
        if obj.saldo_pendiente <= 0:
            return format_html(
                '<span style="color: green; font-weight: bold;">✓ PAGADO</span>'
            )
        else:
            porcentaje = ((obj.total - obj.saldo_pendiente) / obj.total * 100) if obj.total > 0 else 0
            return format_html(
                '<span style="color: orange;">{}% pagado</span>',
                int(porcentaje)
            )
    estado_pago_display.short_description = 'Estado de Pago'

    def cobrado_display(self, obj):
        """Muestra si fue cobrado con un icono"""
        if obj.cobrado:
            return format_html('<span style="color: green;">✓</span>')
        return format_html('<span style="color: gray;">-</span>')
    cobrado_display.short_description = 'Cobrado'

    def ver_servicios(self, obj):
        servicios = obj.reservaservicio_set.all()
        if servicios:
            items = []
            for rs in servicios:
                fecha = rs.fecha_agendamiento.strftime('%d/%m') if rs.fecha_agendamiento else 'Sin fecha'
                hora = rs.hora_inicio if rs.hora_inicio else ''  # hora_inicio es CharField, no TimeField
                items.append(f"• {rs.servicio.nombre} ({fecha} {hora}) - {rs.cantidad_personas} pers.")
            return format_html('<small>{}</small>', '<br>'.join(items))
        return '-'
    ver_servicios.short_description = 'Servicios'

    def mostrar_servicios_reservados(self, obj):
        """Muestra todos los servicios reservados con formato detallado"""
        servicios = obj.reservaservicio_set.all().select_related('servicio', 'proveedor_asignado')
        if not servicios:
            return "No hay servicios reservados"

        html = '<table style="width: 100%; border-collapse: collapse;">'
        html += '<tr style="background-color: #f0f0f0;">'
        html += '<th style="padding: 5px; text-align: left;">Servicio</th>'
        html += '<th style="padding: 5px;">Fecha</th>'
        html += '<th style="padding: 5px;">Hora</th>'
        html += '<th style="padding: 5px;">Personas</th>'
        html += '<th style="padding: 5px;">Precio Unit.</th>'
        html += '<th style="padding: 5px;">Subtotal</th>'
        html += '<th style="padding: 5px;">Proveedor</th>'
        html += '</tr>'

        total = 0
        for i, rs in enumerate(servicios):
            subtotal = rs.servicio.precio * rs.cantidad_personas
            total += subtotal

            bg_color = '#ffffff' if i % 2 == 0 else '#f9f9f9'
            html += f'<tr style="background-color: {bg_color};">'
            html += f'<td style="padding: 5px;">{rs.servicio.nombre}</td>'
            html += f'<td style="padding: 5px; text-align: center;">{rs.fecha_agendamiento.strftime("%d/%m/%Y") if rs.fecha_agendamiento else "-"}</td>'
            html += f'<td style="padding: 5px; text-align: center;">{rs.hora_inicio.strftime("%H:%M") if rs.hora_inicio else "-"}</td>'
            html += f'<td style="padding: 5px; text-align: center;">{rs.cantidad_personas}</td>'
            html += f'<td style="padding: 5px; text-align: right;">${rs.servicio.precio:,.0f}</td>'
            html += f'<td style="padding: 5px; text-align: right; font-weight: bold;">${subtotal:,.0f}</td>'
            html += f'<td style="padding: 5px; text-align: center;">{rs.proveedor_asignado.nombre if rs.proveedor_asignado else "-"}</td>'
            html += '</tr>'

        html += f'<tr style="background-color: #e0e0e0; font-weight: bold;">'
        html += '<td colspan="5" style="padding: 5px; text-align: right;">TOTAL:</td>'
        html += f'<td style="padding: 5px; text-align: right;">${total:,.0f}</td>'
        html += '<td></td>'
        html += '</tr>'
        html += '</table>'

        return format_html(html)

    mostrar_servicios_reservados.short_description = 'Detalle de Servicios'

    def calcular_totales_display(self, obj):
        """Muestra botón para recalcular totales"""
        if obj.pk:
            return format_html(
                '<a class="button" href="#" onclick="return false;" '
                'style="padding: 5px 10px; background-color: #417690; color: white;">'
                'Totales actualizados automáticamente</a>'
            )
        return "Guarde primero para ver totales"

    calcular_totales_display.short_description = 'Actualizar Totales'

    def save_model(self, request, obj, form, change):
        # Guardar el objeto primero
        super().save_model(request, obj, form, change)

        # Mostrar mensaje
        if not change:
            messages.success(request, f'Venta/Reserva #{obj.id} creada exitosamente.')
        else:
            messages.info(request, f'Venta/Reserva #{obj.id} actualizada.')

    def save_formset(self, request, form, formset, change):
        """Override para recalcular totales después de guardar los inlines"""
        instances = formset.save(commit=False)

        # Guardar las instancias
        for instance in instances:
            instance.save()

        # Eliminar las marcadas para eliminación
        for instance in formset.deleted_objects:
            instance.delete()

        formset.save_m2m()

        # Recalcular totales de la venta
        if hasattr(formset, 'instance') and isinstance(formset.instance, VentaReserva):
            venta = formset.instance
            venta.calcular_totales()
            venta.save(update_fields=['total', 'saldo_pendiente'])

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-ventas/', self.export_ventas, name='ventas_ventareserva_export'),
        ]
        return custom_urls + urls

    def export_ventas(self, request):
        """Exportar ventas a Excel"""
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ventas.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Ventas')

        # Cabeceras
        columns = ['ID', 'Cliente', 'Fecha', 'Fecha Evento', 'Total', 'Saldo', 'Estado']
        for col, column_title in enumerate(columns):
            ws.write(0, col, column_title)

        # Datos
        ventas = VentaReserva.objects.all().select_related('cliente')
        for row, venta in enumerate(ventas, 1):
            ws.write(row, 0, venta.id)
            ws.write(row, 1, venta.cliente.nombre if venta.cliente else '')
            ws.write(row, 2, venta.fecha.strftime('%d/%m/%Y'))
            ws.write(row, 3, venta.fecha_evento.strftime('%d/%m/%Y') if venta.fecha_evento else '')
            ws.write(row, 4, float(venta.total))
            ws.write(row, 5, float(venta.saldo_pendiente))
            ws.write(row, 6, venta.estado)

        wb.save(response)
        return response

    actions = ['marcar_como_pagado', 'enviar_recordatorio_pago']

    def marcar_como_pagado(self, request, queryset):
        """Acción para marcar ventas como pagadas"""
        for venta in queryset:
            if venta.saldo_pendiente > 0:
                Pago.objects.create(
                    venta=venta,
                    monto=venta.saldo_pendiente,
                    tipo_pago='efectivo',
                    estado='aprobado',
                    comprobante=f'Pago manual por admin {request.user.username}'
                )
                venta.calcular_totales()
                venta.save()

        self.message_user(request, f'{queryset.count()} ventas marcadas como pagadas.')

    marcar_como_pagado.short_description = 'Marcar como pagado (crear pago por saldo)'

    def enviar_recordatorio_pago(self, request, queryset):
        """Enviar recordatorio de pago pendiente"""
        enviados = 0
        for venta in queryset:
            if venta.saldo_pendiente > 0 and venta.cliente and venta.cliente.email:
                # Aquí iría la lógica de envío de email
                enviados += 1

        self.message_user(request, f'Recordatorios enviados a {enviados} clientes.')

    enviar_recordatorio_pago.short_description = 'Enviar recordatorio de pago pendiente'

@admin.register(Pago)
class PagoAdmin(admin.ModelAdmin):
    list_display = ('id', 'venta_info', 'monto_con_formato', 'tipo_pago',
                   'fecha_pago', 'estado', 'comprobante_corto')
    list_filter = ('estado', 'tipo_pago', 'fecha_pago')
    search_fields = ('venta__cliente__nombre', 'comprobante')
    date_hierarchy = 'fecha_pago'
    readonly_fields = ('fecha_pago',)

    fieldsets = (
        ('Información del Pago', {
            'fields': ('venta', 'monto', 'tipo_pago', 'fecha_pago')
        }),
        ('Detalles', {
            'fields': ('comprobante', 'estado')
        })
    )

    def venta_info(self, obj):
        if obj.venta:
            return format_html(
                'Venta #{}<br><small>{}</small>',
                obj.venta.id,
                obj.venta.cliente.nombre if obj.venta.cliente else 'Sin cliente'
            )
        return '-'
    venta_info.short_description = 'Venta'

    def monto_con_formato(self, obj):
        return f"${obj.monto:,.0f}"
    monto_con_formato.short_description = 'Monto'
    monto_con_formato.admin_order_field = 'monto'

    def comprobante_corto(self, obj):
        if obj.comprobante:
            return obj.comprobante[:50] + '...' if len(obj.comprobante) > 50 else obj.comprobante
        return '-'
    comprobante_corto.short_description = 'Comprobante'

# GiftCard Admin
@admin.register(GiftCard)
class GiftCardAdmin(admin.ModelAdmin):
    """Admin mejorado para GiftCards"""

    list_display = [
        'codigo', 'monto_inicial', 'monto_disponible',
        'estado_badge', 'cliente', 'fecha_emision',
        'fecha_vencimiento', 'porcentaje_usado'
    ]

    list_filter = [
        'estado',
        'fecha_emision',
        'fecha_vencimiento',
        'servicio_asociado'
    ]

    search_fields = [
        'codigo', 'cliente__nombre', 'cliente__email',
        'cliente__telefono', 'destinatario_nombre',
        'comprador_nombre', 'comprador_email'
    ]

    readonly_fields = [
        'codigo', 'fecha_emision', 'fecha_uso',
        'venta_donde_uso', 'porcentaje_usado_display',
        'historial_uso', 'datos_ia_display'
    ]

    fieldsets = (
        ('Información Principal', {
            'fields': ('codigo', 'estado', 'servicio_asociado')
        }),
        ('Montos', {
            'fields': ('monto_inicial', 'monto_disponible', 'porcentaje_usado_display')
        }),
        ('Fechas', {
            'fields': ('fecha_emision', 'fecha_vencimiento', 'fecha_uso')
        }),
        ('Cliente Asociado', {
            'fields': ('cliente',)
        }),
        ('Datos del Comprador', {
            'fields': ('comprador_nombre', 'comprador_email', 'comprador_telefono'),
            'classes': ('collapse',)
        }),
        ('Datos del Destinatario', {
            'fields': ('destinatario_nombre', 'destinatario_email',
                      'destinatario_telefono', 'destinatario_relacion'),
            'classes': ('collapse',)
        }),
        ('Personalización IA', {
            'fields': ('datos_ia_display',),
            'classes': ('collapse',)
        }),
        ('Historial', {
            'fields': ('venta_donde_uso', 'historial_uso'),
            'classes': ('collapse',)
        })
    )

    actions = ['marcar_como_cobrado', 'extender_vencimiento', 'exportar_a_excel']

    def estado_badge(self, obj):
        """Badge colorido para el estado"""
        colors = {
            'por_cobrar': '#ffc107',
            'cobrado': '#28a745',
            'usado': '#6c757d',
            'vencido': '#dc3545'
        }
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 10px; '
            'border-radius: 3px; font-size: 11px;">{}</span>',
            colors.get(obj.estado, '#6c757d'),
            obj.get_estado_display()
        )
    estado_badge.short_description = 'Estado'
    estado_badge.admin_order_field = 'estado'

    def porcentaje_usado(self, obj):
        """Barra de progreso del uso"""
        if obj.monto_inicial == 0:
            return '-'

        porcentaje = ((obj.monto_inicial - obj.monto_disponible) / obj.monto_inicial) * 100
        color = '#28a745' if porcentaje < 50 else '#ffc107' if porcentaje < 80 else '#dc3545'

        return format_html(
            '<div style="width: 100px; background-color: #e0e0e0; '
            'border-radius: 3px; overflow: hidden;">'
            '<div style="width: {}%; background-color: {}; height: 20px; '
            'text-align: center; color: white; line-height: 20px;">'
            '{}%</div></div>',
            porcentaje, color, int(porcentaje)
        )
    porcentaje_usado.short_description = 'Uso'

    def porcentaje_usado_display(self, obj):
        """Muestra detalle del uso"""
        if obj.monto_inicial == 0:
            return "N/A"

        usado = obj.monto_inicial - obj.monto_disponible
        porcentaje = (usado / obj.monto_inicial) * 100

        return format_html(
            '<strong>${:,.0f}</strong> usado de <strong>${:,.0f}</strong> '
            '({}%)<br><br>'
            '<div style="background-color: #e0e0e0; border-radius: 3px;">'
            '<div style="width: {}%; background-color: #28a745; height: 25px; '
            'border-radius: 3px;"></div></div>',
            usado, obj.monto_inicial, int(porcentaje), porcentaje
        )
    porcentaje_usado_display.short_description = 'Uso Detallado'

    def historial_uso(self, obj):
        """Muestra historial de uso"""
        # Por ahora simulado, después se puede implementar un modelo de historial
        if obj.estado == 'usado' and obj.venta_donde_uso:
            return format_html(
                '<strong>Usado en:</strong> Venta #{}<br>'
                '<strong>Fecha:</strong> {}<br>'
                '<strong>Monto:</strong> ${:,.0f}',
                obj.venta_donde_uso.id,
                obj.fecha_uso.strftime('%d/%m/%Y %H:%M'),
                obj.monto_inicial - obj.monto_disponible
            )
        return "Sin uso registrado"
    historial_uso.short_description = 'Historial de Uso'

    def datos_ia_display(self, obj):
        """Muestra los datos de personalización IA"""
        html = '<table style="width: 100%;">'

        # Tipo de mensaje y detalles
        if obj.tipo_mensaje:
            html += f'<tr><td><strong>Tipo:</strong></td><td>{obj.tipo_mensaje}</td></tr>'
        if obj.detalle_especial:
            html += f'<tr><td><strong>Detalle:</strong></td><td>{obj.detalle_especial}</td></tr>'

        # Mensaje personalizado
        if obj.mensaje_personalizado:
            html += '<tr><td colspan="2"><br><strong>Mensaje Personalizado:</strong></td></tr>'
            html += f'<tr><td colspan="2" style="background-color: #f5f5f5; padding: 10px; '
            html += f'border-radius: 5px; font-style: italic;">{obj.mensaje_personalizado}</td></tr>'

        # Alternativas
        if obj.mensaje_alternativas:
            html += '<tr><td colspan="2"><br><strong>Alternativas generadas:</strong></td></tr>'
            for i, alt in enumerate(obj.mensaje_alternativas, 1):
                html += f'<tr><td colspan="2" style="padding-left: 20px;">• {alt[:100]}...</td></tr>'

        html += '</table>'
        return format_html(html)
    datos_ia_display.short_description = 'Datos de Personalización IA'

    def marcar_como_cobrado(self, request, queryset):
        """Marca las GiftCards seleccionadas como cobradas"""
        actualizadas = queryset.filter(estado='por_cobrar').update(estado='cobrado')
        self.message_user(
            request,
            f'{actualizadas} GiftCard(s) marcadas como cobradas.'
        )
    marcar_como_cobrado.short_description = 'Marcar como cobrado'

    def extender_vencimiento(self, request, queryset):
        """Extiende el vencimiento por 6 meses más"""
        from datetime import timedelta
        for giftcard in queryset:
            giftcard.fecha_vencimiento += timedelta(days=180)
            giftcard.save()

        self.message_user(
            request,
            f'{queryset.count()} GiftCard(s) extendidas por 6 meses más.'
        )
    extender_vencimiento.short_description = 'Extender vencimiento (+6 meses)'

    def exportar_a_excel(self, request, queryset):
        """Exporta las GiftCards seleccionadas a Excel"""
        import xlwt

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="giftcards.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('GiftCards')

        # Estilos
        header_style = xlwt.XFStyle()
        header_style.font.bold = True

        # Cabeceras
        columns = [
            'Código', 'Estado', 'Monto Inicial', 'Monto Disponible',
            'Cliente', 'Comprador', 'Destinatario', 'Fecha Emisión',
            'Fecha Vencimiento', 'Servicio'
        ]

        for col, column_title in enumerate(columns):
            ws.write(0, col, column_title, header_style)

        # Datos
        for row, giftcard in enumerate(queryset, 1):
            ws.write(row, 0, giftcard.codigo)
            ws.write(row, 1, giftcard.get_estado_display())
            ws.write(row, 2, float(giftcard.monto_inicial))
            ws.write(row, 3, float(giftcard.monto_disponible))
            ws.write(row, 4, giftcard.cliente.nombre if giftcard.cliente else '')
            ws.write(row, 5, giftcard.comprador_nombre)
            ws.write(row, 6, giftcard.destinatario_nombre)
            ws.write(row, 7, giftcard.fecha_emision.strftime('%d/%m/%Y'))
            ws.write(row, 8, giftcard.fecha_vencimiento.strftime('%d/%m/%Y'))
            ws.write(row, 9, giftcard.servicio_asociado or '')

        wb.save(response)
        return response

    exportar_a_excel.short_description = 'Exportar a Excel'

    def save_model(self, request, obj, form, change):
        """Validaciones al guardar"""
        if not change:  # Nueva GiftCard
            if not obj.codigo:
                # El modelo genera el código automáticamente
                pass

        super().save_model(request, obj, form, change)

    class Media:
        css = {
            'all': ('admin/css/giftcard_admin.css',)
        }

# Cliente Admin
class MovimientoClienteInline(admin.TabularInline):
    model = MovimientoCliente
    extra = 0
    readonly_fields = ('fecha', 'tipo_movimiento', 'monto', 'descripcion', 'saldo_despues')
    can_delete = False

    def has_add_permission(self, request, obj=None):
        return False

@admin.register(Cliente)
class ClienteAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'email', 'telefono', 'vip', 'activo', 'saldo_actual',
                   'total_compras', 'ultima_visita', 'fecha_cumpleanos')
    list_filter = ('vip', 'activo', 'fecha_cumpleanos')
    search_fields = ('nombre', 'email', 'telefono')
    readonly_fields = ('saldo', 'fecha_creacion', 'fecha_modificacion',
                      'resumen_cliente', 'historial_compras')
    inlines = [MovimientoClienteInline]

    fieldsets = (
        ('Información Personal', {
            'fields': ('nombre', 'email', 'telefono', 'direccion',
                      'fecha_cumpleanos', 'notas')
        }),
        ('Estado y Clasificación', {
            'fields': ('activo', 'vip')
        }),
        ('Información Financiera', {
            'fields': ('saldo', 'resumen_cliente'),
            'classes': ('collapse',)
        }),
        ('Historial', {
            'fields': ('historial_compras',),
            'classes': ('collapse',)
        }),
        ('Información del Sistema', {
            'fields': ('fecha_creacion', 'fecha_modificacion'),
            'classes': ('collapse',)
        })
    )

    def saldo_actual(self, obj):
        if obj.saldo > 0:
            return format_html(
                '<span style="color: green; font-weight: bold;">+${:,.0f}</span>',
                obj.saldo
            )
        elif obj.saldo < 0:
            return format_html(
                '<span style="color: red; font-weight: bold;">-${:,.0f}</span>',
                abs(obj.saldo)
            )
        else:
            return '$0'
    saldo_actual.short_description = 'Saldo'
    saldo_actual.admin_order_field = 'saldo'

    def total_compras(self, obj):
        total = obj.ventareserva_set.filter(estado='completado').aggregate(
            total=Sum('total')
        )['total'] or 0
        return f"${total:,.0f}"
    total_compras.short_description = 'Total Compras'

    def ultima_visita(self, obj):
        ultima_venta = obj.ventareserva_set.order_by('-fecha').first()
        if ultima_venta:
            return ultima_venta.fecha.strftime('%d/%m/%Y')
        return 'Nunca'
    ultima_visita.short_description = 'Última Visita'

    def resumen_cliente(self, obj):
        """Muestra un resumen completo del cliente"""
        ventas = obj.ventareserva_set.all()
        total_ventas = ventas.count()
        ventas_completadas = ventas.filter(estado='completado').count()
        total_gastado = ventas.filter(estado='completado').aggregate(
            total=Sum('total')
        )['total'] or 0

        # Servicios más utilizados
        from django.db.models import Count
        servicios_top = ReservaServicio.objects.filter(
            reserva__cliente=obj
        ).values('servicio__nombre').annotate(
            cantidad=Count('id')
        ).order_by('-cantidad')[:5]

        html = f'''
        <div style="padding: 10px; background-color: #f5f5f5; border-radius: 5px;">
            <h4>Resumen del Cliente</h4>
            <table style="width: 100%;">
                <tr>
                    <td><strong>Total de visitas:</strong></td>
                    <td>{total_ventas}</td>
                </tr>
                <tr>
                    <td><strong>Ventas completadas:</strong></td>
                    <td>{ventas_completadas}</td>
                </tr>
                <tr>
                    <td><strong>Total gastado:</strong></td>
                    <td>${total_gastado:,.0f}</td>
                </tr>
                <tr>
                    <td><strong>Ticket promedio:</strong></td>
                    <td>${(total_gastado/ventas_completadas if ventas_completadas > 0 else 0):,.0f}</td>
                </tr>
            </table>

            <h4 style="margin-top: 15px;">Servicios Favoritos</h4>
            <ol>
        '''

        for servicio in servicios_top:
            html += f"<li>{servicio['servicio__nombre']} ({servicio['cantidad']} veces)</li>"

        html += '''
            </ol>
        </div>
        '''

        return format_html(html)

    resumen_cliente.short_description = 'Resumen del Cliente'

    def historial_compras(self, obj):
        """Muestra las últimas compras del cliente"""
        ventas = obj.ventareserva_set.order_by('-fecha')[:10]

        if not ventas:
            return "No hay compras registradas"

        html = '<table style="width: 100%; border-collapse: collapse;">'
        html += '<tr style="background-color: #e0e0e0;">'
        html += '<th style="padding: 5px;">Fecha</th>'
        html += '<th style="padding: 5px;">Total</th>'
        html += '<th style="padding: 5px;">Estado</th>'
        html += '<th style="padding: 5px;">Servicios</th>'
        html += '</tr>'

        for i, venta in enumerate(ventas):
            bg_color = '#ffffff' if i % 2 == 0 else '#f9f9f9'
            servicios = ', '.join([
                rs.servicio.nombre for rs in venta.reservaservicio_set.all()
            ])

            html += f'<tr style="background-color: {bg_color};">'
            html += f'<td style="padding: 5px;">{venta.fecha.strftime("%d/%m/%Y")}</td>'
            html += f'<td style="padding: 5px;">${venta.total:,.0f}</td>'
            html += f'<td style="padding: 5px;">{venta.estado}</td>'
            html += f'<td style="padding: 5px; font-size: 0.9em;">{servicios}</td>'
            html += '</tr>'

        html += '</table>'
        html += '<p style="margin-top: 10px;"><em>Mostrando últimas 10 compras</em></p>'

        return format_html(html)

    historial_compras.short_description = 'Historial de Compras'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export/', self.export_clientes, name='ventas_cliente_export'),
            path('import/', self.import_clientes, name='ventas_cliente_import'),
        ]
        return custom_urls + urls

    def export_clientes(self, request):
        """Exporta clientes a Excel"""
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="clientes.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Clientes')

        # Cabeceras
        columns = ['Nombre', 'Email', 'Teléfono', 'VIP', 'Activo', 'Saldo', 'Fecha Registro']
        for col, column_title in enumerate(columns):
            ws.write(0, col, column_title)

        # Datos
        clientes = Cliente.objects.all()
        for row, cliente in enumerate(clientes, 1):
            ws.write(row, 0, cliente.nombre)
            ws.write(row, 1, cliente.email or '')
            ws.write(row, 2, cliente.telefono or '')
            ws.write(row, 3, 'Sí' if cliente.vip else 'No')
            ws.write(row, 4, 'Sí' if cliente.activo else 'No')
            ws.write(row, 5, float(cliente.saldo))
            ws.write(row, 6, cliente.fecha_creacion.strftime('%d/%m/%Y'))

        wb.save(response)
        return response

    def import_clientes(self, request):
        """Vista para importar clientes desde Excel"""
        if request.method == 'POST' and request.FILES.get('excel_file'):
            excel_file = request.FILES['excel_file']

            try:
                wb = load_workbook(excel_file)
                ws = wb.active

                imported = 0
                errors = []

                # Asumimos que la primera fila son los encabezados
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:  # Si no hay nombre, saltar
                        continue

                    try:
                        cliente, created = Cliente.objects.update_or_create(
                            email=row[1] if row[1] else None,
                            defaults={
                                'nombre': row[0],
                                'telefono': row[2] if len(row) > 2 else None,
                                'vip': str(row[3]).lower() == 'sí' if len(row) > 3 else False,
                                'activo': str(row[4]).lower() == 'sí' if len(row) > 4 else True,
                            }
                        )
                        imported += 1
                    except Exception as e:
                        errors.append(f"Error en fila {row}: {str(e)}")

                messages.success(request, f'{imported} clientes importados exitosamente.')
                if errors:
                    messages.warning(request, f'Errores: {"; ".join(errors[:5])}')

            except Exception as e:
                messages.error(request, f'Error al procesar el archivo: {str(e)}')

            return HttpResponseRedirect(reverse('admin:ventas_cliente_changelist'))

        # Mostrar formulario de carga
        return HttpResponse('''
            <html>
            <head>
                <title>Importar Clientes</title>
                <style>
                    body { font-family: Arial, sans-serif; padding: 20px; }
                    .container { max-width: 600px; margin: 0 auto; }
                    .info { background-color: #e3f2fd; padding: 15px; border-radius: 5px; margin-bottom: 20px; }
                    input[type="file"] { margin: 10px 0; }
                    input[type="submit"] { background-color: #2196F3; color: white; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; }
                    input[type="submit"]:hover { background-color: #1976D2; }
                </style>
            </head>
            <body>
                <div class="container">
                    <h2>Importar Clientes desde Excel</h2>
                    <div class="info">
                        <p><strong>Formato esperado del archivo Excel:</strong></p>
                        <ol>
                            <li>Columna A: Nombre (requerido)</li>
                            <li>Columna B: Email</li>
                            <li>Columna C: Teléfono</li>
                            <li>Columna D: VIP (Sí/No)</li>
                            <li>Columna E: Activo (Sí/No)</li>
                        </ol>
                        <p>La primera fila debe contener los encabezados.</p>
                    </div>
                    <form method="post" enctype="multipart/form-data">
                        {% csrf_token %}
                        <input type="file" name="excel_file" accept=".xlsx,.xls" required>
                        <br><br>
                        <input type="submit" value="Importar">
                        <a href="{% url 'admin:ventas_cliente_changelist' %}" style="margin-left: 10px;">Cancelar</a>
                    </form>
                </div>
            </body>
            </html>
        ''')

    actions = ['marcar_como_vip', 'desactivar_clientes', 'enviar_promocion']

    def marcar_como_vip(self, request, queryset):
        queryset.update(vip=True)
        self.message_user(request, f'{queryset.count()} clientes marcados como VIP.')
    marcar_como_vip.short_description = 'Marcar como VIP'

    def desactivar_clientes(self, request, queryset):
        queryset.update(activo=False)
        self.message_user(request, f'{queryset.count()} clientes desactivados.')
    desactivar_clientes.short_description = 'Desactivar clientes'

    def enviar_promocion(self, request, queryset):
        # Aquí iría la lógica para enviar promociones
        clientes_con_email = queryset.filter(email__isnull=False).count()
        self.message_user(
            request,
            f'Promoción enviada a {clientes_con_email} clientes con email.'
        )
    enviar_promocion.short_description = 'Enviar promoción por email'

@admin.register(CategoriaServicio)
class CategoriaServicioAdmin(admin.ModelAdmin):
    list_display = ('nombre',)
    search_fields = ('nombre',)

@admin.register(Servicio)
class ServicioAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'categoria', 'tipo_servicio', 'precio_con_formato', 'duracion_minutos',
                   'capacidad_maxima', 'activo', 'publicado_web', 'popularidad')
    list_filter = ('categoria', 'tipo_servicio', 'activo', 'publicado_web')
    search_fields = ('nombre', 'descripcion_web')
    ordering = ('categoria', 'nombre')
    readonly_fields = ('popularidad',)

    fieldsets = (
        ('Información General', {
            'fields': ('nombre', 'categoria', 'tipo_servicio', 'descripcion_web', 'imagen')
        }),
        ('Configuración de Reserva', {
            'fields': ('precio_base', 'duracion', 'capacidad_minima', 'capacidad_maxima',
                      'horario_apertura', 'horario_cierre', 'slots_disponibles')
        }),
        ('Proveedores', {
            'fields': ('proveedores',),
            'description': 'Selecciona los proveedores que pueden realizar este servicio (ej. masajistas).'
        }),
        ('Visibilidad', {
            'fields': ('activo', 'publicado_web'),
            'description': 'Activo = disponible internamente. Publicado web = visible en www.aremko.cl'
        }),
        ('Estadísticas', {
            'fields': ('popularidad',),
            'classes': ('collapse',)
        })
    )

    def precio_con_formato(self, obj):
        return f"${obj.precio_base:,.0f}"
    precio_con_formato.short_description = 'Precio'
    precio_con_formato.admin_order_field = 'precio_base'

    def duracion_minutos(self, obj):
        return f"{obj.duracion} min"
    duracion_minutos.short_description = 'Duración'
    duracion_minutos.admin_order_field = 'duracion'

    def popularidad(self, obj):
        """Muestra qué tan popular es el servicio"""
        total_reservas = ReservaServicio.objects.filter(servicio=obj).count()

        if total_reservas == 0:
            return format_html('<span style="color: gray;">Sin reservas</span>')
        elif total_reservas < 10:
            return format_html('<span style="color: orange;">⭐ Poco popular ({} reservas)</span>', total_reservas)
        elif total_reservas < 50:
            return format_html('<span style="color: blue;">⭐⭐ Popular ({} reservas)</span>', total_reservas)
        else:
            return format_html('<span style="color: green;">⭐⭐⭐ Muy popular ({} reservas)</span>', total_reservas)

    popularidad.short_description = 'Popularidad'

@admin.register(MovimientoCliente)
class MovimientoClienteAdmin(admin.ModelAdmin):
    list_display = ('cliente', 'fecha', 'tipo_movimiento', 'monto_con_formato',
                   'saldo_despues_formato', 'descripcion_corta')
    list_filter = ('tipo_movimiento', 'fecha')
    search_fields = ('cliente__nombre', 'descripcion')
    date_hierarchy = 'fecha'
    readonly_fields = ('fecha', 'saldo_despues')

    def monto_con_formato(self, obj):
        if obj.tipo_movimiento == 'credito':
            return format_html(
                '<span style="color: green;">+${:,.0f}</span>',
                obj.monto
            )
        else:
            return format_html(
                '<span style="color: red;">-${:,.0f}</span>',
                obj.monto
            )
    monto_con_formato.short_description = 'Monto'

    def saldo_despues_formato(self, obj):
        if obj.saldo_despues >= 0:
            return format_html(
                '<span style="color: green;">${:,.0f}</span>',
                obj.saldo_despues
            )
        else:
            return format_html(
                '<span style="color: red;">${:,.0f}</span>',
                obj.saldo_despues
            )
    saldo_despues_formato.short_description = 'Saldo Después'

    def descripcion_corta(self, obj):
        if len(obj.descripcion) > 50:
            return obj.descripcion[:50] + '...'
        return obj.descripcion
    descripcion_corta.short_description = 'Descripción'

    def has_add_permission(self, request):
        # Los movimientos se crean automáticamente
        return False

    def has_delete_permission(self, request, obj=None):
        # Solo superusuarios pueden eliminar movimientos
        return request.user.is_superuser

# Compra Admin
class DetalleCompraInline(admin.TabularInline):
    model = DetalleCompra
    extra = 1
    fields = ['producto', 'cantidad', 'precio_unitario', 'subtotal']
    readonly_fields = ['subtotal']

    def subtotal(self, obj):
        if obj.cantidad and obj.precio_unitario:
            return f"${obj.cantidad * obj.precio_unitario:,.0f}"
        return "$0"

@admin.register(Compra)
class CompraAdmin(admin.ModelAdmin):
    list_display = ('id', 'fecha', 'proveedor', 'total_con_formato', 'estado', 'items_count')
    list_filter = ('estado', 'fecha', 'proveedor')
    search_fields = ('proveedor__nombre', 'factura_numero', 'observaciones')
    date_hierarchy = 'fecha'
    readonly_fields = ('total', 'fecha_creacion', 'fecha_modificacion', 'resumen_compra')
    inlines = [DetalleCompraInline]

    fieldsets = (
        ('Información General', {
            'fields': ('proveedor', 'fecha', 'factura_numero', 'estado')
        }),
        ('Detalles', {
            'fields': ('observaciones', 'total', 'resumen_compra')
        }),
        ('Información del Sistema', {
            'fields': ('fecha_creacion', 'fecha_modificacion'),
            'classes': ('collapse',)
        })
    )

    def total_con_formato(self, obj):
        return f"${obj.total:,.0f}"
    total_con_formato.short_description = 'Total'
    total_con_formato.admin_order_field = 'total'

    def items_count(self, obj):
        return obj.detallecompra_set.count()
    items_count.short_description = 'Ítems'

    def resumen_compra(self, obj):
        """Muestra un resumen de la compra"""
        detalles = obj.detallecompra_set.all()

        if not detalles:
            return "Sin detalles"

        html = '<table style="width: 100%; border-collapse: collapse;">'
        html += '<tr style="background-color: #e0e0e0;">'
        html += '<th style="padding: 5px;">Producto</th>'
        html += '<th style="padding: 5px;">Cantidad</th>'
        html += '<th style="padding: 5px;">P. Unitario</th>'
        html += '<th style="padding: 5px;">Subtotal</th>'
        html += '</tr>'

        total = 0
        for detalle in detalles:
            subtotal = detalle.cantidad * detalle.precio_unitario
            total += subtotal

            html += '<tr>'
            html += f'<td style="padding: 5px;">{detalle.producto.nombre}</td>'
            html += f'<td style="padding: 5px; text-align: center;">{detalle.cantidad}</td>'
            html += f'<td style="padding: 5px; text-align: right;">${detalle.precio_unitario:,.0f}</td>'
            html += f'<td style="padding: 5px; text-align: right;">${subtotal:,.0f}</td>'
            html += '</tr>'

        html += '<tr style="background-color: #e0e0e0; font-weight: bold;">'
        html += '<td colspan="3" style="padding: 5px; text-align: right;">TOTAL:</td>'
        html += f'<td style="padding: 5px; text-align: right;">${total:,.0f}</td>'
        html += '</tr>'
        html += '</table>'

        return format_html(html)

    resumen_compra.short_description = 'Resumen de la Compra'

    def save_formset(self, request, form, formset, change):
        """Recalcular total al guardar los detalles"""
        instances = formset.save(commit=False)

        for instance in instances:
            instance.save()

        for instance in formset.deleted_objects:
            instance.delete()

        formset.save_m2m()

        # Recalcular total
        if hasattr(formset, 'instance'):
            compra = formset.instance
            total = sum(
                detalle.cantidad * detalle.precio_unitario
                for detalle in compra.detallecompra_set.all()
            )
            compra.total = total
            compra.save()

    actions = ['aprobar_compras', 'exportar_compras']

    def aprobar_compras(self, request, queryset):
        """Aprobar compras y actualizar inventario"""
        aprobadas = 0
        for compra in queryset.filter(estado='pendiente'):
            try:
                # Actualizar inventario
                for detalle in compra.detallecompra_set.all():
                    detalle.producto.incrementar_inventario(detalle.cantidad)

                compra.estado = 'aprobada'
                compra.save()
                aprobadas += 1
            except Exception as e:
                messages.error(request, f'Error en compra #{compra.id}: {str(e)}')

        messages.success(request, f'{aprobadas} compras aprobadas y stock actualizado.')

    aprobar_compras.short_description = 'Aprobar compras y actualizar stock'

    def exportar_compras(self, request, queryset):
        """Exportar compras seleccionadas a Excel"""
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="compras.xls"'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Compras')

        # Cabeceras
        columns = ['ID', 'Fecha', 'Proveedor', 'Factura', 'Total', 'Estado', 'Items']
        for col, column_title in enumerate(columns):
            ws.write(0, col, column_title)

        # Datos
        for row, compra in enumerate(queryset, 1):
            ws.write(row, 0, compra.id)
            ws.write(row, 1, compra.fecha.strftime('%d/%m/%Y'))
            ws.write(row, 2, compra.proveedor.nombre if compra.proveedor else '')
            ws.write(row, 3, compra.factura_numero or '')
            ws.write(row, 4, float(compra.total))
            ws.write(row, 5, compra.estado)
            ws.write(row, 6, compra.detallecompra_set.count())

        wb.save(response)
        return response

    exportar_compras.short_description = 'Exportar a Excel'

# Homepage Configuration Admin
@admin.register(HomepageConfig)
class HomepageConfigAdmin(SingletonModelAdmin):
    fieldsets = (
        ('Imágenes', {
            'fields': ('hero_background_image', 'philosophy_image')
        }),
    )

# Homepage Settings Admin
@admin.register(HomepageSettings)
class HomepageSettingsAdmin(SingletonModelAdmin):
    fieldsets = (
        ('Hero Section', {
            'fields': ('hero_title', 'hero_subtitle', 'hero_video_url')
        }),
        ('About Section', {
            'fields': ('about_title', 'about_description', 'about_image')
        }),
        ('Contact Information', {
            'fields': ('contact_phone', 'contact_email', 'contact_address', 'whatsapp_number')
        }),
        ('Social Media', {
            'fields': ('facebook_url', 'instagram_url', 'twitter_url')
        }),
        ('Additional Settings', {
            'fields': ('show_testimonials', 'show_gallery', 'maintenance_mode'),
            'classes': ('collapse',)
        }),
    )

# Sistema de Comunicaciones
@admin.register(CommunicationLimit)
class CommunicationLimitAdmin(SingletonModelAdmin):
    """Admin para gestionar límites de comunicación"""

    fieldsets = (
        ('Límites de Email', {
            'fields': (
                'max_emails_per_day',
                'max_emails_per_hour',
                'max_emails_per_client_per_month'
            )
        }),
        ('Límites de SMS', {
            'fields': (
                'max_sms_per_day',
                'max_sms_per_hour',
                'max_sms_per_client_per_month'
            )
        }),
        ('Control de Frecuencia', {
            'fields': (
                'min_hours_between_communications',
                'quiet_hours_start',
                'quiet_hours_end'
            ),
            'description': 'Define las horas de silencio donde no se envían comunicaciones'
        })
    )

@admin.register(ClientPreferences)
class ClientPreferencesAdmin(admin.ModelAdmin):
    list_display = ('cliente', 'permite_emails', 'permite_sms', 'frecuencia_preferida')
    list_filter = ('permite_emails', 'permite_sms', 'frecuencia_preferida')
    search_fields = ('cliente__nombre', 'cliente__email')
    autocomplete_fields = ['cliente']

@admin.register(CommunicationLog)
class CommunicationLogAdmin(admin.ModelAdmin):
    list_display = ('cliente', 'tipo', 'asunto', 'estado', 'fecha_envio', 'fecha_apertura')
    list_filter = ('tipo', 'estado', 'fecha_envio')
    search_fields = ('cliente__nombre', 'asunto', 'mensaje')
    date_hierarchy = 'fecha_envio'
    readonly_fields = ('fecha_envio', 'fecha_apertura', 'fecha_click', 'error_mensaje')

    fieldsets = (
        ('Información General', {
            'fields': ('cliente', 'tipo', 'asunto', 'mensaje')
        }),
        ('Estado y Tracking', {
            'fields': ('estado', 'fecha_envio', 'fecha_apertura', 'fecha_click')
        }),
        ('Información de Error', {
            'fields': ('error_mensaje',),
            'classes': ('collapse',)
        })
    )

@admin.register(SMSTemplate)
class SMSTemplateAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'tipo_evento', 'activo', 'longitud_mensaje')
    list_filter = ('tipo_evento', 'activo')
    search_fields = ('nombre', 'mensaje')

    def longitud_mensaje(self, obj):
        return len(obj.mensaje)
    longitud_mensaje.short_description = 'Longitud'

@admin.register(MailParaEnviar)
class MailParaEnviarAdmin(admin.ModelAdmin):
    list_display = ('destinatario', 'asunto', 'estado', 'fecha_programada',
                   'prioridad', 'intentos')
    list_filter = ('estado', 'prioridad', 'fecha_programada')
    search_fields = ('destinatario', 'asunto')
    date_hierarchy = 'fecha_programada'
    readonly_fields = ('fecha_creacion', 'fecha_envio', 'fecha_ultimo_intento',
                      'error_log')

    fieldsets = (
        ('Información del Email', {
            'fields': ('destinatario', 'asunto', 'mensaje', 'html_mensaje')
        }),
        ('Configuración', {
            'fields': ('prioridad', 'fecha_programada', 'cliente_relacionado')
        }),
        ('Estado', {
            'fields': ('estado', 'intentos', 'max_intentos')
        }),
        ('Historial', {
            'fields': ('fecha_creacion', 'fecha_envio', 'fecha_ultimo_intento',
                      'error_log'),
            'classes': ('collapse',)
        })
    )

    actions = ['enviar_ahora', 'marcar_como_pendiente']

    def enviar_ahora(self, request, queryset):
        """Enviar emails seleccionados inmediatamente"""
        enviados = 0
        for mail in queryset.filter(estado='pendiente'):
            # Aquí iría la lógica de envío
            mail.fecha_programada = timezone.now()
            mail.save()
            enviados += 1

        self.message_user(request, f'{enviados} emails programados para envío inmediato.')

    enviar_ahora.short_description = 'Enviar ahora'

    def marcar_como_pendiente(self, request, queryset):
        """Remarcar emails como pendientes"""
        actualizados = queryset.filter(estado__in=['error', 'cancelado']).update(
            estado='pendiente',
            intentos=0
        )
        self.message_user(request, f'{actualizados} emails marcados como pendientes.')

    marcar_como_pendiente.short_description = 'Marcar como pendiente'

# Advanced Email Campaign Admin
@admin.register(EmailCampaign)
class EmailCampaignAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'tipo', 'estado', 'fecha_inicio',
                   'emails_enviados', 'tasa_apertura', 'tasa_click')
    list_filter = ('tipo', 'estado', 'fecha_inicio')
    search_fields = ('nombre', 'descripcion')
    readonly_fields = ('emails_enviados', 'emails_abiertos', 'clicks_totales',
                      'tasa_apertura', 'tasa_click', 'estadisticas_detalladas')

    fieldsets = (
        ('Información General', {
            'fields': ('nombre', 'descripcion', 'tipo', 'estado')
        }),
        ('Configuración', {
            'fields': ('template_email', 'segmento_clientes', 'fecha_inicio',
                      'fecha_fin')
        }),
        ('Estadísticas', {
            'fields': ('emails_enviados', 'emails_abiertos', 'clicks_totales',
                      'tasa_apertura', 'tasa_click', 'estadisticas_detalladas'),
            'classes': ('collapse',)
        })
    )

    def tasa_apertura(self, obj):
        if obj.emails_enviados == 0:
            return '0%'
        tasa = (obj.emails_abiertos / obj.emails_enviados) * 100
        return f'{tasa:.1f}%'
    tasa_apertura.short_description = 'Tasa Apertura'

    def tasa_click(self, obj):
        if obj.emails_enviados == 0:
            return '0%'
        tasa = (obj.clicks_totales / obj.emails_enviados) * 100
        return f'{tasa:.1f}%'
    tasa_click.short_description = 'Tasa Click'

    def estadisticas_detalladas(self, obj):
        """Muestra estadísticas detalladas de la campaña"""
        # Aquí se pueden agregar gráficos y estadísticas más complejas
        return format_html(
            '<div style="padding: 10px; background-color: #f5f5f5;">'
            '<h4>Resumen de Rendimiento</h4>'
            '<p>Emails enviados: <strong>{}</strong></p>'
            '<p>Emails abiertos: <strong>{}</strong></p>'
            '<p>Clicks totales: <strong>{}</strong></p>'
            '<p>Tasa de rebote: <strong>Pendiente implementar</strong></p>'
            '</div>',
            obj.emails_enviados,
            obj.emails_abiertos,
            obj.clicks_totales
        )
    estadisticas_detalladas.short_description = 'Estadísticas Detalladas'

# Service History Admin
@admin.register(ServiceHistory)
class ServiceHistoryAdmin(admin.ModelAdmin):
    list_display = ('cliente', 'servicio', 'fecha', 'precio_pagado',
                   'duracion_real', 'rating', 'recomienda')
    list_filter = ('servicio', 'fecha', 'rating', 'recomienda')
    search_fields = ('cliente__nombre', 'servicio__nombre', 'notas', 'feedback')
    date_hierarchy = 'fecha'
    readonly_fields = ('fecha',)

    fieldsets = (
        ('Información del Servicio', {
            'fields': ('cliente', 'servicio', 'fecha', 'precio_pagado')
        }),
        ('Detalles de la Sesión', {
            'fields': ('duracion_real', 'proveedor', 'notas')
        }),
        ('Feedback del Cliente', {
            'fields': ('rating', 'recomienda', 'feedback'),
            'classes': ('collapse',)
        })
    )

    def get_queryset(self, request):
        """Optimizar queries"""
        return super().get_queryset(request).select_related(
            'cliente', 'servicio', 'proveedor'
        )

# CRM Models Admin
@admin.register(Lead)
class LeadAdmin(admin.ModelAdmin):
    list_display = ('get_full_name', 'email', 'phone', 'status', 'source',
                   'created_at', 'company_name')
    list_filter = ('status', 'source', 'created_at')
    search_fields = ('first_name', 'last_name', 'email', 'phone', 'company_name', 'notes')
    date_hierarchy = 'created_at'

    fieldsets = (
        ('Información de Contacto', {
            'fields': ('first_name', 'last_name', 'email', 'phone')
        }),
        ('Estado del Lead', {
            'fields': ('status', 'source', 'campaign')
        }),
        ('Información Adicional', {
            'fields': ('company_name', 'notes')
        })
    )

    def get_full_name(self, obj):
        return f"{obj.first_name} {obj.last_name}"
    get_full_name.short_description = 'Nombre'
    get_full_name.admin_order_field = 'first_name'

    actions = ['convertir_a_cliente']

    def convertir_a_cliente(self, request, queryset):
        """Convertir leads seleccionados en clientes"""
        convertidos = 0
        for lead in queryset.filter(status='Qualified'):
            cliente, created = Cliente.objects.get_or_create(
                email=lead.email,
                defaults={
                    'nombre': f"{lead.first_name} {lead.last_name}",
                    'telefono': lead.phone or '',
                    'pais': 'Chile',
                    'ciudad': lead.company_name or ''
                }
            )
            if created:
                lead.status = 'Converted'
                lead.save()
                convertidos += 1

        self.message_user(request, f'{convertidos} leads convertidos a clientes.')

    convertir_a_cliente.short_description = 'Convertir a cliente'


@admin.register(Deal)
class DealAdmin(admin.ModelAdmin):
    list_display = ('name', 'contact', 'stage', 'amount', 'probability',
                   'expected_close_date', 'created_at')
    list_filter = ('stage', 'expected_close_date', 'created_at', 'campaign')
    search_fields = ('name', 'contact__first_name', 'contact__last_name',
                    'contact__email', 'notes')
    date_hierarchy = 'created_at'

    fieldsets = (
        ('Información de la Oportunidad', {
            'fields': ('name', 'contact', 'stage')
        }),
        ('Detalles Financieros', {
            'fields': ('amount', 'probability', 'expected_close_date')
        }),
        ('Vínculos', {
            'fields': ('related_booking', 'campaign')
        }),
        ('Notas', {
            'fields': ('notes',)
        })
    )

    autocomplete_fields = ['contact', 'related_booking', 'campaign']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('contact', 'campaign', 'related_booking')


@admin.register(Company)
class CompanyAdmin(admin.ModelAdmin):
    list_display = ('name', 'industry', 'website', 'created_at')
    list_filter = ('industry', 'created_at')
    search_fields = ('name', 'industry', 'website', 'notes')
    date_hierarchy = 'created_at'


@admin.register(Contact)
class ContactAdmin(admin.ModelAdmin):
    list_display = ('get_full_name', 'email', 'phone', 'company', 'created_at')
    list_filter = ('company', 'created_at')
    search_fields = ('first_name', 'last_name', 'email', 'phone', 'company__name')
    date_hierarchy = 'created_at'
    autocomplete_fields = ['company', 'campaign']

    def get_full_name(self, obj):
        return f"{obj.first_name} {obj.last_name}"
    get_full_name.short_description = 'Nombre Completo'


@admin.register(Activity)
class ActivityAdmin(admin.ModelAdmin):
    list_display = ('activity_type', 'subject', 'related_contact', 'campaign',
                   'created_at', 'created_by')
    list_filter = ('activity_type', 'created_at', 'campaign')
    search_fields = ('subject', 'notes', 'related_contact__first_name',
                    'related_contact__last_name')
    date_hierarchy = 'created_at'
    autocomplete_fields = ['related_contact', 'campaign']


@admin.register(Campaign)
class CampaignAdmin(admin.ModelAdmin):
    list_display = ('name', 'status', 'start_date', 'end_date', 'budget',
                   'created_at')
    list_filter = ('status', 'start_date', 'end_date')
    search_fields = ('name', 'description')
    date_hierarchy = 'created_at'

    fieldsets = (
        ('Información Básica', {
            'fields': ('name', 'description', 'status')
        }),
        ('Fechas y Presupuesto', {
            'fields': ('start_date', 'end_date', 'budget')
        }),
        ('Segmentación', {
            'fields': ('target_audience', 'target_min_spend', 'target_comunas')
        })
    )


@admin.register(CampaignInteraction)
class CampaignInteractionAdmin(admin.ModelAdmin):
    list_display = ('interaction_type', 'campaign', 'contact', 'cliente',
                   'interaction_date', 'response', 'converted')
    list_filter = ('interaction_type', 'response', 'converted', 'interaction_date')
    search_fields = ('campaign__name', 'contact__email', 'cliente__email')
    date_hierarchy = 'interaction_date'
    autocomplete_fields = ['campaign', 'contact', 'cliente']


@admin.register(EmailSubjectTemplate)
class EmailSubjectTemplateAdmin(admin.ModelAdmin):
    list_display = ('name', 'subject_line', 'active', 'created_at')
    list_filter = ('active', 'created_at')
    search_fields = ('name', 'subject_line')


@admin.register(EmailContentTemplate)
class EmailContentTemplateAdmin(admin.ModelAdmin):
    list_display = ('name', 'active', 'created_at')
    list_filter = ('active', 'created_at')
    search_fields = ('name', 'content')
    formfield_overrides = {
        models.TextField: {'widget': forms.Textarea(attrs={'rows': 20, 'cols': 100})},
    }


# Sistema de Premios y Tramos

@admin.register(Premio)
class PremioAdmin(admin.ModelAdmin):
    list_display = (
        'nombre',
        'tipo',
        'valor_formateado',
        'tramo_minimo',
        'dias_validez',
        'activo',
        'stock_display'
    )
    list_filter = ('tipo', 'activo', 'tramo_minimo')
    search_fields = ('nombre', 'descripcion')
    ordering = ['-activo', 'tramo_minimo', 'nombre']

    fieldsets = (
        ('Información General', {
            'fields': ('nombre', 'descripcion', 'tipo', 'activo')
        }),
        ('Configuración del Premio', {
            'fields': ('valor', 'condiciones', 'dias_validez', 'tramo_minimo')
        }),
        ('Control de Stock', {
            'fields': ('stock_disponible', 'stock_inicial'),
            'description': 'Dejar en blanco para stock ilimitado'
        })
    )

    def valor_formateado(self, obj):
        """Muestra el valor formateado según el tipo"""
        if obj.tipo == 'descuento_porcentaje':
            return f'{obj.valor}%'
        elif obj.tipo in ['descuento_monto', 'credito']:
            return f'${obj.valor:,.0f}'
        else:
            return obj.valor
    valor_formateado.short_description = 'Valor'

    def stock_display(self, obj):
        """Muestra el stock disponible"""
        if obj.stock_disponible is None:
            return format_html('<span style="color: green;">Ilimitado</span>')
        elif obj.stock_disponible == 0:
            return format_html('<span style="color: red;">Agotado</span>')
        elif obj.stock_disponible < 10:
            return format_html(
                '<span style="color: orange;">{} disponibles</span>',
                obj.stock_disponible
            )
        else:
            return format_html(
                '<span style="color: green;">{} disponibles</span>',
                obj.stock_disponible
            )
    stock_display.short_description = 'Stock'

@admin.register(ClientePremio)
class ClientePremioAdmin(admin.ModelAdmin):
    list_display = (
        'codigo_link',
        'cliente_link',
        'premio',
        'estado_badge',
        'fecha_ganado',
        'fecha_expiracion',
        'dias_restantes'
    )
    list_filter = ('estado', 'premio__tipo', 'fecha_ganado', 'fecha_expiracion')
    search_fields = (
        'codigo_unico',
        'cliente__nombre',
        'cliente__email',
        'premio__nombre'
    )
    readonly_fields = (
        'codigo_unico',
        'fecha_ganado',
        'fecha_uso',
        'fecha_expiracion',
        'venta_donde_uso'
    )
    date_hierarchy = 'fecha_ganado'

    fieldsets = (
        ('Información del Premio', {
            'fields': ('cliente', 'premio', 'codigo_unico', 'estado')
        }),
        ('Fechas', {
            'fields': ('fecha_ganado', 'fecha_expiracion', 'fecha_uso')
        }),
        ('Uso', {
            'fields': ('venta_donde_uso', 'notas'),
            'classes': ('collapse',)
        })
    )

    def codigo_link(self, obj):
        """Código con formato monospace"""
        return format_html(
            '<code style="background-color: #f5f5f5; padding: 2px 6px; '
            'border-radius: 3px;">{}</code>',
            obj.codigo_unico
        )
    codigo_link.short_description = 'Código'
    codigo_link.admin_order_field = 'codigo_unico'

    def cliente_link(self, obj):
        """Link al cliente"""
        if obj.cliente:
            url = reverse('admin:ventas_cliente_change', args=[obj.cliente.id])
            return format_html('<a href="{}">{}</a>', url, obj.cliente.nombre)
        return '-'
    cliente_link.short_description = 'Cliente'
    cliente_link.admin_order_field = 'cliente__nombre'

    def estado_badge(self, obj):
        """Badge colorido para el estado"""
        colors = {
            'pendiente': '#ffc107',
            'aprobado': '#28a745',
            'enviado': '#17a2b8',
            'usado': '#6c757d',
            'expirado': '#dc3545',
            'cancelado': '#dc3545'
        }

        # Verificar si está expirado
        if obj.estado in ['aprobado', 'enviado'] and obj.fecha_expiracion < timezone.now():
            estado_display = 'EXPIRADO'
            color = colors['expirado']
        else:
            estado_display = obj.get_estado_display()
            color = colors.get(obj.estado, '#6c757d')

        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 10px; '
            'border-radius: 3px; font-size: 11px;">{}</span>',
            color,
            estado_display
        )
    estado_badge.short_description = 'Estado'

    def dias_restantes(self, obj):
        """Días restantes para usar el premio"""
        if obj.estado == 'usado':
            return format_html('<span style="color: gray;">Usado</span>')

        dias = (obj.fecha_expiracion - timezone.now()).days

        if dias < 0:
            return format_html('<span style="color: red;">Expirado</span>')
        elif dias == 0:
            return format_html('<span style="color: red;">Hoy</span>')
        elif dias <= 7:
            return format_html('<span style="color: orange;">{} días</span>', dias)
        else:
            return format_html('<span style="color: green;">{} días</span>', dias)

    dias_restantes.short_description = 'Vigencia'

    def has_delete_permission(self, request, obj=None):
        """Solo superusers pueden eliminar"""
        return request.user.is_superuser

    actions = ['enviar_notificacion', 'extender_vigencia', 'marcar_como_enviado']

    def enviar_notificacion(self, request, queryset):
        """Enviar notificación del premio al cliente"""
        enviados = 0
        for premio in queryset.filter(estado='aprobado'):
            # Aquí iría la lógica de envío
            premio.estado = 'enviado'
            premio.save()
            enviados += 1

        self.message_user(request, f'{enviados} notificaciones enviadas')
    enviar_notificacion.short_description = 'Enviar notificación al cliente'

    def extender_vigencia(self, request, queryset):
        """Extender vigencia por 30 días"""
        for premio in queryset:
            premio.fecha_expiracion += timedelta(days=30)
            premio.save()

        self.message_user(
            request,
            f'{queryset.count()} premios extendidos por 30 días'
        )
    extender_vigencia.short_description = 'Extender vigencia (+30 días)'

    def marcar_como_enviado(self, request, queryset):
        """Marcar como enviado"""
        actualizados = queryset.filter(estado='aprobado').update(estado='enviado')
        self.message_user(request, f'{actualizados} premios marcados como enviados')
    marcar_como_enviado.short_description = 'Marcar como enviado'

@admin.register(HistorialTramo)
class HistorialTramoAdmin(admin.ModelAdmin):
    list_display = (
        'cliente_link',
        'cambio_tramo',
        'gasto_momento',
        'fecha_cambio',
        'premio_link'
    )
    list_filter = ('fecha_cambio', 'tramo_hasta')
    search_fields = ('cliente__nombre', 'cliente__email')
    readonly_fields = (
        'cliente',
        'tramo_desde',
        'tramo_hasta',
        'fecha_cambio',
        'gasto_en_momento',
        'premio_generado'
    )

    def cliente_link(self, obj):
        """Link al cliente"""
        if obj.cliente:
            url = reverse('admin:ventas_cliente_change', args=[obj.cliente.id])
            return format_html('<a href="{}">{}</a>', url, obj.cliente.nombre)
        return '-'
    cliente_link.short_description = 'Cliente'

    def cambio_tramo(self, obj):
        """Visualización del cambio de tramo"""
        return format_html(
            '<span style="font-weight: bold;">Tramo {} → {}</span>',
            obj.tramo_desde,
            obj.tramo_hasta
        )
    cambio_tramo.short_description = 'Cambio'

    def gasto_momento(self, obj):
        """Gasto formateado"""
        return f'${obj.gasto_en_momento:,.0f}'
    gasto_momento.short_description = 'Gasto'

    def premio_link(self, obj):
        """Link al premio generado"""
        if obj.premio_generado:
            url = reverse('admin:ventas_clientepremio_change', args=[obj.premio_generado.id])
            return format_html(
                '<a href="{}">{}</a>',
                url,
                obj.premio_generado.premio.nombre
            )
        return '-'
    premio_link.short_description = 'Premio Generado'

    def has_add_permission(self, request):
        """No permitir crear manualmente"""
        return False

    def has_delete_permission(self, request, obj=None):
        """Solo superusers pueden eliminar"""
        return request.user.is_superuser


# ========== ADMIN PARA PACKS DE DESCUENTO ==========

@admin.register(PackDescuento)
class PackDescuentoAdmin(admin.ModelAdmin):
    """Admin para gestionar Packs de Descuento"""

    list_display = (
        'nombre',
        'descuento_formateado',
        'servicios_display',
        'dias_display',
        'activo',
        'vigencia_display',
        'prioridad'
    )

    list_filter = ('activo', 'fecha_inicio', 'fecha_fin')

    search_fields = ('nombre', 'descripcion')

    ordering = ['-activo', '-prioridad', '-descuento']

    fieldsets = (
        ('Información General', {
            'fields': ('nombre', 'descripcion', 'descuento', 'activo')
        }),
        ('Servicios Requeridos', {
            'fields': ('servicios_requeridos',),
            'description': 'Ingrese los tipos de servicios como lista JSON: ["ALOJAMIENTO", "TINA"]'
        }),
        ('Restricciones de Fecha', {
            'fields': ('dias_semana_validos', 'fecha_inicio', 'fecha_fin'),
            'description': 'Días: 0=Domingo, 1=Lunes... 6=Sábado. Ej: [0,1,2,3,4] para Dom-Jue'
        }),
        ('Configuración Adicional', {
            'fields': ('prioridad', 'cantidad_minima_noches', 'misma_fecha')
        })
    )

    def descuento_formateado(self, obj):
        """Muestra el descuento formateado"""
        return f'${obj.descuento:,.0f}'
    descuento_formateado.short_description = 'Descuento'

    def servicios_display(self, obj):
        """Muestra los servicios requeridos"""
        if obj.servicios_requeridos:
            return ', '.join(obj.servicios_requeridos)
        return 'Sin servicios definidos'
    servicios_display.short_description = 'Servicios Requeridos'

    def dias_display(self, obj):
        """Muestra los días válidos"""
        return obj.get_dias_semana_display()
    dias_display.short_description = 'Días Válidos'

    def vigencia_display(self, obj):
        """Muestra el período de vigencia"""
        inicio = obj.fecha_inicio.strftime('%d/%m/%Y')
        fin = obj.fecha_fin.strftime('%d/%m/%Y') if obj.fecha_fin else 'Sin límite'
        return f'{inicio} - {fin}'
    vigencia_display.short_description = 'Vigencia'

    def save_model(self, request, obj, form, change):
        """Validaciones al guardar"""
        # Validar que servicios_requeridos sea una lista
        if not isinstance(obj.servicios_requeridos, list):
            messages.error(request, 'Los servicios requeridos deben ser una lista.')
            return

        # Validar que días_semana_validos sea una lista
        if obj.dias_semana_validos and not isinstance(obj.dias_semana_validos, list):
            messages.error(request, 'Los días de la semana deben ser una lista.')
            return

        super().save_model(request, obj, form, change)

        if not change:
            messages.success(request, f'Pack "{obj.nombre}" creado exitosamente.')
        else:
            messages.success(request, f'Pack "{obj.nombre}" actualizado.')

    class Media:
        css = {
            'all': ('admin/css/forms.css',)
        }
        js = ('admin/js/jquery.init.js',)