# -*- coding: utf-8 -*-
"""Servicios de finanzas: consumidores automáticos de datos de otras fuentes.

P-22 F2-B: el fetch de la conciliación (`conciliacion.services_mp.traer_pagos_mp`)
separa los pagos donde Aremko COBRA de los donde Aremko PAGA. Los segundos son
ruido para la cola de Deborah, pero para finanzas son exactamente los gastos
pagados vía Mercado Pago — un solo fetch, dos consumidores.

P-22 F3: las transferencias SALIENTES de MP (pagos a trabajadores, barridos a
Scotiabank) no aparecen en /v1/payments/search — pero MP manda un correo por
cada una («Tu transferencia fue enviada», a abonosaremko, reenviado a ecolonco
por el filtro del 2026-08-06). `parsear_transferencia_mp` +
`registrar_transferencia_mp` convierten ese correo en el movimiento.
"""
import hashlib
import html as html_lib
import logging
import re
from datetime import date

from django.utils.dateparse import parse_datetime

logger = logging.getLogger(__name__)

# Decisión de Jorge 2026-08-08: la cobertura de gastos parte en julio 2026, el
# primer mes completo y que se recuerda. Aplica igual a lo que llega por API.
COBERTURA_GASTOS_DESDE = date(2026, 7, 1)

# payment_type_id de MP → clave de CuentaFinanciera: CON QUÉ PLATA se pagó.
# Si la compra se fondeó con la Visa, el gasto es de la Visa (el saldo de MP no
# se movió) — importa para que el cierre por saldos (F4) cuadre por cuenta.
CUENTA_POR_PAYMENT_TYPE = {
    'account_money': 'mercado_pago',
    'credit_card': 'visa_2936',
}


def registrar_compras_mp(pagos_api):
    """Convierte pagos de /v1/payments/search donde Aremko es el PAGADOR en
    MovimientoFinanciero de gasto (categoría «Por clasificar», fuente api).

    Idempotente por referencia ``mp:<payment_id>``. Devuelve cuántos gastos
    creó. NUNCA lanza: un problema acá no puede romper el fetch de Deborah —
    el que llama ya envuelve en try/except, y acá además se aísla por pago.
    """
    from .models import CategoriaFinanciera, CuentaFinanciera, MovimientoFinanciero

    try:
        cuentas = {c.clave: c for c in CuentaFinanciera.objects.all()}
        cat = CategoriaFinanciera.objects.get(clave='por_clasificar')
    except CategoriaFinanciera.DoesNotExist:
        logger.warning('finanzas sin sembrar: compras MP no registradas')
        return 0

    creados = 0
    for p in pagos_api:
        try:
            pid = str(p.get('id') or '')
            monto = int(p.get('transaction_amount') or 0)
            if not pid or monto <= 0:
                continue
            f = parse_datetime(p.get('date_approved') or p.get('date_created') or '')
            if not f or f.date() < COBERTURA_GASTOS_DESDE:
                continue
            ref = f'mp:{pid}'
            if MovimientoFinanciero.objects.filter(referencia=ref).exists():
                continue
            clave = CUENTA_POR_PAYMENT_TYPE.get(p.get('payment_type_id'), 'mercado_pago')
            cuenta = cuentas.get(clave)
            if cuenta is None:
                logger.warning('compra MP %s: cuenta %s no sembrada', pid, clave)
                continue
            MovimientoFinanciero.objects.create(
                fecha=f.date(), cuenta=cuenta, clase='gasto', sentido='sale',
                monto=monto, categoria=cat, fuente='api', referencia=ref,
                descripcion=f"Compra vía MP: {p.get('description') or 'sin glosa'}"[:255],
            )
            creados += 1
        except Exception:
            logger.exception('compra MP %s no registrada', p.get('id'))

    if creados:
        logger.info('finanzas: %s compras MP registradas como gasto por clasificar', creados)
    return creados


# ── F3: transferencias salientes de Mercado Pago, desde el correo ────────────

def _texto_plano(html_crudo):
    """HTML del correo → texto plano de una línea (tags fuera, entidades resueltas)."""
    txt = re.sub(r'<[^>]+>', ' ', html_crudo or '')
    txt = html_lib.unescape(txt)
    return re.sub(r'\s+', ' ', txt).strip()


def parsear_transferencia_mp(html_crudo):
    """Extrae monto/beneficiario/entidad del correo «Tu transferencia fue enviada».

    Función pura (testeable sin red). Formato real verificado 2026-08-08:
    «Ya enviamos tu transferencia de $ 10.000 … Datos del beneficiario
    Nombre y apellido: Martin Aguilera Entidad: Banco Estado Número de cuenta: …»
    Devuelve None si el correo no calza — mejor no registrar que adivinar.
    """
    txt = _texto_plano(html_crudo)
    m = re.search(r'transferencia de\s*\$\s*([\d.]+)', txt)
    b = re.search(r'Nombre y apellido:\s*(.*?)\s*Entidad:', txt)
    if not m or not b:
        return None
    monto = int(m.group(1).replace('.', ''))
    if monto <= 0 or not b.group(1).strip():
        return None
    e = re.search(r'Entidad:\s*(.*?)\s*N[úu]mero de cuenta:', txt)
    return {'monto': monto, 'beneficiario': b.group(1).strip(),
            'entidad': e.group(1).strip() if e else ''}


def referencia_correo(message_id, respaldo=''):
    """Referencia idempotente para un correo: hash del Message-ID (estable)."""
    base = (message_id or respaldo or '').strip()
    return 'correo:mp:' + hashlib.sha1(base.encode()).hexdigest()[:24]


def registrar_transferencia_mp(datos, fecha, referencia):
    """Crea el/los MovimientoFinanciero de una transferencia saliente de MP.

    Clasificación (las mismas reglas validadas con Jorge para el histórico):
    - beneficiario con «aremko» → TRASPASO MP→Scotiabank, dos piernas enlazadas
    - «insumos sur» → gasto insumos · resto → gasto remuneraciones

    Guardias: idempotencia por referencia, y anti-solape con la carga histórica
    (si existe un hist:mp con la misma fecha+monto, este correo ya está cargado).
    Devuelve ('creado'|'ya_existe'|'en_historico', movimientos_creados).
    """
    from django.db import transaction

    from .models import CategoriaFinanciera, CuentaFinanciera, MovimientoFinanciero

    if MovimientoFinanciero.objects.filter(
            referencia__in=(referencia, f'{referencia}:sale')).exists():
        return 'ya_existe', 0
    if MovimientoFinanciero.objects.filter(
            referencia__startswith='hist:mp', fecha=fecha,
            monto=datos['monto']).exists():
        return 'en_historico', 0

    cuentas = {c.clave: c for c in CuentaFinanciera.objects.all()}
    benef = datos['beneficiario']
    with transaction.atomic():
        if 'aremko' in benef.lower():
            sale = MovimientoFinanciero.objects.create(
                fecha=fecha, cuenta=cuentas['mercado_pago'], clase='traspaso',
                sentido='sale', monto=datos['monto'], fuente='correo',
                referencia=f'{referencia}:sale',
                descripcion='Barrido MP → Scotiabank (correo)')
            entra = MovimientoFinanciero.objects.create(
                fecha=fecha, cuenta=cuentas['scotiabank'], clase='traspaso',
                sentido='entra', monto=datos['monto'], fuente='correo',
                referencia=f'{referencia}:entra', traspaso_par=sale,
                descripcion='Barrido MP → Scotiabank (correo)')
            sale.traspaso_par = entra
            sale.save(update_fields=['traspaso_par'])
            return 'creado', 2

        # Plan de cuentas 2026-08-08: el destinatario decide la categoría por
        # reglas (masajistas, sueldos, personales…); un nombre desconocido va
        # a «por clasificar» — mejor a mano que adivinado.
        from .reglas import PLAN_CUENTAS, clasificar_por_reglas
        clave_cat = clasificar_por_reglas(benef) or 'por_clasificar'
        defn = PLAN_CUENTAS.get(clave_cat, ('Por clasificar', 'gasto', 'otros'))
        cat, _ = CategoriaFinanciera.objects.get_or_create(
            clave=clave_cat, defaults={'nombre': defn[0], 'clase': defn[1],
                                       'grupo': defn[2]})
        MovimientoFinanciero.objects.create(
            fecha=fecha, cuenta=cuentas['mercado_pago'], clase='gasto',
            sentido='sale', monto=datos['monto'],
            categoria=cat,
            fuente='correo', referencia=referencia,
            descripcion=f'Transferencia MP a {benef}'[:255])
        return 'creado', 1


# ── F4 paso 2: comisiones que MP descuenta por cada cobro ────────────────────

def _comision_cobro_mp(pago):
    """Comisión total que paga Aremko (collector) en un pago, en CLP enteros.

    `fee_details` viene por pago en /v1/payments/search; las transferencias
    simples suelen traer 0 y los cobros con tarjeta/link el % de MP.
    """
    total = 0.0
    for f in (pago.get('fee_details') or []):
        if (f.get('fee_payer') or 'collector') == 'collector':
            total += float(f.get('amount') or 0)
    return int(round(total))


def registrar_comisiones_mp(cobros_api):
    """Registra la comisión de cada cobro MP como gasto «comisiones» (fuente api).

    Sin esto el saldo MP calculado queda inflado: los cobros entran BRUTOS al
    registro comercial, pero a la cuenta le llega el neto. Idempotente por
    referencia mp:fee:<payment_id>; respeta el corte de julio. Nunca lanza.
    """
    from .models import CategoriaFinanciera, CuentaFinanciera, MovimientoFinanciero

    try:
        cuenta = CuentaFinanciera.objects.get(clave='mercado_pago')
        cat = CategoriaFinanciera.objects.get(clave='comisiones')
    except (CuentaFinanciera.DoesNotExist, CategoriaFinanciera.DoesNotExist):
        logger.warning('finanzas sin sembrar: comisiones MP no registradas')
        return 0

    creados = 0
    for p in cobros_api:
        try:
            pid = str(p.get('id') or '')
            monto = _comision_cobro_mp(p)
            if not pid or monto <= 0:
                continue
            f = parse_datetime(p.get('date_approved') or p.get('date_created') or '')
            if not f or f.date() < COBERTURA_GASTOS_DESDE:
                continue
            ref = f'mp:fee:{pid}'
            if MovimientoFinanciero.objects.filter(referencia=ref).exists():
                continue
            MovimientoFinanciero.objects.create(
                fecha=f.date(), cuenta=cuenta, clase='gasto', sentido='sale',
                monto=monto, categoria=cat, fuente='api', referencia=ref,
                descripcion=f"Comisión MP del cobro {pid} ({p.get('description') or 'sin glosa'})"[:255])
            creados += 1
        except Exception:
            logger.exception('comisión MP %s no registrada', p.get('id'))

    if creados:
        logger.info('finanzas: %s comisiones MP registradas', creados)
    return creados


# ── F4 paso 3: cartola BancoEstado (export XLSX del portal) ──────────────────
# Formato real verificado 2026-08-08 (Excel_Cartola_Historica_Chequera_
# Electronica.xlsx): hoja "Resumen" con rótulos en col A y valores en col E;
# hoja "Movimientos" con encabezado Fecha DD/MM (sin año — el año sale del
# rango del Resumen) | ... | Descripción | Cheques / Cargos | Depósitos /
# Abonos | Saldo (encadenado fila a fila). Montos mezclados: enteros crudos
# y strings '$1.234.567'.

CATEGORIAS_CARTOLA = {
    # clave → (nombre, clase). Se crean al confirmar si no existen (sin Shell).
    'liquidacion_flow': ('Liquidación Flow', 'ingreso'),
    'liquidacion_sumup': ('Liquidación SumUp', 'ingreso'),
    'transferencias_recibidas': ('Transferencias recibidas', 'ingreso'),
}


def _monto_celda(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(round(v))
    s = str(v).replace('$', '').replace('.', '').replace(' ', '').strip()
    return int(s) if s.lstrip('-').isdigit() else 0


def clasificar_fila_cartola(descripcion, cargo, abono):
    """(clase, sentido, categoria_clave) para una fila de la cartola."""
    d = (descripcion or '').upper()
    if abono > 0:
        if 'SUMUP' in d:
            return 'ingreso', 'entra', 'liquidacion_sumup'
        if 'FLOW' in d:
            return 'ingreso', 'entra', 'liquidacion_flow'
        return 'ingreso', 'entra', 'transferencias_recibidas'
    # Los cargos pasan por el MISMO plan de cuentas que todo lo demás. Sin
    # esto, un «TEF A TOLOZA POBLETE ALDA» entraba como por clasificar y ni
    # se veía como retiro ni aparecía de candidato en Calzar retiros, aunque
    # la regla existiera desde el principio (visto 2026-08-10).
    from .reglas import clasificar_por_reglas
    return 'gasto', 'sale', clasificar_por_reglas(descripcion) or 'por_clasificar'


def parsear_cartola_bancoestado(archivo):
    """Lee el XLSX del portal y devuelve dict con resumen, filas y chequeos.

    No escribe nada. Cada fila queda con su referencia idempotente
    be:<hash(fecha|descripcion|cargo|abono|saldo)> — el saldo encadenado hace
    única incluso a la segunda transferencia idéntica del mismo día.
    """
    import warnings
    from datetime import datetime

    import openpyxl

    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)

    if 'Resumen' in wb.sheetnames and 'Registros' in wb.sheetnames:
        # Variante «Cartola en Línea» (verificada 2026-08-10): otra hoja, sin
        # fechas de período y con las filas DESORDENADAS.
        return _parsear_bancoestado_en_linea(wb)

    if 'Resumen' not in wb.sheetnames or 'Movimientos' not in wb.sheetnames:
        raise ValueError(
            'El archivo no tiene las hojas del export de BancoEstado: se '
            'esperan Resumen + Movimientos (Cartola Histórica) o Resumen + '
            'Registros (Cartola en Línea).')

    # ── Resumen: rótulo en col A, valor en la primera celda no vacía después ──
    resumen = {}
    for fila in wb['Resumen'].iter_rows(values_only=True):
        rotulo = str(fila[0] or '').strip()
        if not rotulo:
            continue
        valor = next((c for c in fila[1:] if c is not None), None)
        resumen[rotulo] = valor

    def _fecha_resumen(rotulo):
        crudo = str(resumen.get(rotulo) or '').strip()
        return datetime.strptime(crudo, '%d/%m/%Y').date()

    try:
        f_inicio = _fecha_resumen('Fecha Inicio')
        f_final = _fecha_resumen('Fecha Final')
    except ValueError:
        raise ValueError('El Resumen no trae Fecha Inicio/Fecha Final — '
                         '¿es el export correcto?')
    saldo_inicial = _monto_celda(resumen.get('Saldo Inicial'))
    saldo_final_resumen = _monto_celda(resumen.get('Saldo Final'))

    # ── Movimientos ──────────────────────────────────────────────────────────
    ws = wb['Movimientos']
    filas_crudas = list(ws.iter_rows(values_only=True))
    if not filas_crudas:
        raise ValueError('La hoja Movimientos viene vacía.')
    encabezado = [str(c or '') for c in filas_crudas[0]]

    def _col(nombre):
        for i, h in enumerate(encabezado):
            if nombre.lower() in h.lower():
                return i
        raise ValueError(f'No encuentro la columna «{nombre}» en Movimientos.')

    c_fecha, c_desc = _col('Fecha'), _col('Descripción')
    c_cargo, c_abono, c_saldo = _col('Cargos'), _col('Abonos'), _col('Saldo')

    filas, saldo_prev, cadena_rota = [], saldo_inicial, 0
    anio, mes_prev = f_inicio.year, f_inicio.month
    for cruda in filas_crudas[1:]:
        crudo_fecha = cruda[c_fecha]
        if crudo_fecha is None:
            continue
        if hasattr(crudo_fecha, 'year'):          # celda ya viene como fecha
            fecha = crudo_fecha.date() if hasattr(crudo_fecha, 'date') else crudo_fecha
        else:
            try:
                dia, mes = (int(x) for x in str(crudo_fecha).strip().split('/')[:2])
            except ValueError:
                continue
            if mes < mes_prev:                    # cruce de año (dic → ene)
                anio += 1
            mes_prev = mes
            fecha = date(anio, mes, dia)

        desc = str(cruda[c_desc] or '').strip()
        cargo, abono = _monto_celda(cruda[c_cargo]), _monto_celda(cruda[c_abono])
        saldo = _monto_celda(cruda[c_saldo])
        if cargo == 0 and abono == 0:
            continue
        if saldo_prev + abono - cargo != saldo:
            cadena_rota += 1
        saldo_prev = saldo

        clase, sentido, cat = clasificar_fila_cartola(desc, cargo, abono)
        huella = f'{fecha.isoformat()}|{desc}|{cargo}|{abono}|{saldo}'
        filas.append({
            'fecha': fecha.isoformat(), 'descripcion': desc,
            'cargo': cargo, 'abono': abono, 'saldo': saldo,
            'clase': clase, 'sentido': sentido, 'categoria': cat,
            'referencia': 'be:' + hashlib.sha1(huella.encode()).hexdigest()[:24],
        })

    # ── Cierres de mes cubiertos por el archivo ──────────────────────────────
    # El saldo de la última fila de un mes ES el cierre de ese mes, siempre que
    # el archivo siga en el mes siguiente (si no, el mes quedó a medias).
    cierres = {}
    for i, f in enumerate(filas):
        mes_fila = f['fecha'][:7]
        hay_despues = any(g['fecha'][:7] > mes_fila for g in filas[i + 1:])
        if hay_despues:
            cierres[mes_fila] = f['saldo']

    return {
        'cuenta_numero': str(resumen.get('N° Cuenta') or ''),
        'fecha_inicio': f_inicio.isoformat(), 'fecha_final': f_final.isoformat(),
        'saldo_inicial': saldo_inicial, 'saldo_final_resumen': saldo_final_resumen,
        'saldo_final_calculado': saldo_prev,
        'total_cargos': sum(f['cargo'] for f in filas),
        'total_abonos': sum(f['abono'] for f in filas),
        'cadena_rota': cadena_rota,
        'cuadra': (cadena_rota == 0 and saldo_prev == saldo_final_resumen),
        'cierres_mes': cierres,
        'filas': filas,
    }


def registrar_filas_cartola(filas, cierres_mes=None, cuenta_clave='bancoestado'):
    """Escribe las filas confirmadas de una cartola (fuente=captura) + cierres.

    Idempotente por referencia; respeta el corte de julio; y re-verifica el
    estado de cada fila (defensa doble contra dobles conteos: histórico
    mes-nivel de Scotiabank, barridos ya registrados como traspaso).
    Devuelve (creados, saltados).
    """
    from django.db import transaction

    from .models import (CategoriaFinanciera, CuentaFinanciera,
                         MovimientoFinanciero, SaldoMensual)

    cuenta = CuentaFinanciera.objects.get(clave=cuenta_clave)
    cats = {c.clave: c for c in CategoriaFinanciera.objects.all()}
    for clave, (nombre, clase) in CATEGORIAS_CARTOLA.items():
        if clave not in cats:
            cats[clave], _ = CategoriaFinanciera.objects.get_or_create(
                clave=clave, defaults={'nombre': nombre, 'clase': clase})

    def _categoria(clave):
        # Nunca dejar un gasto sin categoría: si la regla nombró una que no
        # existe, cae en «por clasificar» y se ve en el triaje.
        return cats.get(clave) or cats.get('por_clasificar')

    # Guardia contra la DOBLE CARGA desde dos exports distintos del mismo
    # período (Jorge 2026-08-10: los mismos gastos aparecían dos veces en
    # Scotiabank). La referencia incluye el saldo, y ese valor cambia entre el
    # export de movimientos y el estado mensual, así que no reconocía la fila
    # como repetida. Acá se compara por lo que NO cambia —fecha, monto y
    # glosa— y se permite crear solo las que falten: si el archivo trae dos
    # compras idénticas y en la base ya hay una, se crea una.
    from collections import Counter
    pendientes = Counter()
    for f in filas:
        firma = (f['fecha'], f['abono'] or f['cargo'], f['descripcion'])
        pendientes[firma] += 1
    for firma, en_archivo in list(pendientes.items()):
        fecha_iso, monto, desc = firma
        ya = MovimientoFinanciero.objects.filter(
            cuenta=cuenta, fecha=date.fromisoformat(fecha_iso), monto=monto,
            descripcion__endswith=desc[:180]).count()
        pendientes[firma] = max(0, en_archivo - ya)

    creados = saltados = 0
    with transaction.atomic():
        for f in filas:
            firma = (f['fecha'], f['abono'] or f['cargo'], f['descripcion'])
            if pendientes[firma] <= 0:
                saltados += 1
                continue
            pendientes[firma] -= 1
            fecha = date.fromisoformat(f['fecha'])
            monto = f['abono'] or f['cargo']
            if (fecha < COBERTURA_GASTOS_DESDE or monto <= 0
                    or estado_fila_cartola(cuenta_clave, f) != 'nuevo'):
                saltados += 1
                continue

            destino = (destino_puente(f['descripcion'], cuenta_clave)
                       if f['clase'] == 'gasto' else None)
            if destino and _traspaso_a_puente(
                    fecha, monto, f['descripcion'], f['referencia'],
                    cuenta, destino):
                creados += 1
                continue

            MovimientoFinanciero.objects.create(
                fecha=fecha, cuenta=cuenta, clase=f['clase'],
                sentido=f['sentido'], monto=monto,
                categoria=_categoria(f['categoria']),
                fuente='captura', referencia=f['referencia'],
                descripcion=f"Cartola {cuenta_clave}: {f['descripcion']}"[:255])
            creados += 1

        for mes_iso, saldo in (cierres_mes or {}).items():
            anio, mes = (int(x) for x in mes_iso.split('-'))
            SaldoMensual.objects.update_or_create(
                cuenta=cuenta, periodo=date(anio, mes, 1),
                defaults={'saldo_cierre': saldo, 'fuente': 'cartola',
                          'notas': 'Cierre derivado del export de cartola'})

    return creados, saltados


def _traspaso_a_puente(fecha, monto, descripcion, referencia, cuenta_origen,
                       destino_clave):
    """Registra Aremko → cuenta puente como traspaso, con sus dos piernas.

    No es un retiro: la plata sigue existiendo, solo cambió de bolsillo. Si
    la cartola de la cuenta puente ya se cargó y ese abono quedó esperando
    («por calzar»), se usa ESE como pierna que entra en vez de crear otra —
    si no, la misma plata entraría dos veces a su cuenta.
    Devuelve True si registró; False si la cuenta destino no existe todavía
    (mejor caer al camino normal que inventar una pierna suelta).
    """
    from datetime import timedelta

    from .models import CuentaFinanciera, MovimientoFinanciero

    destino = CuentaFinanciera.objects.filter(clave=destino_clave).first()
    if destino is None:
        return False

    ventana = (fecha - timedelta(days=2), fecha + timedelta(days=2))
    entra = MovimientoFinanciero.objects.filter(
        cuenta=destino, clase='traspaso', sentido='entra', monto=monto,
        traspaso_par__isnull=True, fecha__range=ventana).first()
    if entra is None:
        entra = MovimientoFinanciero.objects.filter(
            cuenta=destino, clase='ingreso', monto=monto,
            categoria__clave=CLAVE_POR_CALZAR, fecha__range=ventana).first()

    sale = MovimientoFinanciero.objects.create(
        fecha=fecha, cuenta=cuenta_origen, clase='traspaso', sentido='sale',
        monto=monto, fuente='captura', referencia=referencia,
        descripcion=f'Traspaso a {destino.nombre}: {descripcion}'[:255])

    if entra is not None:
        entra.clase = 'traspaso'
        entra.sentido = 'entra'
        entra.categoria = None
        entra.traspaso_par = sale
        entra.save(update_fields=['clase', 'sentido', 'categoria',
                                  'traspaso_par'])
    else:
        entra = MovimientoFinanciero.objects.create(
            fecha=fecha, cuenta=destino, clase='traspaso', sentido='entra',
            monto=monto, fuente='captura',
            referencia=f'{referencia}:entra'[:120], traspaso_par=sale,
            descripcion=f'Traspaso desde {cuenta_origen.nombre}: '
                        f'{descripcion}'[:255])
    sale.traspaso_par = entra
    sale.save(update_fields=['traspaso_par'])
    return True


def ordenar_por_cadena_de_saldos(movs, saldo_inicial):
    """Ordena movimientos usando el saldo como hilo. Devuelve (orden, sueltos).

    La Cartola en Línea de BancoEstado entrega las filas DESORDENADAS y con
    varias del mismo día, así que ordenar por fecha no basta: el saldo de cada
    fila dice cuál va después de cuál. Se arma la cadena desde el saldo
    inicial; lo que no encaja se devuelve aparte en vez de inventarle un
    lugar — una fila que no calza es justamente la señal de que falta algo.
    """
    pendientes = list(movs)
    orden, saldo = [], saldo_inicial
    while pendientes:
        for i, m in enumerate(pendientes):
            if saldo + m['abono'] - m['cargo'] == m['saldo']:
                orden.append(m)
                saldo = m['saldo']
                pendientes.pop(i)
                break
        else:
            break
    return orden, pendientes


def _parsear_bancoestado_en_linea(wb):
    """Cartola en Línea de la Chequera: hojas Resumen + Registros."""
    from datetime import datetime

    resumen = {}
    for fila in wb['Resumen'].iter_rows(values_only=True):
        rotulo = str(fila[0] or '').strip()
        if not rotulo:
            continue
        valor = next((c for c in fila[1:] if c not in (None, '')), None)
        if valor is not None:
            resumen[rotulo] = valor

    saldo_inicial = _monto_celda(resumen.get('Inicial'))
    saldo_final_resumen = _monto_celda(resumen.get('Saldo Contable')
                                       or resumen.get('Disponible'))

    crudas = list(wb['Registros'].iter_rows(values_only=True))
    if len(crudas) < 2:
        raise ValueError('La hoja Registros viene sin movimientos.')
    encabezado = [str(c or '').lower() for c in crudas[0]]

    def _col(pedazo):
        for i, h in enumerate(encabezado):
            if pedazo in h:
                return i
        raise ValueError(f'No encuentro la columna «{pedazo}» en Registros.')

    c_f, c_d = _col('fecha'), _col('descrip')
    c_c, c_a, c_s = _col('cargo'), _col('abono'), _col('saldo')

    movs = []
    for cruda in crudas[1:]:
        crudo = cruda[c_f]
        if crudo in (None, ''):
            continue
        if hasattr(crudo, 'year'):
            fecha = crudo.date() if hasattr(crudo, 'date') else crudo
        else:
            try:
                fecha = datetime.strptime(str(crudo).strip(), '%d/%m/%Y').date()
            except ValueError:
                continue
        cargo, abono = _monto_celda(cruda[c_c]), _monto_celda(cruda[c_a])
        if cargo == 0 and abono == 0:
            continue
        movs.append({'fecha': fecha,
                     'desc': str(cruda[c_d] or '').strip(),
                     'cargo': cargo, 'abono': abono,
                     'saldo': _monto_celda(cruda[c_s])})

    orden, sueltos = ordenar_por_cadena_de_saldos(movs, saldo_inicial)
    # Los que no encajaron van al final por fecha: se muestran igual, pero la
    # página avisa que la cadena está rota.
    orden.extend(sorted(sueltos, key=lambda m: m['fecha']))

    filas = []
    for m in orden:
        clase, sentido, cat = clasificar_fila_cartola(m['desc'], m['cargo'],
                                                      m['abono'])
        huella = (f"{m['fecha'].isoformat()}|{m['desc']}|{m['cargo']}|"
                  f"{m['abono']}|{m['saldo']}")
        filas.append({
            'fecha': m['fecha'].isoformat(), 'descripcion': m['desc'],
            'cargo': m['cargo'], 'abono': m['abono'], 'saldo': m['saldo'],
            'clase': clase, 'sentido': sentido, 'categoria': cat,
            'referencia': 'be:' + hashlib.sha1(huella.encode()).hexdigest()[:24],
        })
    if not filas:
        raise ValueError('El archivo no trae movimientos.')

    cierres = {}
    for i, f in enumerate(filas):
        mes_fila = f['fecha'][:7]
        if any(g['fecha'][:7] > mes_fila for g in filas[i + 1:]):
            cierres[mes_fila] = f['saldo']

    saldo_final = filas[-1]['saldo']
    return {
        'cuenta_numero': str(resumen.get('Chequera Electrónica') or ''),
        'fecha_inicio': min(f['fecha'] for f in filas),
        'fecha_final': max(f['fecha'] for f in filas),
        'saldo_inicial': saldo_inicial,
        'saldo_final_resumen': saldo_final_resumen or saldo_final,
        'saldo_final_calculado': saldo_final,
        'total_cargos': sum(f['cargo'] for f in filas),
        'total_abonos': sum(f['abono'] for f in filas),
        'cadena_rota': len(sueltos),
        'cuadra': (not sueltos and
                   (not saldo_final_resumen or saldo_final_resumen == saldo_final)),
        'cierres_mes': cierres,
        'filas': filas,
    }


# ── F4 paso 3b: cartola Scotiabank (export .xls del portal) ──────────────────
# Formato real verificado 2026-08-08 (typeDesc.xls): hoja única 'Data' con
# metadatos rótulo/valor arriba (Saldo Disponible, Fecha Desde/Hasta) y luego
# encabezado Fecha DD-MM-AAAA | Descripción | Sucursal | N° Doc. | Cargos
# (NEGATIVOS) | Abonos | Saldo — en orden DESCENDENTE (lo más nuevo primero).

RUT_AREMKO_SIN_DV = '76485192'


def clasificar_fila_scotiabank(descripcion, cargo, abono):
    """(clase, sentido, categoria_clave, propio) para una fila Scotiabank."""
    d = (descripcion or '').upper()
    propio = RUT_AREMKO_SIN_DV in d or 'AREMKO' in d
    if abono > 0:
        # Abono propio = barrido desde otra cuenta de Aremko (traspaso, no
        # ingreso). El match contra el traspaso ya registrado se hace aparte.
        if propio:
            return 'traspaso', 'entra', '', True
        return 'ingreso', 'entra', 'transferencias_recibidas', False
    if 'SEGURO' in d:
        return 'gasto', 'sale', 'seguros', False
    if 'COMISION' in d or d.startswith('IVA'):
        return 'gasto', 'sale', 'comisiones', False
    return 'gasto', 'sale', 'por_clasificar', False


def parsear_filas_scotiabank(filas_crudas):
    """Núcleo puro: recibe las filas de la hoja como listas y devuelve el
    mismo shape que la cartola BancoEstado. Testeable sin xlrd.

    Soporta las DOS variantes reales del portal (verificadas 2026-08-08):
    - Movimientos de la línea (typeDesc, 7 columnas, DESCENDENTE, meta
      'Saldo Disponible' / 'Número Línea')
    - Estado de cuenta mensual (6 columnas, ASCENDENTE, meta 'Saldo
      Anterior' / 'Saldo Actual' / 'Numero Cuenta')
    Las columnas se mapean POR NOMBRE del encabezado y el orden se detecta
    comparando la primera y la última fecha.
    """
    from datetime import datetime

    meta, encabezado_en = {}, None
    for i, fila in enumerate(filas_crudas):
        primera = str(fila[0] or '').strip()
        if primera == 'Fecha' and any('Descripci' in str(c or '') for c in fila):
            encabezado_en = i
            break
        if primera and len(fila) > 1 and fila[1] not in ('', None):
            meta[primera] = fila[1]
    if encabezado_en is None:
        raise ValueError('No encuentro el encabezado de movimientos — '
                         '¿es el export de Scotiabank?')

    encabezado = [str(c or '').lower() for c in filas_crudas[encabezado_en]]

    def _col(pedazo):
        for j, h in enumerate(encabezado):
            if pedazo in h:
                return j
        raise ValueError(f'No encuentro la columna «{pedazo}» en el export.')

    c_desc, c_cargo = _col('descripci'), _col('cargo')
    c_abono, c_saldo = _col('abono'), _col('saldo')

    movimientos = []
    for fila in filas_crudas[encabezado_en + 1:]:
        crudo_fecha = str(fila[0] or '').strip()
        if not crudo_fecha:
            continue
        try:
            fecha = datetime.strptime(crudo_fecha, '%d-%m-%Y').date()
        except ValueError:
            continue
        desc = str(fila[c_desc] or '').strip()
        cargo = abs(_monto_celda(fila[c_cargo]))
        abono = _monto_celda(fila[c_abono])
        saldo = _monto_celda(fila[c_saldo])
        if cargo == 0 and abono == 0:
            continue
        movimientos.append((fecha, desc, cargo, abono, saldo))

    # Orden: la variante typeDesc viene de lo más nuevo a lo más viejo; el
    # estado de cuenta mensual ya viene cronológico. Se detecta por fechas.
    if len(movimientos) > 1 and movimientos[0][0] > movimientos[-1][0]:
        movimientos.reverse()

    # Ancla del inicio si la trae la cabecera (estado de cuenta mensual).
    saldo_anterior = _monto_celda(meta.get('Saldo Anterior'))

    filas, cadena_rota = [], 0
    saldo_prev = saldo_anterior if saldo_anterior else None
    for fecha, desc, cargo, abono, saldo in movimientos:
        if saldo_prev is not None and saldo_prev + abono - cargo != saldo:
            cadena_rota += 1
        saldo_prev = saldo
        clase, sentido, cat, propio = clasificar_fila_scotiabank(desc, cargo, abono)
        huella = f'{fecha.isoformat()}|{desc}|{cargo}|{abono}|{saldo}'
        filas.append({
            'fecha': fecha.isoformat(), 'descripcion': desc,
            'cargo': cargo, 'abono': abono, 'saldo': saldo,
            'clase': clase, 'sentido': sentido, 'categoria': cat,
            'propio': propio,
            'referencia': 'sc:' + hashlib.sha1(huella.encode()).hexdigest()[:24],
        })
    if not filas:
        raise ValueError('El export no trae movimientos.')

    saldo_inicial = (saldo_anterior or
                     filas[0]['saldo'] - filas[0]['abono'] + filas[0]['cargo'])
    saldo_final = filas[-1]['saldo']

    cierres = {}
    for i, f in enumerate(filas):
        mes_fila = f['fecha'][:7]
        if any(g['fecha'][:7] > mes_fila for g in filas[i + 1:]):
            cierres[mes_fila] = f['saldo']
    # El estado de cuenta MENSUAL cerrado (Fecha Hasta = último día del mes)
    # también ancla el cierre de ese mes, aunque no haya filas del siguiente.
    hasta = str(meta.get('Fecha Hasta') or '').strip()
    if hasta:
        try:
            from datetime import datetime as _dt
            from datetime import timedelta as _td
            f_hasta = _dt.strptime(hasta, '%d-%m-%Y').date()
            if (f_hasta + _td(days=1)).month != f_hasta.month:
                cierres.setdefault(f'{f_hasta.year}-{f_hasta.month:02d}',
                                   saldo_final)
        except ValueError:
            pass

    # Saldo final declarado por la cabecera, según la variante.
    declarado = (_monto_celda(meta.get('Saldo Disponible')) or
                 _monto_celda(meta.get('Saldo Actual')))
    n_cuenta = meta.get('Número Línea') or meta.get('Numero Cuenta') or ''
    if isinstance(n_cuenta, float) and n_cuenta.is_integer():
        n_cuenta = int(n_cuenta)
    return {
        'cuenta_numero': str(n_cuenta),
        'fecha_inicio': filas[0]['fecha'], 'fecha_final': filas[-1]['fecha'],
        'saldo_inicial': saldo_inicial, 'saldo_final_resumen': declarado or saldo_final,
        'saldo_final_calculado': saldo_final,
        'total_cargos': sum(f['cargo'] for f in filas),
        'total_abonos': sum(f['abono'] for f in filas),
        'cadena_rota': cadena_rota,
        'cuadra': (cadena_rota == 0 and
                   (not declarado or declarado == saldo_final)),
        'cierres_mes': cierres,
        'filas': filas,
    }


def parsear_cartola_scotiabank(archivo):
    """Capa fina: lee el .xls con xlrd y delega en el núcleo puro."""
    import xlrd

    wb = xlrd.open_workbook(file_contents=archivo.read())
    hoja = wb.sheet_by_index(0)
    crudas = [[hoja.cell_value(i, j) for j in range(hoja.ncols)]
              for i in range(hoja.nrows)]
    return parsear_filas_scotiabank(crudas)


# ── F7: cuenta personal de Alda — cartola BSA.dat de Scotia Connect ──────────
# Formato real verificado 2026-08-08 (archivo de Jorge): texto ASCII con ';',
# cabecera de metadatos (';Numero Cuenta : 99-00138-96', ';Fecha Hasta :
# DD/MM/AAAA'), encabezado 'Fecha;Descripcion;NroDoc.;Cargos;Abonos;Saldo' y
# filas 'DDMMYYYY;desc;nrodoc;cargo;abono;saldo' con montos rellenos de ceros
# y coma decimal ('0000001000000,00') y saldo con signo ('+0000001061836,00').

CUENTA_ALDA_NUMERO = '99-00138'   # 99-00138-96, se compara por prefijo

# clave → (nombre, clase, grupo). El default de un cargo es PERSONAL (decisión
# F7): lo que sea de la empresa (p. ej. el vale vista de impuestos) se
# reclasifica o cae por regla propia.
RUT_ALDA_SIN_DV = '11744727'

CATEGORIAS_ALDA = {
    'tarjeta_alda': ('Tarjeta de crédito (Alda)', 'gasto', 'personales_alda'),
    'banco_alda': ('Banco: comisiones, seguros e intereses (Alda)', 'gasto',
                   'personales_alda'),
    'traslado_alda': ('Traslado a otras cuentas de Alda', 'gasto',
                      'personales_alda'),
    'personales_alda_pc': ('Personales Alda (por clasificar)', 'gasto',
                           'personales_alda'),
    # Un vale vista es un INSTRUMENTO, no un destino: el de julio 2026 era
    # impuestos de Aremko (dato de Jorge), pero el siguiente puede ser otra
    # cosa. Queda por clasificar y VISIBLE, no adivinado.
    'vale_vista_alda': ('Vale vista emitido (asignar destino)', 'gasto',
                        'otros'),
    # Los abonos que NO vienen de Aremko (AFP, IPS, línea de crédito, plata
    # de ella o de él) sí están en el saldo del banco: se registran para que
    # el saldo de la cuenta puente cuadre con la app. No tocan el resultado
    # del negocio — los ingresos del tablero salen de las reservas.
    'abonos_personales': ('Abonos personales (no son de Aremko)', 'ingreso',
                          'ingresos'),
    # Abono que SÍ viene de Aremko pero no calzó con ningún retiro por monto
    # exacto (pasa seguido: la glosa del banco redondea o la transferencia se
    # partió). Entra igual —si no, el saldo de la cuenta queda mal— y espera
    # el calce a mano en /finanzas/calzar-retiros/.
    'abono_aremko_por_calzar': ('Abono desde Aremko por calzar', 'ingreso',
                                'ingresos'),
}

CLAVE_POR_CALZAR = 'abono_aremko_por_calzar'

# Transferencias de Aremko a una cuenta que SÍ seguimos: no son retiro, son
# traspaso — la plata cambió de bolsillo y desde ahí se pagan cuentas de
# Aremko y gastos personales (Jorge 2026-08-10). El retiro de verdad ocurre
# después, cuando esa plata se gasta en algo personal, y ahí lo dice la
# cartola de esa cuenta. A una cuenta que NO seguimos (el BCI de Alda, Mach,
# Martín) sí es retiro: la plata sale de nuestra vista.
# Se evalúan en orden: lo que va a None es retiro aunque el nombre calce.
DESTINO_PUENTE = [
    ('ALDA BCI', None), ('BCI ALDA', None), ('MACH', None),
    ('TOLOZA POBLETE ALDA', 'scotiabank_alda'),
    ('ALDA ANGELICA TOLOZA', 'scotiabank_alda'),
    ('ANGELICA TOLOZA', 'scotiabank_alda'),
    ('ALDA TOLOZA', 'scotiabank_alda'),
    ('JORGE ANTONIO AGUILERA', 'cuentarut_jorge'),
    ('JORGE AGUILERA', 'cuentarut_jorge'),
    # BancoEstado escribe la TEF dentro del mismo banco con los apellidos a
    # secas ("TEF BANCOESTADO A AGUILERA GONZALEZ"), sin el nombre — y con
    # eso no calzaba con 'JORGE AGUILERA' y el retiro quedaba como gasto por
    # clasificar (caso real 20/08/2026: $300.000 que Jorge se traspasó para
    # pagar Google Ads desde su CuentaRUT). Jorge confirmó el 22/08 que nadie
    # más «Aguilera González» recibe plata de Aremko en una cuenta BancoEstado,
    # así que la variante INTRA-banco es inequívocamente su CuentaRUT. La
    # interbancaria ("TRANSFERENCIA A AGUILERA GONZALEZ") NO se toca: ahí el
    # apellido es ambiguo con Cristian (infraestructura) y Martín (retiro).
    ('TEF BANCOESTADO A AGUILERA GONZALEZ', 'cuentarut_jorge'),
]


def destino_puente(descripcion, cuenta_origen=None):
    """A qué cuenta puente va este cargo, o None si es retiro de verdad.

    Si el destino es la MISMA cuenta que lo origina, no hay traspaso: es un
    traslado a otra cuenta de esa persona que no seguimos.
    """
    d = (descripcion or '').upper()
    for patron, destino in DESTINO_PUENTE:
        if patron in d:
            return None if destino == cuenta_origen else destino
    return None


def marcar_traspaso_puente(fila, cuenta_clave, nombres_cta=None):
    """Marca una fila de cartola como «esto va a una cuenta puente», SOLO para
    mostrarla en la propuesta.

    REGLA DURA: no toca `clase` ni `categoria`. Ese mismo diccionario viaja
    firmado en el payload hasta la confirmación, y `registrar_filas_cartola`
    decide si arma el traspaso de DOS piernas preguntando
    `if f['clase'] == 'gasto'`. El 22/08/2026 la vista previa ponía
    `f['clase'] = 'traspaso'` para pintarlo bonito y esa condición dejaba de
    cumplirse: el movimiento se guardaba como una pierna «sale» huérfana, sin
    par, y los traspasos dejaban de cuadrar ($809.331 en tres filas).
    """
    if fila.get('clase') != 'gasto':
        return fila
    destino = destino_puente(fila.get('descripcion'), cuenta_clave)
    if destino:
        fila['es_traspaso_puente'] = True
        fila['destino_puente'] = (nombres_cta or {}).get(destino, destino)
    return fila


def clasificar_fila_alda(descripcion, cargo, abono):
    """(clase, sentido, categoria_clave, propio) para una fila del BSA.dat.

    - Abono desde Aremko → 'traspaso' (el retiro ya registrado se convierte).
    - Abono de otra fuente (incluida su línea de crédito) → 'personal':
      plata de ella, no se registra.
    - Cargo → gasto, con el default PERSONAL: lo de la empresa se reclasifica
      a mano (los comercios REDCOMPRA son ambiguos por naturaleza).
    """
    from .reglas_glosa import categoria_por_glosa

    d = (descripcion or '').upper()
    if abono > 0:
        if RUT_AREMKO_SIN_DV in d or 'AREMKO' in d:
            return 'traspaso', 'entra', '', True
        return 'personal', 'entra', '', False
    # La lista de Jorge manda sobre todo lo demás: si él declaró que esa
    # glosa es de Aremko, lo es aunque la haya pagado ella.
    de_aremko = categoria_por_glosa(d)
    if de_aremko:
        return 'gasto', 'sale', de_aremko, False
    if 'PAGO TARJ' in d:
        return 'gasto', 'sale', 'tarjeta_alda', False
    # Scotiabank llama «EMISION DE VIGENTE» a la emisión de un vale vista.
    if 'VALE VISTA' in d or 'EMISION DE VIGENTE' in d:
        return 'gasto', 'sale', 'vale_vista_alda', False
    if RUT_ALDA_SIN_DV in d or 'ALDA TOLOZA' in d:
        return 'gasto', 'sale', 'traslado_alda', False
    if ('SEG' in d or 'DESGRAVAMEN' in d or 'COMISION' in d
            or d.startswith('IVA') or 'LINEA' in d or 'L.CREDITO' in d):
        return 'gasto', 'sale', 'banco_alda', False
    return 'gasto', 'sale', 'personales_alda_pc', False


def _monto_bsa(texto):
    """'+0000001061836,00' → 1061836 · '' → 0 (los decimales son siempre ,00)."""
    t = str(texto or '').strip()
    if not t:
        return 0
    signo = -1 if t.startswith('-') else 1
    entero = t.lstrip('+-').split(',')[0].lstrip('0')
    return signo * int(entero or 0)


def parsear_cartola_alda(archivo):
    """Parsea el BSA.dat al mismo shape de las otras cartolas."""
    from datetime import datetime

    crudo = archivo.read()
    texto = crudo.decode('latin-1') if isinstance(crudo, bytes) else crudo

    meta, movimientos = {}, []
    for linea in texto.splitlines():
        limpia = linea.strip()
        if not limpia:
            continue
        if limpia.startswith(';'):
            clave, _, valor = limpia[1:].partition(':')
            if valor.strip():
                meta[clave.strip()] = valor.strip()
            continue
        if limpia.startswith('Fecha;'):
            continue
        partes = linea.split(';')
        if len(partes) < 6:
            continue
        try:
            fecha = datetime.strptime(partes[0].strip(), '%d%m%Y').date()
        except ValueError:
            continue
        desc = partes[1].strip()
        cargo = abs(_monto_bsa(partes[3]))
        abono = _monto_bsa(partes[4])
        saldo = _monto_bsa(partes[5])
        if cargo == 0 and abono == 0:
            continue
        movimientos.append((fecha, desc, cargo, abono, saldo))

    if not movimientos:
        raise ValueError('El BSA.dat no trae movimientos — '
                         '¿es el export de Scotia Connect?')
    if len(movimientos) > 1 and movimientos[0][0] > movimientos[-1][0]:
        movimientos.reverse()

    filas, cadena_rota = [], 0
    saldo_prev = None
    for fecha, desc, cargo, abono, saldo in movimientos:
        if saldo_prev is not None and saldo_prev + abono - cargo != saldo:
            cadena_rota += 1
        saldo_prev = saldo
        clase, sentido, cat, propio = clasificar_fila_alda(desc, cargo, abono)
        huella = f'{fecha.isoformat()}|{desc}|{cargo}|{abono}|{saldo}'
        filas.append({
            'fecha': fecha.isoformat(), 'descripcion': desc,
            'cargo': cargo, 'abono': abono, 'saldo': saldo,
            'clase': clase, 'sentido': sentido, 'categoria': cat,
            'propio': propio,
            'referencia': 'alda:' + hashlib.sha1(huella.encode()).hexdigest()[:24],
        })

    saldo_inicial = filas[0]['saldo'] - filas[0]['abono'] + filas[0]['cargo']
    saldo_final = filas[-1]['saldo']

    cierres = {}
    for i, f in enumerate(filas):
        mes_fila = f['fecha'][:7]
        if any(g['fecha'][:7] > mes_fila for g in filas[i + 1:]):
            cierres[mes_fila] = f['saldo']
    hasta = meta.get('Fecha Hasta', '')
    if hasta:
        try:
            from datetime import timedelta as _td
            f_hasta = datetime.strptime(hasta, '%d/%m/%Y').date()
            if (f_hasta + _td(days=1)).month != f_hasta.month:
                cierres.setdefault(f'{f_hasta.year}-{f_hasta.month:02d}',
                                   saldo_final)
        except ValueError:
            pass

    return {
        'cuenta_numero': meta.get('Numero Cuenta', ''),
        'fecha_inicio': filas[0]['fecha'], 'fecha_final': filas[-1]['fecha'],
        'saldo_inicial': saldo_inicial, 'saldo_final_resumen': saldo_final,
        'saldo_final_calculado': saldo_final,
        'total_cargos': sum(f['cargo'] for f in filas),
        'total_abonos': sum(f['abono'] for f in filas),
        'cadena_rota': cadena_rota,
        'cuadra': cadena_rota == 0,
        'cierres_mes': cierres,
        'filas': filas,
    }


# ── F7b: CuentaRUT de Jorge — sin export, se cargan movimientos pegados ──────
# El portal Personas de BancoEstado no da el mismo .xlsx que Empresas, así que
# la cuenta se alimenta pegando los movimientos tal como se leen en la app
# (negativo = cargo, positivo = abono).

def clasificar_fila_cuentarut(descripcion, cargo, abono):
    """(clase, sentido, categoria_clave, propio) para la CuentaRUT de Jorge.

    Es una cuenta MIXTA: paga publicidad e infraestructura de Aremko y a la
    vez el presupuesto de Martín y gastos personales. Lo que no calza con una
    regla clara queda POR CLASIFICAR (grupo otros) — ni se infla el gasto de
    la empresa ni se esconde como personal.
    """
    from .reglas_glosa import categoria_por_glosa

    d = (descripcion or '').upper()
    if abono > 0:
        if RUT_AREMKO_SIN_DV in d or 'AREMKO' in d:
            return 'traspaso', 'entra', '', True
        return 'personal', 'entra', '', False
    # Martín es distinto de Jorge por una razón estructural, no personal: su
    # cuenta NO la seguimos. La plata que va hacia él sale de nuestra vista y
    # no hay cartola donde ver en qué se gastó, así que el retiro se registra
    # acá mismo. Un cargo con el nombre del PROPIO Jorge, en cambio, queda por
    # clasificar: él decide si fue retiro o no (decisión suya 2026-08-10).
    if 'MARTIN AGUILERA' in d:
        return 'gasto', 'sale', 'personales_martin', False
    # Las tres jubilaciones de Jorge entran acá («Abono Convenio …») y salen
    # casi íntegras al día siguiente: parte a su otra cuenta y parte a
    # Datamatic, su otra empresa (confirmado por él el 22/08/2026). Sin estas
    # dos reglas caían como «gasto por clasificar» y sumaban ~$400.000
    # mensuales de gasto fantasma al resultado de Aremko. Van por GLOSA y no
    # por monto porque las jubilaciones están en UF: el monto cambia todos los
    # meses, la glosa no.
    #
    # Esto es ESTRUCTURA, no criterio: plata que sale del bolsillo de Jorge
    # hacia sí mismo o hacia otra empresa suya jamás es costo del spa. Y si
    # alguna de esas transferencias fuera en realidad un retiro de plata de
    # Aremko, el resultado no cambia: ambas categorías viven en el grupo
    # `personales_jorge`, igual que «Retiros Jorge».
    #
    # Solo rige para ESTA cuenta: si Aremko le paga a Datamatic desde sus
    # propias cuentas, eso sí es un gasto del spa y lo clasifica su cartola.
    if 'JORGE ANTONIO AGUILERA' in d or 'JORGE AGUILERA' in d:
        return 'gasto', 'sale', 'traslado_cuenta_propia', False
    if 'DATAMATIC' in d:
        return 'gasto', 'sale', 'aporte_datamatic', False
    de_aremko = categoria_por_glosa(d)
    if de_aremko:
        return 'gasto', 'sale', de_aremko, False
    return 'gasto', 'sale', 'por_clasificar', False


# cuenta → (función de clasificación, grupo del retiro que se convierte)
# Las cuentas del NEGOCIO. Un abono que dice «llegó desde Aremko» solo puede
# tener su origen acá: si sale de la tarjeta de Alda o de la CuentaRUT de
# Jorge, no es plata de Aremko por definición (22/08/2026).
CUENTAS_AREMKO = ('bancoestado', 'scotiabank', 'mercado_pago', 'efectivo')

CUENTAS_PUENTE = {
    'scotiabank_alda': (clasificar_fila_alda, 'personales_alda'),
    'cuentarut_jorge': (clasificar_fila_cuentarut, 'personales_jorge'),
}


def _monto_pegado(texto):
    """'-$22.687' → -22687 · '200.000' → 200000 · '' → None si no es número."""
    t = str(texto or '').strip().replace('$', '').replace(' ', '')
    t = t.replace('.', '').replace(',', '')
    if not t or t in ('-', '+'):
        return None
    try:
        return int(t)
    except ValueError:
        return None


def parsear_pegado(texto):
    """Movimientos pegados → (filas_crudas, errores).

    Una línea por movimiento: `fecha ; descripción ; monto`, separadas por
    `;` o tabulador. Fecha DD-MM-AAAA, DD/MM/AAAA o AAAA-MM-DD. Monto tal
    como se ve en la app: NEGATIVO es cargo, positivo es abono.
    Las líneas que no se entienden se DEVUELVEN como error con su número —
    saltarlas en silencio sería perder plata sin avisar.
    """
    from datetime import datetime

    filas, errores = [], []
    for n, linea in enumerate(texto.splitlines(), start=1):
        cruda = linea.strip()
        if not cruda or cruda.startswith('#'):
            continue
        partes = [p.strip() for p in
                  (cruda.split('\t') if '\t' in cruda else cruda.split(';'))]
        if len(partes) < 3:
            errores.append((n, cruda, 'faltan campos (fecha ; glosa ; monto)'))
            continue
        fecha = None
        for formato in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d'):
            try:
                fecha = datetime.strptime(partes[0], formato).date()
                break
            except ValueError:
                continue
        if fecha is None:
            errores.append((n, cruda, f'no entiendo la fecha «{partes[0]}»'))
            continue
        monto = _monto_pegado(partes[2])
        if monto is None or monto == 0:
            errores.append((n, cruda, f'no entiendo el monto «{partes[2]}»'))
            continue
        filas.append({'fecha': fecha, 'descripcion': partes[1],
                      'cargo': -monto if monto < 0 else 0,
                      'abono': monto if monto > 0 else 0})
    return filas, errores


def preparar_filas_manual(texto, cuenta_clave):
    """Pegado → filas con clase/categoría/referencia, listas para la vista."""
    clasificar, _ = CUENTAS_PUENTE.get(cuenta_clave,
                                       (clasificar_fila_cuentarut, ''))
    crudas, errores = parsear_pegado(texto)
    vistas, filas = {}, []
    for f in crudas:
        clase, sentido, cat, propio = clasificar(f['descripcion'],
                                                 f['cargo'], f['abono'])
        # Dos movimientos idénticos el mismo día son posibles: el contador de
        # repetición los hace distintos sin romper la idempotencia al repegar.
        base = f"{f['fecha'].isoformat()}|{f['descripcion']}|{f['cargo']}|{f['abono']}"
        vistas[base] = vistas.get(base, 0) + 1
        huella = f'{base}|{vistas[base]}'
        filas.append({
            'fecha': f['fecha'].isoformat(), 'descripcion': f['descripcion'],
            'cargo': f['cargo'], 'abono': f['abono'], 'saldo': 0,
            'clase': clase, 'sentido': sentido, 'categoria': cat,
            'propio': propio,
            'referencia': f'man:{cuenta_clave}:' +
                          hashlib.sha1(huella.encode()).hexdigest()[:20],
        })
    return filas, errores


def registrar_filas_puente(filas, cuenta_clave, cierres_mes=None):
    """Escribe filas de una cuenta puente (Alda o CuentaRUT).

    Devuelve (creados, saltados, convertidos).

    - Cargos → gastos en esa cuenta, ya clasificados.
    - Abonos desde Aremko → buscan el RETIRO ya registrado (gasto del grupo
      familia de esa persona, mismo monto, ±2 días) y lo CONVIERTEN en
      traspaso con la pierna entra en la cuenta puente: la plata deja de
      contarse dos veces y el gasto real pasa a ser lo que ahí se paga.
    - Abonos propios de la persona ('personal') no se tocan.
    """
    from datetime import timedelta

    from django.db import transaction

    from .models import (CategoriaFinanciera, CuentaFinanciera,
                         MovimientoFinanciero, SaldoMensual)
    from .reglas import PLAN_CUENTAS

    cuenta = CuentaFinanciera.objects.get(clave=cuenta_clave)
    _, grupo_retiro = CUENTAS_PUENTE[cuenta_clave]
    etiqueta = cuenta_clave.replace('_', ' ')

    definiciones = dict(PLAN_CUENTAS)
    definiciones.update(CATEGORIAS_ALDA)
    cats = {}

    def _categoria(clave):
        if not clave:
            return None
        if clave not in cats:
            nombre, clase, grupo = definiciones.get(
                clave, (clave.replace('_', ' ').title(), 'gasto', 'otros'))
            cats[clave], _ = CategoriaFinanciera.objects.get_or_create(
                clave=clave,
                defaults={'nombre': nombre, 'clase': clase, 'grupo': grupo})
        return cats[clave]

    creados = saltados = convertidos = 0
    with transaction.atomic():
        for f in filas:
            fecha = date.fromisoformat(f['fecha'])
            monto = f['abono'] or f['cargo']
            if (monto <= 0 or fecha < COBERTURA_GASTOS_DESDE
                    or MovimientoFinanciero.objects.filter(
                        referencia=f['referencia']).exists()):
                saltados += 1
                continue

            if f.get('clase') == 'personal':
                MovimientoFinanciero.objects.create(
                    fecha=fecha, cuenta=cuenta, clase='ingreso',
                    sentido='entra', monto=monto,
                    categoria=_categoria('abonos_personales'),
                    fuente='captura', referencia=f['referencia'],
                    descripcion=f"Cartola {etiqueta}: {f['descripcion']}"[:255])
                creados += 1
                continue

            if f.get('propio') and f['clase'] == 'traspaso':
                ventana = (fecha - timedelta(days=2), fecha + timedelta(days=2))
                # Si la cartola de Aremko ya se cargó, la pierna que entra
                # existe: repetirla metería la misma plata dos veces.
                if MovimientoFinanciero.objects.filter(
                        cuenta=cuenta, clase='traspaso', sentido='entra',
                        monto=monto, fecha__range=ventana).exists():
                    saltados += 1
                    continue
                candidatos = list(MovimientoFinanciero.objects.filter(
                    clase='gasto', sentido='sale', monto=monto,
                    categoria__grupo=grupo_retiro, fecha__range=ventana)
                    .exclude(cuenta=cuenta))
                if not candidatos:
                    # Sin calce exacto NO se descarta: la plata entró de
                    # verdad. Queda como abono por calzar para no dejar el
                    # saldo de la cuenta corto (Jorge 2026-08-09).
                    MovimientoFinanciero.objects.create(
                        fecha=fecha, cuenta=cuenta, clase='ingreso',
                        sentido='entra', monto=monto,
                        categoria=_categoria(CLAVE_POR_CALZAR),
                        fuente='captura', referencia=f['referencia'],
                        descripcion=f"Cartola {etiqueta}: {f['descripcion']}"[:255])
                    creados += 1
                    continue
                retiro = min(candidatos,
                             key=lambda m: abs((m.fecha - fecha).days))
                entra = MovimientoFinanciero.objects.create(
                    fecha=fecha, cuenta=cuenta, clase='traspaso',
                    sentido='entra', monto=monto, fuente='captura',
                    referencia=f['referencia'], traspaso_par=retiro,
                    descripcion=f"Cartola {etiqueta}: {f['descripcion']}"[:255])
                retiro.clase = 'traspaso'
                retiro.categoria = None
                retiro.traspaso_par = entra
                retiro.save(update_fields=['clase', 'categoria',
                                           'traspaso_par'])
                convertidos += 1
                continue

            MovimientoFinanciero.objects.create(
                fecha=fecha, cuenta=cuenta, clase='gasto', sentido='sale',
                monto=monto, categoria=_categoria(f['categoria']),
                fuente='captura', referencia=f['referencia'],
                descripcion=f"Cartola {etiqueta}: {f['descripcion']}"[:255])
            creados += 1

        for mes_iso, saldo in (cierres_mes or {}).items():
            anio, mes = (int(x) for x in mes_iso.split('-'))
            SaldoMensual.objects.update_or_create(
                cuenta=cuenta, periodo=date(anio, mes, 1),
                defaults={'saldo_cierre': saldo, 'fuente': 'cartola',
                          'notas': 'Cierre derivado de la cartola'})

    return creados, saltados, convertidos


def registrar_filas_alda(filas, cierres_mes=None):
    """Compatibilidad: la cartola de Alda por el camino genérico."""
    return registrar_filas_puente(filas, 'scotiabank_alda', cierres_mes)


# ── F7c: tarjetas de crédito de Alda (PDF del portal Scotiabank) ─────────────
# Formato real verificado 2026-08-10: impresión a PDF del portal. Cada línea es
# «DD/MM/AAAA  COMERCIO  CIUDAD  $MONTO», las fechas vienen DESORDENADAS y las
# descripciones largas se parten en dos renglones. Los montos NEGATIVOS («PAGO
# EN EFECTIVO») son los pagos de la tarjeta, no compras. El archivo NO dice a
# qué tarjeta corresponde: eso lo elige quien la sube.

TARJETAS_ALDA = ('tarjeta_alda_1', 'tarjeta_alda_2')

_RE_MOV_TARJETA = re.compile(
    r'^(\d{2}/\d{2}/\d{4})\s+(.+?)\s+\$\s*(-?[\d.]+)\s*$')
# Ruido de la impresión del navegador: encabezado repetido, hora y pie.
_RE_RUIDO_TARJETA = re.compile(
    r'^(about:blank|.*Página \d+ de \d+|\d{2}-\d{2}-\d{2},.*|F+E+C+H+A+.*)$')


def clasificar_compra_tarjeta(descripcion):
    """Categoría de una compra de la tarjeta.

    Es una tarjeta MIXTA (Aremko y personal). Solo se asigna a Aremko lo que
    esté en la lista de Jorge; el resto queda POR CLASIFICAR. Antes consultaba
    las reglas generales del plan de cuentas, y eso metía a Aremko cualquier
    compra en un comercio que también aparece en su vida personal (regla de
    Jorge 2026-08-10).
    """
    from .reglas_glosa import categoria_por_glosa

    d = (descripcion or '').upper()
    # "TRASPASO DEUDA INTERNAC(IONAL)" la escribe el BANCO al pasar al estado
    # nacional la deuda de las compras en dólares. No es un traspaso entre
    # cuentas nuestras (pese al nombre) ni una compra identificable: es el
    # gasto internacional del ciclo, agregado. Va a su propia categoría para
    # que no se confunda con traspaso ni se pierda entre lo por clasificar —
    # Jorge la reparte (publicidad / infraestructura / personal) mirando el
    # detalle del portal. Decidido el 22/08/2026.
    if 'TRASPASO DEUDA' in d:
        return 'tarjeta_internacional'
    de_aremko = categoria_por_glosa(d)
    if de_aremko:
        return de_aremko
    if any(p in d for p in ('COMISION', 'INTERES', 'SERVICIO DE ACTIVIDAD',
                            'SEGURO', 'DESGRAVAMEN')):
        return 'banco_alda'
    return 'por_clasificar'


def parsear_lineas_tarjeta(texto, cuenta_clave):
    """Núcleo puro: el texto del PDF → filas. Testeable sin pdfplumber.

    Las líneas sin fecha son la continuación de la descripción anterior (el
    portal parte «PUERTO VARAS» en dos renglones); se pegan a la fila de
    arriba en vez de descartarse, si no el comercio queda a medias.
    """
    from datetime import datetime

    filas, vistas = [], {}
    for cruda in (texto or '').splitlines():
        linea = ' '.join(cruda.split())
        if not linea or _RE_RUIDO_TARJETA.match(linea):
            continue
        m = _RE_MOV_TARJETA.match(linea)
        if not m:
            if filas:                      # continuación de la anterior
                filas[-1]['descripcion'] = (
                    f"{filas[-1]['descripcion']} {linea}")[:200]
            continue
        try:
            fecha = datetime.strptime(m.group(1), '%d/%m/%Y').date()
        except ValueError:
            continue
        monto = int(m.group(3).replace('.', ''))
        if monto == 0:
            continue
        filas.append({'fecha': fecha.isoformat(),
                      'descripcion': ' '.join(m.group(2).split())[:200],
                      'monto': monto})

    # Referencia idempotente: la misma compra que hoy está «por facturar» y
    # mañana aparece «facturada» tiene la misma fecha, glosa y monto — se
    # reconoce y no entra dos veces. El contador distingue dos compras
    # idénticas el mismo día (dos cafés seguidos, por ejemplo).
    for f in filas:
        base = f"{cuenta_clave}|{f['fecha']}|{f['descripcion']}|{f['monto']}"
        vistas[base] = vistas.get(base, 0) + 1
        f['es_pago'] = f['monto'] < 0
        f['cargo'] = 0 if f['es_pago'] else f['monto']
        f['abono'] = -f['monto'] if f['es_pago'] else 0
        f['categoria'] = ('' if f['es_pago']
                          else clasificar_compra_tarjeta(f['descripcion']))
        f['referencia'] = 'tar:' + hashlib.sha1(
            f'{base}|{vistas[base]}'.encode()).hexdigest()[:24]
    filas.sort(key=lambda f: f['fecha'])
    return filas


def parsear_tarjeta_alda(archivo, cuenta_clave):
    """Capa fina: saca el texto del PDF y delega en el núcleo puro."""
    import pdfplumber

    with pdfplumber.open(archivo) as pdf:
        texto = '\n'.join((p.extract_text() or '') for p in pdf.pages)
    filas = parsear_lineas_tarjeta(texto, cuenta_clave)
    if not filas:
        raise ValueError('No encontré movimientos en el PDF — ¿es el listado '
                         'de la tarjeta del portal Scotiabank?')
    compras = [f for f in filas if not f['es_pago']]
    pagos = [f for f in filas if f['es_pago']]
    return {
        'filas': filas,
        'n_compras': len(compras), 'n_pagos': len(pagos),
        'total_compras': sum(f['cargo'] for f in compras),
        'total_pagos': sum(f['abono'] for f in pagos),
        'fecha_inicio': filas[0]['fecha'], 'fecha_final': filas[-1]['fecha'],
    }


def registrar_filas_tarjeta(filas, cuenta_clave):
    """Escribe las compras de la tarjeta y calza sus pagos.

    Devuelve (creados, saltados, pagos_calzados, pagos_sin_calce).

    - Compra → gasto de la tarjeta, con su categoría.
    - Pago (monto negativo) → busca el cargo «PAGO TARJ.CRED.» en la cuenta
      corriente de Alda (mismo monto, ±3 días) y lo convierte en TRASPASO
      hacia la tarjeta. Sin ese cargo no se inventa nada: se informa, porque
      significa que falta cargar su cartola.
    """
    from datetime import timedelta

    from django.db import transaction

    from .models import (CategoriaFinanciera, CuentaFinanciera,
                         MovimientoFinanciero)
    from .reglas import PLAN_CUENTAS

    tarjeta = CuentaFinanciera.objects.get(clave=cuenta_clave)
    corriente = CuentaFinanciera.objects.filter(
        clave='scotiabank_alda').first()

    definiciones = dict(PLAN_CUENTAS)
    definiciones.update(CATEGORIAS_ALDA)
    cats = {}

    def _categoria(clave):
        if clave not in cats:
            nombre, clase, grupo = definiciones.get(
                clave, (clave.replace('_', ' ').title(), 'gasto', 'otros'))
            cats[clave], _ = CategoriaFinanciera.objects.get_or_create(
                clave=clave,
                defaults={'nombre': nombre, 'clase': clase, 'grupo': grupo})
        return cats[clave]

    creados = saltados = calzados = sin_calce = 0
    with transaction.atomic():
        for f in filas:
            fecha = date.fromisoformat(f['fecha'])
            monto = f['cargo'] or f['abono']
            if (fecha < COBERTURA_GASTOS_DESDE or monto <= 0
                    or MovimientoFinanciero.objects.filter(
                        referencia=f['referencia']).exists()):
                saltados += 1
                continue

            if f['es_pago']:
                pago_cta = None
                if corriente is not None:
                    pago_cta = MovimientoFinanciero.objects.filter(
                        cuenta=corriente, clase='gasto', monto=monto,
                        fecha__range=(fecha - timedelta(days=3),
                                      fecha + timedelta(days=3))
                    ).filter(descripcion__icontains='TARJ').first()
                if pago_cta is None:
                    sin_calce += 1
                    continue
                entra = MovimientoFinanciero.objects.create(
                    fecha=fecha, cuenta=tarjeta, clase='traspaso',
                    sentido='entra', monto=monto, fuente='captura',
                    referencia=f['referencia'], traspaso_par=pago_cta,
                    descripcion=f"Pago de la tarjeta: {f['descripcion']}"[:255])
                pago_cta.clase = 'traspaso'
                pago_cta.categoria = None
                pago_cta.traspaso_par = entra
                pago_cta.save(update_fields=['clase', 'categoria',
                                             'traspaso_par'])
                calzados += 1
                continue

            MovimientoFinanciero.objects.create(
                fecha=fecha, cuenta=tarjeta, clase='gasto', sentido='sale',
                monto=monto, categoria=_categoria(f['categoria']),
                fuente='captura', referencia=f['referencia'],
                descripcion=f"Tarjeta: {f['descripcion']}"[:255])
            creados += 1

    return creados, saltados, calzados, sin_calce


# ── Calce a mano de abonos con retiros (Jorge 2026-08-09) ───────────────────
# El calce automático exige monto idéntico y a veces no coincide ($499.001
# contra $500.000). Sin calce, la misma plata se cuenta dos veces: como
# retiro al salir de Aremko y como gasto al gastarse desde la cuenta puente.

def abonos_por_calzar():
    """Los abonos desde Aremko que esperan que alguien diga a qué retiro
    corresponden, del más nuevo al más viejo."""
    from .models import MovimientoFinanciero
    return (MovimientoFinanciero.objects
            .filter(categoria__clave=CLAVE_POR_CALZAR, clase='ingreso')
            .select_related('cuenta').order_by('-fecha', '-id'))


def candidatos_de_calce(abono, dias=7):
    """Retiros que podrían ser el otro lado de este abono.

    Se ofrecen los gastos de la familia cercanos en fecha, con el monto más
    parecido primero — que es como uno los reconoce a ojo.
    """
    from datetime import timedelta

    from django.db.models import Q

    from .models import MovimientoFinanciero
    from .reglas import GRUPOS_FAMILIA

    # Dos ampliaciones sobre «solo gastos del grupo familia», ambas de casos
    # reales del 20-22/08/2026:
    #
    # 1. Los «por clasificar»: un retiro que el banco glosó con el nombre a
    #    medias («TEF A AGUILERA GONZALEZ») queda justo ahí y era invisible.
    # 2. Las piernas de traspaso HUÉRFANAS (`traspaso_par` nulo): son
    #    literalmente la pareja perdida de un abono. Quedaron así por el bug de
    #    la vista previa de cartolas —que mutaba `clase` antes de escribir— y
    #    eran incalzables para siempre, porque acá solo se buscaba clase
    #    'gasto'. Tres de ellas sumaban $809.331, el descuadre exacto entre las
    #    piernas de traspaso.
    #
    # Sigue siendo Jorge quien elige la pareja: acá solo se le muestra.
    #
    # Y DOS GUARDAS, de un caso real que corrompió datos (22/08/2026): un
    # abono de $567.500 que Aremko le mandó a Alda se consumió en seis calces
    # parciales seguidos contra un consumo de restorán de SU tarjeta, un
    # retiro a Martín desde la CuentaRUT de Jorge y compras con tarjeta. Nada
    # de eso podía ser el origen, pero el filtro solo miraba monto y fecha.
    #
    # (a) El origen tiene que ser una cuenta del NEGOCIO. El abono declara en
    #     su glosa que vino de Aremko; una tarjeta o la cuenta puente de otra
    #     persona no pueden serlo.
    cercanos = (MovimientoFinanciero.objects
                .filter(sentido='sale',
                        cuenta__clave__in=CUENTAS_AREMKO,
                        fecha__range=(abono.fecha - timedelta(days=dias),
                                      abono.fecha + timedelta(days=dias)))
                .filter(Q(clase='gasto',
                          categoria__grupo__in=GRUPOS_FAMILIA)
                        | Q(clase='gasto', categoria__clave='por_clasificar')
                        | Q(clase='traspaso', traspaso_par__isnull=True))
                .exclude(cuenta=abono.cuenta)
                .select_related('cuenta', 'categoria'))

    # (b) Si la glosa del retiro dice a QUIÉN iba, tiene que ser el dueño de
    #     esta cuenta. «TEF A AGUILERA GONZALEZ» (CuentaRUT de Jorge) no puede
    #     ser el origen de un abono en la cuenta de Alda. Cuando la glosa no
    #     identifica a nadie no se descarta: ahí decide Jorge, que es el
    #     criterio de siempre.
    def _va_a_otra_persona(mov):
        destino = destino_puente(mov.descripcion, mov.cuenta.clave)
        return destino is not None and destino != abono.cuenta.clave

    candidatos = [m for m in cercanos if not _va_a_otra_persona(m)]
    return sorted(candidatos,
                  key=lambda m: (abs(int(m.monto) - int(abono.monto)),
                                 abs((m.fecha - abono.fecha).days)))


def calzar_abono_con_retiro(abono, retiro):
    """Convierte el par en un traspaso. Devuelve (común, resto_abono,
    resto_retiro).

    Si los montos no son iguales, la parte común pasa a traspaso (las dos
    piernas siguen sumando cero) y la diferencia queda VISIBLE como resto:
    un pedazo sin explicar es un dato, taparlo es un invento.
    """
    from django.db import transaction

    from .models import CategoriaFinanciera, MovimientoFinanciero

    comun = min(int(abono.monto), int(retiro.monto))
    resto_abono = int(abono.monto) - comun
    resto_retiro = int(retiro.monto) - comun

    with transaction.atomic():
        if resto_abono:
            MovimientoFinanciero.objects.create(
                fecha=abono.fecha, cuenta=abono.cuenta, clase='ingreso',
                sentido='entra', monto=resto_abono, categoria=abono.categoria,
                fuente=abono.fuente,
                referencia=f'{abono.referencia}:resto'[:120],
                descripcion=f'{abono.descripcion} (resto sin calzar)'[:255])
        if resto_retiro:
            # Un gasto SIEMPRE lleva categoría (lo exige el modelo). Si el
            # retiro venía de una pierna de traspaso huérfana, su categoría
            # puede ser nula: el resto cae en «por clasificar», que es donde
            # se ve y se resuelve, y nunca en un gasto sin categoría.
            cat_resto = retiro.categoria or CategoriaFinanciera.objects.filter(
                clave='por_clasificar').first()
            MovimientoFinanciero.objects.create(
                fecha=retiro.fecha, cuenta=retiro.cuenta, clase='gasto',
                sentido='sale', monto=resto_retiro, categoria=cat_resto,
                fuente=retiro.fuente,
                referencia=f'{retiro.referencia}:resto'[:120],
                descripcion=f'{retiro.descripcion} (resto sin calzar)'[:255])

        for mov, par in ((abono, retiro), (retiro, abono)):
            mov.monto = comun
            # La categoría se guarda ANTES de vaciarla: es lo único que hace
            # reversible el calce (ver descalzar_par). Un traspaso no lleva
            # categoría, pero perderla al convertir dejaba el deshacer manco.
            mov.categoria_previa = mov.categoria
            mov.clase = 'traspaso'
            mov.categoria = None
            mov.traspaso_par = par
            mov.save(update_fields=['monto', 'clase', 'categoria',
                                    'categoria_previa', 'traspaso_par'])

    return comun, resto_abono, resto_retiro


def pares_calzados(dias=60):
    """Los calces de RETIRO hechos, uno por par (la pierna que SALE), del más
    nuevo al más viejo. Para revisarlos y deshacer el que esté mal.

    Solo traspasos Aremko → cuenta puente, que es el dominio de esta página.
    Los barridos entre cuentas del negocio (MP → Scotiabank) y los pagos de
    tarjeta de Alda son traspasos automáticos y legítimos: NUNCA fueron un
    calce, y ofrecer «Deshacer» sobre ellos es peligroso — romper un barrido
    de $1.000.000 inventa un gasto y un ingreso que no existieron. La primera
    versión los listaba a todos (37 filas donde solo ~12 correspondían) y
    Jorge lo cazó antes de apretar nada (22/08/2026).
    """
    from datetime import timedelta

    from django.utils import timezone

    from .models import MovimientoFinanciero

    desde = timezone.localdate() - timedelta(days=dias)
    return (MovimientoFinanciero.objects
            .filter(clase='traspaso', sentido='sale',
                    traspaso_par__isnull=False, fecha__gte=desde,
                    cuenta__clave__in=CUENTAS_AREMKO,
                    traspaso_par__cuenta__clave__in=tuple(CUENTAS_PUENTE))
            .select_related('cuenta', 'traspaso_par', 'traspaso_par__cuenta',
                            'categoria_previa',
                            'traspaso_par__categoria_previa')
            .order_by('-fecha', '-id'))


def pares_sospechosos(dias=90):
    """Traspasos que ENTRAN a una cuenta puente sin salir de una cuenta de
    Aremko. Por definición no pueden ser un calce de retiro: la plata que llega
    «desde Aremko» tiene que salir de Aremko.

    Salieron de la cascada del 22/08/2026 — un abono de $567.500 consumido en
    seis calces parciales contra un consumo de restorán de la tarjeta de Alda,
    un retiro a Martín desde la CuentaRUT de Jorge y compras con tarjeta. Las
    guardas nuevas impiden que vuelva a pasar, pero los pares ya hechos siguen
    en la base y hay que poder verlos para deshacerlos: acotar la lista
    principal a lo que corresponde los dejaba escondidos.
    """
    from datetime import timedelta

    from django.utils import timezone

    from .models import MovimientoFinanciero

    desde = timezone.localdate() - timedelta(days=dias)
    return (MovimientoFinanciero.objects
            .filter(clase='traspaso', sentido='sale',
                    traspaso_par__isnull=False, fecha__gte=desde,
                    traspaso_par__cuenta__clave__in=tuple(CUENTAS_PUENTE))
            .exclude(cuenta__clave__in=CUENTAS_AREMKO)
            .select_related('cuenta', 'traspaso_par', 'traspaso_par__cuenta',
                            'categoria_previa')
            .order_by('-fecha', '-id'))


def descalzar_par(movimiento):
    """Rompe un calce y deja las DOS piernas como estaban. Devuelve la lista
    de movimientos tocados.

    Nace del 22/08/2026: se calzó por error un pago real a Previred contra un
    abono de Alda —$567.500 los dos, a nueve días— y deshacerlo exigió un
    comando escrito a mano. La herramienta sabía crear el par pero no romperlo,
    y eso vuelve peligroso el botón de calzar.

    Cada pierna vuelve a su clase por el SENTIDO, que nunca se tocó: lo que
    entra era un ingreso, lo que sale un gasto. La categoría sale de
    `categoria_previa`; si la fila es anterior a ese campo, cae en el default
    honesto de su lado (por clasificar / abono por calzar) en vez de inventar.

    NO deshace la partición de un calce parcial: los «restos» quedan como
    filas propias y se resuelven a mano — juntarlos automáticamente sería
    adivinar cuál resto era de cuál.
    """
    from django.db import transaction

    from .models import CategoriaFinanciera, MovimientoFinanciero

    if movimiento is None or movimiento.clase != 'traspaso':
        return []
    par = movimiento.traspaso_par
    piernas = [p for p in (movimiento, par) if p is not None]

    def _default(mov):
        clave = (CLAVE_POR_CALZAR if mov.sentido == 'entra'
                 else 'por_clasificar')
        return CategoriaFinanciera.objects.filter(clave=clave).first()

    with transaction.atomic():
        for mov in piernas:
            mov.clase = 'ingreso' if mov.sentido == 'entra' else 'gasto'
            mov.categoria = mov.categoria_previa or _default(mov)
            mov.categoria_previa = None
            mov.traspaso_par = None
            mov.save(update_fields=['clase', 'categoria', 'categoria_previa',
                                    'traspaso_par'])
    return piernas


def estado_fila_cartola(cuenta_clave, fila):
    """'nuevo' | 'ya_existe' | 'en_historico' | 'revisar' para una fila.

    La usan la vista previa (mostrar) Y el registro (defensa doble):
    - referencia ya escrita → ya_existe
    - Scotiabank, gasto: el histórico de julio era mes-nivel con fecha
      estimada → cualquier hist: del MISMO MES con el mismo monto lo cubre
      (conservador: mejor saltar que contar dos veces).
    - BancoEstado: hist: exacto por fecha+monto.
    - Traspaso propio (barrido que llega): si ya existe la pierna 'entra'
      con el mismo monto a ±2 días → ya_existe; si no → 'revisar' (no se
      crea una pierna suelta que rompería la suma cero).
    """
    from .models import MovimientoFinanciero

    fecha = date.fromisoformat(fila['fecha'])
    monto = fila['abono'] or fila['cargo']

    if MovimientoFinanciero.objects.filter(referencia=fila['referencia']).exists():
        return 'ya_existe'

    if fila.get('propio') and fila['clase'] == 'traspaso':
        from datetime import timedelta
        hay_par = MovimientoFinanciero.objects.filter(
            clase='traspaso', sentido='entra', cuenta__clave=cuenta_clave,
            monto=monto, fecha__range=(fecha - timedelta(days=2),
                                       fecha + timedelta(days=2))).exists()
        if hay_par:
            return 'ya_existe'
        # En una cuenta puente el abono desde Aremko SIEMPRE se escribe: con
        # retiro que calce se convierte en traspaso, y sin él queda como
        # «por calzar» (a mano). Lo que no se puede es descartarlo: la plata
        # entró y el saldo tiene que reflejarlo.
        if cuenta_clave in CUENTAS_PUENTE:
            return 'nuevo'
        return 'revisar'

    if cuenta_clave == 'scotiabank' and fila['clase'] == 'gasto':
        if MovimientoFinanciero.objects.filter(
                referencia__startswith='hist:', cuenta__clave=cuenta_clave,
                monto=monto, fecha__year=fecha.year,
                fecha__month=fecha.month).exists():
            return 'en_historico'
    else:
        if MovimientoFinanciero.objects.filter(
                referencia__startswith='hist:', cuenta__clave=cuenta_clave,
                fecha=fecha, monto=monto).exists():
            return 'en_historico'

    return 'nuevo'


# ── F5: comisiones de SumUp por la API ───────────────────────────────────────
# Estructura real verificada 2026-08-08 (sonda): /v0.1/me/financials/payouts
# devuelve un item POR TRANSACCIÓN liquidada: {"amount": neto, "fee": comisión,
# "date": "AAAA-MM-DD", "id": único, "type": "PAYOUT", "status": "SUCCESSFUL"}.
# El neto ya entra a BancoEstado vía cartola (ingreso liquidacion_sumup), así
# que acá solo se registra la COMISIÓN — en la cuenta puente "SumUp (en
# tránsito)", que no está en el flujo de caja (no mueve saldos) pero sí suma
# al gasto del mes en el tablero. Mismo patrón que las comisiones MP.


def registrar_payouts_sumup(items):
    """Núcleo puro-DB: convierte items de payouts en gastos de comisión.

    Idempotente por referencia sumup:fee:<id>; corte julio; solo PAYOUT
    SUCCESSFUL con fee > 0. Devuelve (creados, saltados).
    """
    from .models import (CategoriaFinanciera, CuentaFinanciera,
                         MovimientoFinanciero)

    try:
        cuenta = CuentaFinanciera.objects.get(clave='sumup_transito')
        cat = CategoriaFinanciera.objects.get(clave='comisiones')
    except (CuentaFinanciera.DoesNotExist, CategoriaFinanciera.DoesNotExist):
        logger.warning('finanzas sin sembrar (sumup_transito): comisiones '
                       'SumUp no registradas')
        return 0, 0

    creados = saltados = 0
    for p in items:
        try:
            if (p.get('type') != 'PAYOUT'
                    or p.get('status') != 'SUCCESSFUL'):
                saltados += 1
                continue
            fee = int(round(float(p.get('fee') or 0)))
            pid = p.get('id')
            fecha = date.fromisoformat(str(p.get('date') or ''))
            if not pid or fee <= 0 or fecha < COBERTURA_GASTOS_DESDE:
                saltados += 1
                continue
            ref = f'sumup:fee:{pid}'
            if MovimientoFinanciero.objects.filter(referencia=ref).exists():
                saltados += 1
                continue
            MovimientoFinanciero.objects.create(
                fecha=fecha, cuenta=cuenta, clase='gasto', sentido='sale',
                monto=fee, categoria=cat, fuente='api', referencia=ref,
                descripcion=(f"Comisión SumUp venta {p.get('transaction_code') or pid} "
                             f"(neto ${int(float(p.get('amount') or 0)):,})"
                             .replace(',', '.'))[:255])
            creados += 1
        except Exception:
            logger.exception('payout SumUp %s no registrado', p.get('id'))
    return creados, saltados


def traer_comisiones_sumup(dias=7):
    """Capa fina: consulta payouts a la API de SumUp y registra comisiones.

    Devuelve (creados, total_api). Lanza RuntimeError si no hay clave —
    el que llama decide si eso es un salto limpio o un error.
    """
    import os

    import requests

    clave = os.environ.get('SUMUP_API_KEY')
    if not clave:
        raise RuntimeError('SUMUP_API_KEY no configurada')

    from datetime import timedelta
    desde = (date.today() - timedelta(days=dias)).isoformat()
    r = requests.get(
        'https://api.sumup.com/v0.1/me/financials/payouts',
        headers={'Authorization': f'Bearer {clave}'},
        params={'start_date': desde, 'end_date': date.today().isoformat()},
        timeout=30,
    )
    r.raise_for_status()
    datos = r.json()
    items = datos if isinstance(datos, list) else (datos.get('items') or [])
    creados, _ = registrar_payouts_sumup(items)
    if creados:
        logger.info('finanzas: %s comisiones SumUp registradas', creados)
    return creados, len(items)
