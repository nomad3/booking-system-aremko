# -*- coding: utf-8 -*-
"""Tests de finanzas (P-22 F1).

Sin fixtures de ventas: la app es aislada y el tablero funciona con Pago vacío.
"""
from datetime import date, datetime, timedelta
from decimal import Decimal

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.db.models import Sum
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from conciliacion.models import MOTIVO_NO_ES_COBRO, MovimientoMP

from .models import (CategoriaFinanciera, CuentaFinanciera,
                     MovimientoFinanciero, SaldoMensual)
from .services import (parsear_cartola_bancoestado, parsear_transferencia_mp,
                       registrar_compras_mp, registrar_comisiones_mp,
                       registrar_transferencia_mp)


class SiembraTest(TestCase):
    def test_sembrar_es_idempotente(self):
        call_command('sembrar_finanzas')
        cuentas, cats = CuentaFinanciera.objects.count(), CategoriaFinanciera.objects.count()
        self.assertGreaterEqual(cuentas, 6)
        self.assertGreaterEqual(cats, 13)
        call_command('sembrar_finanzas')
        self.assertEqual(CuentaFinanciera.objects.count(), cuentas)
        self.assertEqual(CategoriaFinanciera.objects.count(), cats)


class CargaHistoricaTest(TestCase):
    def test_carga_idempotente_y_traspasos_cuadran(self):
        call_command('sembrar_finanzas')
        # Modo lectura no escribe nada.
        call_command('cargar_historico_finanzas')
        self.assertEqual(MovimientoFinanciero.objects.count(), 0)

        call_command('cargar_historico_finanzas', '--aplicar')
        total = MovimientoFinanciero.objects.count()
        self.assertGreater(total, 100)

        # Decisión 2026-08-08: se parte de julio, el primer mes completo y que
        # se recuerda. Nada anterior debe entrar con el default.
        self.assertEqual(
            MovimientoFinanciero.objects.filter(fecha__lt=date(2026, 7, 1)).count(), 0)

        # Idempotencia: segunda corrida no duplica.
        call_command('cargar_historico_finanzas', '--aplicar')
        self.assertEqual(MovimientoFinanciero.objects.count(), total)

        # El control central del diseño: los traspasos suman cero.
        agg = {r['sentido']: int(r['t'] or 0)
               for r in MovimientoFinanciero.objects.filter(clase='traspaso')
               .values('sentido').annotate(t=Sum('monto'))}
        self.assertEqual(agg.get('entra', 0), agg.get('sale', 0))
        self.assertGreater(agg.get('entra', 0), 0)

        # Las dos piernas quedan enlazadas y en cuentas distintas.
        una = MovimientoFinanciero.objects.filter(clase='traspaso', sentido='sale').first()
        self.assertIsNotNone(una.traspaso_par)
        self.assertNotEqual(una.cuenta_id, una.traspaso_par.cuenta_id)
        self.assertEqual(una.monto, una.traspaso_par.monto)

        # El SII quedó con fecha real y categoría impuestos.
        sii = MovimientoFinanciero.objects.get(monto=Decimal('1956770'))
        self.assertEqual(sii.categoria.clave, 'impuestos')
        self.assertEqual(sii.fecha, date(2026, 7, 20))
        self.assertFalse(sii.fecha_estimada)


class CoherenciaTest(TestCase):
    def test_un_gasto_que_entra_no_pasa(self):
        call_command('sembrar_finanzas')
        m = MovimientoFinanciero(
            fecha=date(2026, 8, 1),
            cuenta=CuentaFinanciera.objects.get(clave='efectivo'),
            clase='gasto', sentido='entra', monto=1000,
            categoria=CategoriaFinanciera.objects.get(clave='insumos'))
        with self.assertRaises(ValidationError):
            m.full_clean()

    def test_traspaso_con_categoria_no_pasa(self):
        call_command('sembrar_finanzas')
        m = MovimientoFinanciero(
            fecha=date(2026, 8, 1),
            cuenta=CuentaFinanciera.objects.get(clave='efectivo'),
            clase='traspaso', sentido='sale', monto=1000,
            categoria=CategoriaFinanciera.objects.get(clave='insumos'))
        with self.assertRaises(ValidationError):
            m.full_clean()


class TableroTest(TestCase):
    def test_solo_superusuario(self):
        url = reverse('finanzas:tablero')
        # Anónimo → redirect a login.
        self.assertEqual(self.client.get(url).status_code, 302)
        # Staff no superusuario → tampoco entra.
        User.objects.create_user('staff', password='x', is_staff=True)
        self.client.login(username='staff', password='x')
        self.assertEqual(self.client.get(url).status_code, 302)

    def test_superusuario_ve_el_tablero(self):
        call_command('sembrar_finanzas')
        call_command('cargar_historico_finanzas', '--aplicar')
        User.objects.create_superuser('duenio', 'x@x.cl', 'x')
        self.client.login(username='duenio', password='x')
        r = self.client.get(reverse('finanzas:tablero'))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Resumen mensual')
        self.assertContains(r, 'Los traspasos cuadran')

        # Regresión (visto en prod 2026-08-08): un mes SIN gastos cargados
        # muestra «—» y no "$0 + resultado" — la trampa era que sum(gastos[m])
        # sobre un defaultdict creaba la llave antes del chequeo `m in gastos`.
        por_mes = {f['mes']: f for f in r.context['resumen']}
        junio = date(2026, 6, 1)
        self.assertIn(junio, por_mes)
        self.assertEqual(por_mes[junio]['gastos'], '—')
        self.assertEqual(por_mes[junio]['resultado'], '—')
        # Julio sí tiene gastos cargados → números de verdad.
        julio = date(2026, 7, 1)
        self.assertNotEqual(por_mes[julio]['gastos'], '—')
        self.assertNotEqual(por_mes[julio]['resultado'], '—')


class ComprasMPTest(TestCase):
    """Compras vía MP (Aremko pagador) → gasto automático (P-22 F2-B)."""

    def test_registra_con_cuenta_corte_e_idempotencia(self):
        call_command('sembrar_finanzas')
        pagos = [
            # Con saldo MP → cuenta mercado_pago.
            dict(id=111, transaction_amount=25990, payment_type_id='account_money',
                 date_approved='2026-08-05T10:00:00.000-04:00',
                 description='Mercado Libre - cable HDMI'),
            # Con la Visa a través de MP → el gasto es de la Visa.
            dict(id=222, transaction_amount=49990, payment_type_id='credit_card',
                 date_approved='2026-07-15T09:00:00.000-04:00',
                 description='Retail'),
            # Antes del corte de julio → fuera.
            dict(id=333, transaction_amount=10000, payment_type_id='account_money',
                 date_approved='2026-06-20T09:00:00.000-04:00'),
            # Monto cero y sin fecha → fuera, sin explotar.
            dict(id=444, transaction_amount=0,
                 date_approved='2026-08-05T10:00:00.000-04:00'),
            dict(id=555, transaction_amount=5000, payment_type_id='account_money'),
        ]
        self.assertEqual(registrar_compras_mp(pagos), 2)

        m1 = MovimientoFinanciero.objects.get(referencia='mp:111')
        self.assertEqual((m1.cuenta.clave, m1.clase, m1.sentido, m1.fuente),
                         ('mercado_pago', 'gasto', 'sale', 'api'))
        self.assertEqual(m1.categoria.clave, 'por_clasificar')
        self.assertEqual(m1.fecha, date(2026, 8, 5))
        self.assertIn('Mercado Libre', m1.descripcion)

        m2 = MovimientoFinanciero.objects.get(referencia='mp:222')
        self.assertEqual(m2.cuenta.clave, 'visa_2936')

        # Segunda corrida: nada nuevo.
        self.assertEqual(registrar_compras_mp(pagos), 0)
        self.assertEqual(MovimientoFinanciero.objects.count(), 2)

    def test_sin_sembrar_no_explota(self):
        self.assertEqual(registrar_compras_mp([dict(
            id=1, transaction_amount=1000, payment_type_id='account_money',
            date_approved='2026-08-05T10:00:00.000-04:00')]), 0)


class ComisionesMPTest(TestCase):
    """F4 paso 2: la comisión que MP descuenta de cada cobro, como gasto."""

    def test_registra_con_corte_e_idempotencia(self):
        call_command('sembrar_finanzas')
        cobros = [
            # Cobro con comisión del lado de Aremko → gasto comisiones.
            dict(id=901, date_approved='2026-08-06T12:00:00.000-04:00',
                 description='Reserva Ana',
                 fee_details=[{'type': 'mercadopago_fee', 'amount': 1990.0,
                               'fee_payer': 'collector'}]),
            # Comisión que paga el CLIENTE → no es gasto de Aremko.
            dict(id=902, date_approved='2026-08-06T12:00:00.000-04:00',
                 fee_details=[{'amount': 500.0, 'fee_payer': 'payer'}]),
            # Transferencia simple sin comisión → nada.
            dict(id=903, date_approved='2026-08-06T12:00:00.000-04:00',
                 fee_details=[]),
            # Antes del corte de julio → fuera.
            dict(id=904, date_approved='2026-06-06T12:00:00.000-04:00',
                 fee_details=[{'amount': 1000.0, 'fee_payer': 'collector'}]),
        ]
        self.assertEqual(registrar_comisiones_mp(cobros), 1)
        m = MovimientoFinanciero.objects.get(referencia='mp:fee:901')
        self.assertEqual((m.clase, m.cuenta.clave, m.categoria.clave, int(m.monto)),
                         ('gasto', 'mercado_pago', 'comisiones', 1990))
        self.assertIn('Reserva Ana', m.descripcion)
        # Segunda corrida: no duplica.
        self.assertEqual(registrar_comisiones_mp(cobros), 0)
        self.assertEqual(MovimientoFinanciero.objects.filter(
            referencia__startswith='mp:fee:').count(), 1)


# Estructura real del correo «Tu transferencia fue enviada» (2026-08-08),
# anonimizada: mismos rótulos, mismo anidado de tags alrededor del monto.
HTML_TRANSFERENCIA = (
    '<html><body><span>Ya enviamos tu transferencia de '
    '<span style="white-space: nowrap;">$ {monto}</span></span>'
    '<h1> Nombre y apellido: <strong>{nombre}</strong><p></p> '
    'Entidad: <strong>Banco Estado</strong><p></p> '
    'N&uacute;mero de cuenta: <strong>12345678</strong></h1></body></html>'
)


class CorreosMPTest(TestCase):
    """F3: transferencias salientes de MP desde el correo."""

    def _html(self, monto='250.000', nombre='Maria Prueba'):
        return HTML_TRANSFERENCIA.format(monto=monto, nombre=nombre)

    def test_parser_extrae_los_tres_campos(self):
        d = parsear_transferencia_mp(self._html())
        self.assertEqual(d, {'monto': 250000, 'beneficiario': 'Maria Prueba',
                             'entidad': 'Banco Estado'})
        # Correo que no calza → None, no un registro inventado.
        self.assertIsNone(parsear_transferencia_mp('<html>Recibiste un pago</html>'))

    def test_registro_clasifica_e_idempotente(self):
        call_command('sembrar_finanzas')
        f = date(2026, 8, 7)

        # Nombre desconocido → por clasificar (plan de cuentas 2026-08-08:
        # ya no se asume que toda transferencia a persona es remuneración).
        d = parsear_transferencia_mp(self._html())
        self.assertEqual(registrar_transferencia_mp(d, f, 'correo:mp:t1'), ('creado', 1))
        m = MovimientoFinanciero.objects.get(referencia='correo:mp:t1')
        self.assertEqual((m.clase, m.fuente, m.cuenta.clave, m.categoria.clave),
                         ('gasto', 'correo', 'mercado_pago', 'por_clasificar'))

        # Masajista conocida por regla → honorarios.
        dm = parsear_transferencia_mp(self._html(monto='120.000', nombre='Sofia Plaza Cue'))
        registrar_transferencia_mp(dm, f, 'correo:mp:tm')
        self.assertEqual(MovimientoFinanciero.objects.get(
            referencia='correo:mp:tm').categoria.clave, 'honorarios_masajistas')
        # Mismo correo de nuevo → no duplica.
        self.assertEqual(registrar_transferencia_mp(d, f, 'correo:mp:t1'), ('ya_existe', 0))

        # Insumos Sur → insumos.
        d2 = parsear_transferencia_mp(self._html(monto='89.990', nombre='Insumos Sur'))
        registrar_transferencia_mp(d2, f, 'correo:mp:t2')
        self.assertEqual(MovimientoFinanciero.objects.get(
            referencia='correo:mp:t2').categoria.clave, 'insumos')

        # Barrido a cuenta propia → traspaso con dos piernas enlazadas.
        d3 = parsear_transferencia_mp(self._html(monto='1.500.000',
                                                 nombre='Aremko Spa Scotiabank'))
        self.assertEqual(registrar_transferencia_mp(d3, f, 'correo:mp:t3'), ('creado', 2))
        sale = MovimientoFinanciero.objects.get(referencia='correo:mp:t3:sale')
        self.assertEqual(sale.clase, 'traspaso')
        self.assertEqual(sale.traspaso_par.cuenta.clave, 'scotiabank')
        self.assertEqual(sale.traspaso_par.traspaso_par_id, sale.id)

    def test_no_duplica_lo_que_cargo_el_historico(self):
        call_command('sembrar_finanzas')
        f = date(2026, 8, 2)
        MovimientoFinanciero.objects.create(
            fecha=f, cuenta=CuentaFinanciera.objects.get(clave='mercado_pago'),
            clase='gasto', sentido='sale', monto=40000,
            categoria=CategoriaFinanciera.objects.get(clave='remuneraciones'),
            fuente='correo', referencia='hist:mp:77')
        d = parsear_transferencia_mp(self._html(monto='40.000', nombre='Javiera Perez'))
        self.assertEqual(registrar_transferencia_mp(d, f, 'correo:mp:t9'),
                         ('en_historico', 0))


class VerificacionMPTest(TestCase):
    """La sección sistema-vs-API del tablero (P-22 F2-A)."""

    def test_lado_mp_cuenta_lo_correcto(self):
        # La tabla agrega POR DÍA: cada caso va en su propio día para poder
        # afirmar su monto, y el ajeno comparte día con v1 — si se colara,
        # el día de v1 mostraría la suma y no $987.654.
        ahora = timezone.now()
        # Cobro normal → cuenta.
        MovimientoMP.objects.create(mp_payment_id='v1', fecha=ahora, monto=987654)
        # Ignorado POR DEBORAH (irrelevante para su tarea, pero plata que entró) → cuenta.
        MovimientoMP.objects.create(mp_payment_id='v2', monto=345678,
                                    fecha=ahora - timedelta(days=1),
                                    estado='ignorado')
        # Ajeno confirmado por la API (Aremko pagador) → NO cuenta.
        MovimientoMP.objects.create(mp_payment_id='v3', fecha=ahora, monto=111222,
                                    estado='ignorado',
                                    sugerencia_motivo=MOTIVO_NO_ES_COBRO)

        User.objects.create_superuser('duenio', 'x@x.cl', 'x')
        self.client.login(username='duenio', password='x')
        r = self.client.get(reverse('finanzas:tablero'))
        self.assertContains(r, 'Verificación Mercado Pago')
        self.assertContains(r, '$987.654')
        self.assertContains(r, '$345.678')
        self.assertNotContains(r, '$111.222')
        # Sin pagos en Django, la diferencia de la ventana es todo el lado MP.
        self.assertContains(r, '+$1.333.332')


def _xlsx_cartola():
    """Un export de BancoEstado en memoria con la estructura real: Resumen con
    rótulos en col A y valores en col E; Movimientos con fechas DD/MM (sin
    año), cargos como enteros crudos y abonos/saldos como strings '$1.234'.
    Incluye dos transferencias GEMELAS el mismo día (el saldo encadenado las
    distingue) y el cruce julio→agosto (cierre de julio derivable)."""
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    res = wb.active
    res.title = 'Resumen'
    for i, (k, v) in enumerate([
            ('Fecha Inicio', '14/07/2026'), ('Fecha Final', '04/08/2026'),
            ('Saldo Inicial', '$100.000'), ('Saldo Final', '$100.000'),
            ('N° Cuenta', '82370351925')], 1):
        res.cell(row=i, column=1, value=k)
        res.cell(row=i, column=5, value=v)
    mov = wb.create_sheet('Movimientos')
    mov.append(['Fecha', 'Sucursal', 'N° Cuenta', 'Alias', 'N° Cartola',
                'N° Operación', 'Descripción', 'Cheques / Cargos',
                'Depósitos / Abonos', 'Saldo'])
    mov.append(['15/07', 'STGO', '82370351925', 'CHEQ', 9, '0001077',
                'TEF DE SUMUP CHILE PAYMENTS S A', 0, '$50.000', '$150.000'])
    mov.append(['20/07', 'STGO', '82370351925', 'CHEQ', 9, '7023294',
                'TEF A PEREZ SOTO MARIA', 30000, '$0', '$120.000'])
    mov.append(['20/07', 'STGO', '82370351925', 'CHEQ', 9, '7023295',
                'TEF A PEREZ SOTO MARIA', 30000, '$0', '$90.000'])
    mov.append(['02/08', 'STGO', '82370351925', 'CHEQ', 9, '0001077',
                'TEF DE FLOW S A', 0, '$10.000', '$100.000'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class CartolaBancoEstadoTest(TestCase):
    """F4 paso 3: parser del export XLSX y página de carga con confirmación."""

    def test_parser_cuadra_clasifica_y_deriva_cierre_julio(self):
        r = parsear_cartola_bancoestado(_xlsx_cartola())
        self.assertTrue(r['cuadra'])
        self.assertEqual(len(r['filas']), 4)
        self.assertEqual(r['cierres_mes'], {'2026-07': 90000})
        # El año se asigna bien cruzando julio→agosto.
        self.assertEqual(r['filas'][-1]['fecha'], '2026-08-02')
        # Las gemelas (mismo día, mismo monto) tienen referencias distintas.
        refs = [f['referencia'] for f in r['filas']]
        self.assertEqual(len(refs), len(set(refs)))
        # Clasificación por descripción.
        self.assertEqual(r['filas'][0]['categoria'], 'liquidacion_sumup')
        self.assertEqual(r['filas'][1]['categoria'], 'por_clasificar')
        self.assertEqual(r['filas'][3]['categoria'], 'liquidacion_flow')

    def test_pagina_propone_confirma_y_no_duplica(self):
        call_command('sembrar_finanzas')
        User.objects.create_superuser('duenio', 'x@x.cl', 'x')
        self.client.login(username='duenio', password='x')
        url = reverse('finanzas:cargar_cartola')

        self.assertContains(self.client.get(url), 'Subir el export')

        # Propuesta: nada se escribe todavía.
        r = self.client.post(url, {'archivo': _xlsx_cartola()})
        self.assertContains(r, 'Confirmar y escribir 4')
        self.assertEqual(MovimientoFinanciero.objects.count(), 0)

        # Confirmar: se escriben los 4 + el cierre de julio.
        r2 = self.client.post(url, {'confirmar': '1',
                                    'payload': r.context['payload']})
        self.assertContains(r2, 'movimientos nuevos escritos')
        en_be = MovimientoFinanciero.objects.filter(cuenta__clave='bancoestado')
        self.assertEqual(en_be.count(), 4)
        self.assertEqual(en_be.filter(clase='ingreso').count(), 2)
        self.assertEqual(en_be.filter(fuente='captura').count(), 4)
        cierre = SaldoMensual.objects.get(cuenta__clave='bancoestado',
                                          periodo=date(2026, 7, 1))
        self.assertEqual((int(cierre.saldo_cierre), cierre.fuente),
                         (90000, 'cartola'))
        # Las categorías de ingreso se crearon solas.
        self.assertTrue(CategoriaFinanciera.objects.filter(
            clave='liquidacion_sumup', clase='ingreso').exists())

        # Re-subir el mismo archivo: todo «ya está», nada que confirmar.
        r3 = self.client.post(url, {'archivo': _xlsx_cartola()})
        self.assertContains(r3, 'Nada nuevo que escribir')
        # Y re-confirmar el payload viejo tampoco duplica.
        self.client.post(url, {'confirmar': '1', 'payload': r.context['payload']})
        self.assertEqual(en_be.count(), 4)

    def test_solo_superusuario(self):
        self.assertEqual(self.client.get(
            reverse('finanzas:cargar_cartola')).status_code, 302)


class CartolaScotiabankTest(TestCase):
    """F4 paso 3b: núcleo puro del parser Scotiabank + estados anti-doble-conteo."""

    # Estructura real de typeDesc.xls: metadatos rótulo/valor, encabezado, y
    # filas DESCENDENTES con cargos NEGATIVOS. Cruza julio→agosto.
    FILAS = [
        ['Nombre Empresa', 'AREMKO HOTEL SPA', '', '', '', '', ''],
        ['Saldo Disponible', 700000.0, '', '', '', '', ''],
        ['Fecha', 'Descripción', 'Sucursal', 'N° Doc.', 'Cargos', 'Abonos', 'Saldo'],
        ['02-08-2026', 'PAC SEGUROS GENERALES  51927', 'PM', 1.0, -100000.0, '', 700000.0],
        ['02-08-2026', 'REDCOMPRA COVEPA SPA', 'PM', 2.0, -50000.0, '', 800000.0],
        ['20-07-2026', 'TEF 76485192-7 AREMKO HOTEL SP', 'NM', 3.0, '', 500000.0, 850000.0],
        ['20-07-2026', 'TEF 12343982-1 Nancy Mansilla', 'PM', 4.0, -150000.0, '', 350000.0],
    ]

    def _parsear(self):
        from finanzas.services import parsear_filas_scotiabank
        return parsear_filas_scotiabank([list(f) for f in self.FILAS])

    def test_nucleo_invierte_encadena_y_clasifica(self):
        r = self._parsear()
        self.assertTrue(r['cuadra'])
        self.assertEqual(r['cadena_rota'], 0)
        # Invertido: la más vieja primero; saldo inicial derivado de ella.
        self.assertEqual(r['filas'][0]['fecha'], '2026-07-20')
        self.assertEqual(r['saldo_inicial'], 500000)   # 350.000 + 150.000
        self.assertEqual(r['saldo_final_calculado'], 700000)
        # Cierre de julio derivado (el archivo sigue en agosto).
        self.assertEqual(r['cierres_mes'], {'2026-07': 850000})
        # Clasificación: barrido propio → traspaso; TEF persona → gasto;
        # PAC SEGUROS → seguros; REDCOMPRA → por clasificar.
        por_desc = {f['descripcion']: f for f in r['filas']}
        self.assertEqual(por_desc['TEF 76485192-7 AREMKO HOTEL SP']['clase'], 'traspaso')
        self.assertTrue(por_desc['TEF 76485192-7 AREMKO HOTEL SP']['propio'])
        self.assertEqual(por_desc['TEF 12343982-1 Nancy Mansilla']['categoria'],
                         'por_clasificar')
        self.assertEqual(por_desc['PAC SEGUROS GENERALES  51927']['categoria'], 'seguros')

    def test_estados_evitan_dobles_conteos(self):
        from finanzas.services import estado_fila_cartola, registrar_filas_cartola
        call_command('sembrar_finanzas')
        r = self._parsear()
        por_desc = {f['descripcion']: f for f in r['filas']}

        # Barrido propio SIN traspaso registrado → revisar (no se crea suelto).
        barrido = por_desc['TEF 76485192-7 AREMKO HOTEL SP']
        self.assertEqual(estado_fila_cartola('scotiabank', barrido), 'revisar')

        # Con el traspaso registrado (fecha ±2 días) → ya_existe.
        mp = CuentaFinanciera.objects.get(clave='mercado_pago')
        sc = CuentaFinanciera.objects.get(clave='scotiabank')
        sale = MovimientoFinanciero.objects.create(
            fecha=date(2026, 7, 19), cuenta=mp, clase='traspaso', sentido='sale',
            monto=500000, fuente='correo', referencia='t:s')
        MovimientoFinanciero.objects.create(
            fecha=date(2026, 7, 19), cuenta=sc, clase='traspaso', sentido='entra',
            monto=500000, fuente='correo', referencia='t:e', traspaso_par=sale)
        self.assertEqual(estado_fila_cartola('scotiabank', barrido), 'ya_existe')

        # Gasto que el histórico mes-nivel ya cubre (mismo MES + monto) → en_historico.
        MovimientoFinanciero.objects.create(
            fecha=date(2026, 7, 1), cuenta=sc, clase='gasto', sentido='sale',
            monto=150000, fecha_estimada=True, fuente='correo',
            categoria=CategoriaFinanciera.objects.get(clave='por_clasificar'),
            referencia='hist:scotia:99')
        nancy = por_desc['TEF 12343982-1 Nancy Mansilla']
        self.assertEqual(estado_fila_cartola('scotiabank', nancy), 'en_historico')

        # El registro respeta los estados: solo entran los 2 gastos de agosto.
        creados, saltados = registrar_filas_cartola(
            r['filas'], r['cierres_mes'], cuenta_clave='scotiabank')
        self.assertEqual((creados, saltados), (2, 2))
        self.assertEqual(MovimientoFinanciero.objects.filter(
            fuente='captura', cuenta=sc).count(), 2)
        self.assertEqual(int(SaldoMensual.objects.get(
            cuenta=sc, periodo=date(2026, 7, 1)).saldo_cierre), 850000)


class FlujoCajaTest(TestCase):
    """F4 paso 4: saldo = ancla de julio + movimientos, día a día."""

    def test_ancla_mas_movimientos_da_el_saldo(self):
        call_command('sembrar_finanzas')
        be = CuentaFinanciera.objects.get(clave='bancoestado')
        sc = CuentaFinanciera.objects.get(clave='scotiabank')
        SaldoMensual.objects.create(cuenta=be, periodo=date(2026, 7, 1),
                                    saldo_cierre=1000000, fuente='cartola')
        SaldoMensual.objects.create(cuenta=sc, periodo=date(2026, 7, 1),
                                    saldo_cierre=500000, fuente='cartola')
        oi = CategoriaFinanciera.objects.get(clave='otros_ingresos')
        pc = CategoriaFinanciera.objects.get(clave='por_clasificar')
        M = MovimientoFinanciero
        M.objects.create(fecha=date(2026, 8, 2), cuenta=be, clase='ingreso',
                         sentido='entra', monto=100000, categoria=oi,
                         fuente='captura', referencia='fc:1')
        M.objects.create(fecha=date(2026, 8, 3), cuenta=sc, clase='gasto',
                         sentido='sale', monto=50000, categoria=pc,
                         fuente='captura', referencia='fc:2')
        # Traspaso BE → Scotia: mueve saldos, pero NO es entrada ni salida.
        M.objects.create(fecha=date(2026, 8, 4), cuenta=be, clase='traspaso',
                         sentido='sale', monto=200000, fuente='manual',
                         referencia='fc:3')
        M.objects.create(fecha=date(2026, 8, 4), cuenta=sc, clase='traspaso',
                         sentido='entra', monto=200000, fuente='manual',
                         referencia='fc:4')
        # Cobro MP por API (cuenta sin ancla: cuenta como entrada del día,
        # pero su saldo muestra «—»).
        MovimientoMP.objects.create(
            mp_payment_id='fc1', monto=30000,
            fecha=timezone.make_aware(datetime(2026, 8, 5, 12, 0)))

        User.objects.create_superuser('duenio', 'x@x.cl', 'x')
        self.client.login(username='duenio', password='x')
        r = self.client.get(reverse('finanzas:flujo_caja'))
        self.assertEqual(r.status_code, 200)

        # La primera fila es HOY (orden descendente pedido por Jorge).
        hoy_fila = r.context['filas'][0]
        self.assertEqual(hoy_fila['dia'], date.today())
        # Orden: MP, BancoEstado, Scotiabank, Efectivo + las dos puente
        # (sin ancla todavía → «—» y fuera de los totales).
        self.assertEqual(hoy_fila['saldos'],
                         ['—', '$900.000', '$650.000', '—', '—', '—'])
        self.assertEqual(hoy_fila['total_caja'], '$1.550.000')
        self.assertEqual(hoy_fila['total_puente'], '—')
        self.assertEqual(hoy_fila['total'], '$1.550.000')

        por_dia = {f['dia']: f for f in r.context['filas']}
        self.assertEqual(por_dia[date(2026, 8, 2)]['entradas'], '$100.000')
        self.assertEqual(por_dia[date(2026, 8, 4)]['entradas'], '')
        self.assertEqual(por_dia[date(2026, 8, 4)]['salidas'], '')
        self.assertEqual(por_dia[date(2026, 8, 5)]['entradas'], '$30.000')
        self.assertIn('Mercado Pago', ' '.join(r.context['sin_ancla']))

        # Detalle por día: cada fila trae sus movimientos uno a uno.
        self.assertEqual(por_dia[date(2026, 8, 2)]['n_movs'], 1)
        self.assertEqual(por_dia[date(2026, 8, 2)]['movs'][0]['sentido'], 'entra')
        # El día del traspaso muestra sus dos piernas.
        self.assertEqual(por_dia[date(2026, 8, 4)]['n_movs'], 2)
        self.assertEqual(por_dia[date(2026, 8, 4)]['movs'][0]['extra'], 'traspaso')
        # El cobro MP aparece con su glosa/etiqueta y cuenta.
        mp_mov = por_dia[date(2026, 8, 5)]['movs'][0]
        self.assertIn('Cobro MP', mp_mov['desc'])
        self.assertEqual(mp_mov['cuenta'], 'Mercado Pago')

    def test_traspaso_a_cuenta_puente_no_hace_desaparecer_la_plata(self):
        """El caso que pilló Jorge (2026-08-09): sacar plata del negocio a la
        CuentaRUT baja Caja Aremko pero NO el total disponible — la plata
        sigue existiendo, solo cambió de bolsillo."""
        call_command('sembrar_finanzas')
        sc = CuentaFinanciera.objects.get(clave='scotiabank')
        rut = CuentaFinanciera.objects.get(clave='cuentarut_jorge')
        SaldoMensual.objects.create(cuenta=sc, periodo=date(2026, 7, 1),
                                    saldo_cierre=3000000, fuente='cartola')
        SaldoMensual.objects.create(cuenta=rut, periodo=date(2026, 7, 1),
                                    saldo_cierre=344977, fuente='cartola')
        M = MovimientoFinanciero
        sale = M.objects.create(fecha=date(2026, 8, 2), cuenta=sc,
                                clase='traspaso', sentido='sale',
                                monto=1000000, fuente='manual',
                                referencia='fp:1')
        M.objects.create(fecha=date(2026, 8, 2), cuenta=rut, clase='traspaso',
                         sentido='entra', monto=1000000, fuente='manual',
                         referencia='fp:2', traspaso_par=sale)

        User.objects.create_superuser('duenio', 'x@x.cl', 'x')
        self.client.login(username='duenio', password='x')
        r = self.client.get(reverse('finanzas:flujo_caja'))
        hoy_fila = r.context['filas'][0]
        self.assertEqual(hoy_fila['total_caja'], '$2.000.000')
        self.assertEqual(hoy_fila['total_puente'], '$1.344.977')
        self.assertEqual(hoy_fila['total'], '$3.344.977')
        # Y el día del traspaso no muestra salida: no se gastó nada.
        por_dia = {f['dia']: f for f in r.context['filas']}
        self.assertEqual(por_dia[date(2026, 8, 2)]['salidas'], '')

    def test_gasto_desde_la_cuenta_puente_si_baja_el_total(self):
        """Contrapartida: cuando esa plata se gasta de verdad, el total baja."""
        call_command('sembrar_finanzas')
        call_command('aplicar_plan_cuentas', '--aplicar')
        rut = CuentaFinanciera.objects.get(clave='cuentarut_jorge')
        SaldoMensual.objects.create(cuenta=rut, periodo=date(2026, 7, 1),
                                    saldo_cierre=344977, fuente='cartola')
        MovimientoFinanciero.objects.create(
            fecha=date(2026, 8, 3), cuenta=rut, clase='gasto', sentido='sale',
            monto=119443, fuente='captura', referencia='fp:3',
            categoria=CategoriaFinanciera.objects.get(clave='infraestructura'))

        User.objects.create_superuser('duenio', 'x@x.cl', 'x')
        self.client.login(username='duenio', password='x')
        r = self.client.get(reverse('finanzas:flujo_caja'))
        hoy_fila = r.context['filas'][0]
        self.assertEqual(hoy_fila['total_puente'], '$225.534')
        por_dia = {f['dia']: f for f in r.context['filas']}
        self.assertEqual(por_dia[date(2026, 8, 3)]['salidas'], '$119.443')

    def test_solo_superusuario(self):
        self.assertEqual(self.client.get(
            reverse('finanzas:flujo_caja')).status_code, 302)


class CartolaScotiabankMensualTest(TestCase):
    """La segunda variante real del portal: estado de cuenta mensual —
    6 columnas, orden ASCENDENTE, Saldo Anterior/Actual en la cabecera."""

    FILAS = [
        ['Nombre Empresa', 'AREMKO HOTEL SPA', '', '', '', ''],
        ['Numero Cuenta', 973080644.0, '', '', '', ''],
        ['Fecha Desde', '01-07-2026', '', '', '', ''],
        ['Fecha Hasta', '31-07-2026', '', '', '', ''],
        ['Saldo Anterior', 882699.0, '', '', '', ''],
        ['Saldo Actual', 490434.0, '', '', '', ''],
        ['Fecha', 'Descripción', 'Numero Documento', 'Cargo', 'Abono', 'Saldo Diario'],
        ['01-07-2026', 'COMISION MANTENCION PLAN', 0.0, -40823.0, '', 841876.0],
        ['03-07-2026', 'TEF 18883207-5 PAULA ANDREA IB', 0.0, '', 155000.0, 996876.0],
        ['31-07-2026', 'REDCOMPRA COVEPA SPA', 0.0, -506442.0, '', 490434.0],
    ]

    def test_variante_mensual_cuadra_y_ancla_el_cierre(self):
        from finanzas.services import parsear_filas_scotiabank
        r = parsear_filas_scotiabank([list(f) for f in self.FILAS])
        self.assertTrue(r['cuadra'])
        self.assertEqual(r['cadena_rota'], 0)
        # NO se invierte (ya viene cronológico) y el inicio sale de la cabecera.
        self.assertEqual(r['filas'][0]['fecha'], '2026-07-01')
        self.assertEqual(r['saldo_inicial'], 882699)
        self.assertEqual(r['saldo_final_calculado'], 490434)
        # Mes cerrado (Fecha Hasta = 31-07) → ancla el cierre aunque no haya
        # filas de agosto.
        self.assertEqual(r['cierres_mes'], {'2026-07': 490434})
        self.assertEqual(r['cuenta_numero'], '973080644')
        # El abono de una clienta cuenta como ingreso (transferencia recibida).
        self.assertEqual(r['filas'][1]['categoria'], 'transferencias_recibidas')


class ComisionesSumUpTest(TestCase):
    """F5: la comisión de cada payout SumUp como gasto en la cuenta puente."""

    def test_registra_solo_payouts_exitosos_con_fee(self):
        from finanzas.services import registrar_payouts_sumup
        call_command('sembrar_finanzas')
        items = [
            # Payout real con comisión (estructura de la sonda 2026-08-08).
            {'id': 111, 'type': 'PAYOUT', 'status': 'SUCCESSFUL', 'fee': 1998.0,
             'amount': 58002.0, 'date': '2026-07-27', 'transaction_code': 'TAAA1'},
            # Sin comisión → nada.
            {'id': 222, 'type': 'PAYOUT', 'status': 'SUCCESSFUL', 'fee': 0,
             'amount': 10000.0, 'date': '2026-08-01'},
            # No es payout → fuera.
            {'id': 333, 'type': 'CHARGE_BACK', 'status': 'SUCCESSFUL',
             'fee': 500.0, 'date': '2026-08-01'},
            # Antes del corte de julio → fuera.
            {'id': 444, 'type': 'PAYOUT', 'status': 'SUCCESSFUL', 'fee': 900.0,
             'amount': 5000.0, 'date': '2026-06-15'},
        ]
        self.assertEqual(registrar_payouts_sumup(items), (1, 3))
        m = MovimientoFinanciero.objects.get(referencia='sumup:fee:111')
        self.assertEqual(
            (m.cuenta.clave, m.categoria.clave, int(m.monto), str(m.fecha)),
            ('sumup_transito', 'comisiones', 1998, '2026-07-27'))
        # La cuenta puente NO participa del flujo de caja.
        from finanzas.views import CUENTAS_FLUJO
        self.assertNotIn('sumup_transito', CUENTAS_FLUJO)
        # Idempotente.
        self.assertEqual(registrar_payouts_sumup(items), (0, 4))


class PlanCuentasTest(TestCase):
    """El plan de cuentas de Jorge (2026-08-08): grupos, reglas y reclasificación."""

    def test_reglas_puras(self):
        from finanzas.reglas import clasificar_por_reglas
        self.assertEqual(clasificar_por_reglas('TEF A VIDAL PANTOJA CAROLINA ANDREA'),
                         'honorarios_masajistas')
        self.assertEqual(clasificar_por_reglas('Cintia Brinzo'), 'personales_alda')
        self.assertEqual(clasificar_por_reglas('TEF 7604892-4 jorge aguilera'),
                         'personales_jorge')
        self.assertEqual(clasificar_por_reglas('TEF A AGUILERA GONZALEZ CRISTIAN AN'),
                         'infraestructura')
        self.assertEqual(clasificar_por_reglas('PAGO CRELL PUERTO VARAS'),
                         'energia_electrica')
        self.assertEqual(clasificar_por_reglas('Rafael Perez Quintero'),
                         'remuneraciones')
        self.assertIsNone(clasificar_por_reglas('REDCOMPRA COMERCIO CUALQUIERA'))

    def test_comando_sincroniza_y_reclasifica(self):
        call_command('sembrar_finanzas')
        call_command('cargar_historico_finanzas', '--aplicar')

        # Lectura: no crea ni cambia nada.
        call_command('aplicar_plan_cuentas')
        self.assertFalse(CategoriaFinanciera.objects.filter(
            clave='honorarios_masajistas').exists())

        call_command('aplicar_plan_cuentas', '--aplicar')
        self.assertEqual(CategoriaFinanciera.objects.get(
            clave='honorarios_masajistas').grupo, 'masajistas')
        self.assertEqual(CategoriaFinanciera.objects.get(
            clave='remuneraciones').grupo, 'personal')

        # Del histórico: Angélica/Alda → personales_alda; Martín → personales_martin.
        angelica = MovimientoFinanciero.objects.filter(
            descripcion__icontains='Angelica Toloza').first()
        self.assertEqual(angelica.categoria.clave, 'personales_alda')
        martin = MovimientoFinanciero.objects.filter(
            descripcion__icontains='Martin Aguilera').first()
        self.assertEqual(martin.categoria.clave, 'personales_martin')
        # Rafael era y sigue siendo remuneraciones (decisión de Jorge).
        rafael = MovimientoFinanciero.objects.filter(
            descripcion__icontains='Rafael Perez').first()
        self.assertEqual(rafael.categoria.clave, 'remuneraciones')

        # Idempotente: segunda corrida no cambia nada más.
        antes = list(MovimientoFinanciero.objects.values_list('id', 'categoria__clave'))
        call_command('aplicar_plan_cuentas', '--aplicar')
        self.assertEqual(antes, list(
            MovimientoFinanciero.objects.values_list('id', 'categoria__clave')))

    def test_tablero_separa_negocio_de_retiros(self):
        call_command('sembrar_finanzas')
        call_command('cargar_historico_finanzas', '--aplicar')
        call_command('aplicar_plan_cuentas', '--aplicar')
        User.objects.create_superuser('duenio', 'x@x.cl', 'x')
        self.client.login(username='duenio', password='x')
        r = self.client.get(reverse('finanzas:tablero'))
        self.assertContains(r, 'Resultado operacional')
        self.assertContains(r, 'Retiros familia')
        por_mes = {f['mes']: f for f in r.context['resumen']}
        julio = por_mes[date(2026, 7, 1)]
        # Julio tiene retiros de la familia (Alda/Jorge del histórico) > 0.
        self.assertNotEqual(julio['retiros'], '')
        self.assertNotEqual(julio['gastos'], '—')
        # La tabla agrupada trae subtotales de grupo.
        grupos = [f['nombre'] for f in r.context['tabla_gastos'] if f['es_grupo']]
        self.assertIn('Personales Alda', grupos)
        self.assertIn('Sueldos de personal', grupos)


class DevolucionesTest(TestCase):
    """Devoluciones = contra-ingreso: restan de ingresos, no suman a gastos."""

    def test_resumen_resta_de_ingresos(self):
        call_command('sembrar_finanzas')
        call_command('aplicar_plan_cuentas', '--aplicar')
        dev = CategoriaFinanciera.objects.get(clave='devoluciones')
        self.assertEqual(dev.grupo, 'devoluciones')
        MovimientoFinanciero.objects.create(
            fecha=date(2026, 8, 6),
            cuenta=CuentaFinanciera.objects.get(clave='bancoestado'),
            clase='gasto', sentido='sale', monto=50000, categoria=dev,
            fuente='manual', referencia='dev:test:1',
            descripcion='Devolución reserva anulada')

        User.objects.create_superuser('duenio', 'x@x.cl', 'x')
        self.client.login(username='duenio', password='x')
        r = self.client.get(reverse('finanzas:tablero'))
        por_mes = {f['mes']: f for f in r.context['resumen']}
        agosto = por_mes[date(2026, 8, 1)]
        # La devolución aparece en su columna, NO en gastos del negocio,
        # y los ingresos netos la restan (sin Pago en el test: 0 − 50.000).
        self.assertEqual(agosto['devoluciones'], '$50.000')
        self.assertEqual(agosto['gastos'], '$0')
        self.assertEqual(agosto['ingresos'], '$-50.000')
        self.assertEqual(agosto['resultado'], '$-50.000')
        # En el flujo de caja SÍ es plata que sale (día con salida real).
        rf = self.client.get(reverse('finanzas:flujo_caja'))
        por_dia = {f['dia']: f for f in rf.context['filas']}
        self.assertEqual(por_dia[date(2026, 8, 6)]['salidas'], '$50.000')


class AccesoColaboradorTest(TestCase):
    """El grupo «Finanzas colaborador» (Alda): ve las 3 páginas y solo edita
    la categoría de movimientos; el staff común sigue sin ver nada."""

    def setUp(self):
        call_command('sembrar_finanzas')
        from django.contrib.auth.models import Group
        self.grupo = Group.objects.create(name='Finanzas colaborador')
        self.alda = User.objects.create_user('alda', password='x', is_staff=True,
                                             first_name='Alda')
        self.alda.groups.add(self.grupo)

    def test_colaboradora_ve_las_tres_paginas(self):
        self.client.login(username='alda', password='x')
        for nombre in ('finanzas:tablero', 'finanzas:flujo_caja',
                       'finanzas:cargar_cartola'):
            self.assertEqual(self.client.get(reverse(nombre)).status_code, 200,
                             nombre)

    def test_staff_comun_sigue_afuera(self):
        User.objects.create_user('deborah', password='x', is_staff=True)
        self.client.login(username='deborah', password='x')
        for nombre in ('finanzas:tablero', 'finanzas:flujo_caja',
                       'finanzas:cargar_cartola'):
            self.assertEqual(self.client.get(reverse(nombre)).status_code, 302,
                             nombre)

    def test_admin_movimientos_solo_categoria_editable(self):
        from django.contrib import admin as dj_admin

        from .admin import MovimientoFinancieroAdmin
        from .models import MovimientoFinanciero

        ma = MovimientoFinancieroAdmin(MovimientoFinanciero, dj_admin.site)

        class Req:
            pass
        r = Req()
        r.user = self.alda
        self.assertTrue(ma.has_view_permission(r))
        self.assertTrue(ma.has_change_permission(r))
        self.assertFalse(ma.has_add_permission(r))
        self.assertFalse(ma.has_delete_permission(r))
        readonly = ma.get_readonly_fields(r)
        self.assertIn('monto', readonly)
        self.assertIn('cuenta', readonly)
        self.assertNotIn('categoria', readonly)

    def test_comando_configura_al_usuario_existente(self):
        from django.contrib.auth.models import Group
        Group.objects.filter(name='Finanzas colaborador').delete()
        self.alda.groups.clear()
        call_command('configurar_acceso_alda')
        self.alda.refresh_from_db()
        self.assertTrue(self.alda.groups.filter(
            name='Finanzas colaborador').exists())
        # La contraseña sigue funcionando (no se tocó).
        self.assertTrue(self.client.login(username='alda', password='x'))


class ReportesGastosTest(TestCase):
    """Reportes de Jorge 2026-08-09: plan de cuentas x cuenta financiera
    (mes elegido) y plan de cuentas x meses del año."""

    def setUp(self):
        call_command('sembrar_finanzas')
        from django.contrib.auth.models import Group
        grupo = Group.objects.create(name='Finanzas colaborador')
        self.alda = User.objects.create_user('alda', password='x', is_staff=True)
        self.alda.groups.add(grupo)

        from .models import (CategoriaFinanciera, CuentaFinanciera,
                             MovimientoFinanciero)
        be = CuentaFinanciera.objects.get(clave='bancoestado')
        mp = CuentaFinanciera.objects.get(clave='mercado_pago')
        luz = CategoriaFinanciera.objects.create(
            clave='test-luz', nombre='Luz Crell', clase='gasto', grupo='energia')
        sueldo = CategoriaFinanciera.objects.create(
            clave='test-sueldo', nombre='Sueldo Nancy', clase='gasto',
            grupo='personal')
        crear = MovimientoFinanciero.objects.create
        crear(fecha=date(2026, 7, 15), cuenta=be, clase='gasto', sentido='sale',
              monto=111000, descripcion='luz julio', categoria=luz,
              fuente='manual', referencia='manual:test:1')
        crear(fecha=date(2026, 8, 5), cuenta=be, clase='gasto', sentido='sale',
              monto=222000, descripcion='luz agosto', categoria=luz,
              fuente='manual', referencia='manual:test:2')
        crear(fecha=date(2026, 8, 10), cuenta=mp, clase='gasto', sentido='sale',
              monto=55000, descripcion='sueldo agosto', categoria=sueldo,
              fuente='manual', referencia='manual:test:3')
        # Un traspaso NO debe aparecer en ninguno de los dos reportes.
        crear(fecha=date(2026, 8, 12), cuenta=be, clase='traspaso',
              sentido='sale', monto=999999, descripcion='barrido',
              fuente='manual', referencia='manual:test:4')
        self.be, self.mp = be, mp

    def test_gastos_mes_cruza_plan_con_cuentas(self):
        self.client.login(username='alda', password='x')
        r = self.client.get(reverse('finanzas:gastos_mes'),
                            {'ano': 2026, 'mes': 8})
        self.assertEqual(r.status_code, 200)
        # Columnas: solo cuentas con gastos del mes, la mayor primero, con
        # nombre corto para que la tabla quepa en pantalla.
        self.assertEqual([c['largo'] for c in r.context['cuentas']],
                         [self.be.nombre, self.mp.nombre])
        self.assertEqual(r.context['cuentas'][0]['corto'],
                         'BancoEstado Chequera')
        # Grupo con una sola categoría → se muestra la fila del grupo
        # (mismo criterio del tablero).
        self.assertContains(r, 'Energía eléctrica')
        self.assertContains(r, 'Sueldos de personal')
        self.assertContains(r, '$222.000')
        self.assertContains(r, '$55.000')
        self.assertContains(r, '$277.000')          # total general del mes
        self.assertNotContains(r, '$111.000')       # julio queda fuera
        self.assertNotContains(r, '999.999')        # traspasos no son gasto

    def test_gastos_mes_selector_de_mes(self):
        self.client.login(username='alda', password='x')
        r = self.client.get(reverse('finanzas:gastos_mes'),
                            {'ano': 2026, 'mes': 7})
        self.assertContains(r, '$111.000')
        self.assertNotContains(r, '$222.000')

    def test_gastos_ano_mes_a_mes(self):
        self.client.login(username='alda', password='x')
        r = self.client.get(reverse('finanzas:gastos_ano'), {'ano': 2026})
        self.assertEqual(r.status_code, 200)
        # 2026 parte en julio (corte de datos).
        # Encabezado abreviado (cabe en pantalla), nombre completo en el title.
        self.assertEqual(r.context['columnas'][0]['nombre'], 'Jul')
        self.assertEqual(r.context['columnas'][0]['largo'], 'Julio')
        self.assertEqual(len(r.context['columnas']), 6)   # jul..dic
        self.assertContains(r, '$333.000')   # fila luz: 111 + 222
        self.assertContains(r, '$388.000')   # total general del anio
        self.assertNotContains(r, '999.999')

    def test_staff_comun_no_ve_los_reportes(self):
        User.objects.create_user('deborah', password='x', is_staff=True)
        self.client.login(username='deborah', password='x')
        for nombre in ('finanzas:gastos_mes', 'finanzas:gastos_ano'):
            self.assertEqual(self.client.get(reverse(nombre)).status_code,
                             302, nombre)

    def test_tarjeta_repo_en_el_panel(self):
        """La tarjeta REPO del panel: Alda ve los enlaces de finanzas
        adentro; el staff común la ve pero solo con Artesanías."""
        self.client.login(username='alda', password='x')
        r = self.client.get('/admin/')
        self.assertContains(r, 'REPO')
        self.assertContains(r, 'Gastos del mes')
        self.assertContains(r, 'Tablero financiero')

        User.objects.create_user('deborah', password='x', is_staff=True)
        self.client.login(username='deborah', password='x')
        r = self.client.get('/admin/')
        self.assertContains(r, 'REPO')
        self.assertNotContains(r, 'Tablero financiero')
        self.assertContains(r, 'Artesan')


BSA_ALDA = ("\r\n".join([
    ";Cartola ",
    ";Numero Cuenta : 99-00138-96",
    ";Fecha Desde : 24/06/2026",
    ";Fecha Hasta : 31/07/2026",
    ";Ejecutivo : EJECUTIVA DE PRUEBA",
    "Fecha;Descripcion;NroDoc.;Cargos;Abonos;Saldo",
    "   28062026;CARGO SEG.Fraude              ;00000000;0000000005000,00;;+0000000061836,00",
    "   02072026;TEF 76485192-7 AREMKO HOTEL SP;00000000;;0000001000000,00;+0000001061836,00",
    "   02072026;820407_PAGO TARJ.CRED. POR SWE;00000000;0000001000000,00;;+0000000061836,00",
    "   06072026;CARGO SEG.Fraude              ;00000000;0000000010618,00;;+0000000051218,00",
    "   10072026;EMISION DE VIGENTE.     298494;00000000;0000000020000,00;;+0000000031218,00",
    "   13072026;TEF 11744727-8 ALDA ANGELICA T;00000000;;0000000400000,00;+0000000431218,00",
]) + "\r\n").encode("latin-1")


class CartolaAldaTest(TestCase):
    """F7: la cuenta personal de Alda — BSA.dat de Scotia Connect."""

    def setUp(self):
        call_command('sembrar_finanzas')

    def _retiro_aremko(self):
        cat = CategoriaFinanciera.objects.create(
            clave='retiros-alda-test', nombre='Retiros Alda', clase='gasto',
            grupo='personales_alda')
        return MovimientoFinanciero.objects.create(
            fecha=date(2026, 7, 2),
            cuenta=CuentaFinanciera.objects.get(clave='mercado_pago'),
            clase='gasto', sentido='sale', monto=1000000, categoria=cat,
            descripcion='Transferencia a ALDA ANGELICA TOLOZA',
            fuente='correo', referencia='correo:mp:test-retiro')

    def test_parser_bsa(self):
        import io

        from .services import parsear_cartola_alda
        datos = parsear_cartola_alda(io.BytesIO(BSA_ALDA))
        self.assertEqual(datos['cuenta_numero'], '99-00138-96')
        self.assertEqual(len(datos['filas']), 6)
        self.assertTrue(datos['cuadra'])            # cadena de saldos sana
        self.assertEqual(datos['cierres_mes']['2026-07'], 431218)
        self.assertEqual(datos['cierres_mes']['2026-06'], 61836)

        por_desc = {f['descripcion']: f for f in datos['filas']}
        aremko = por_desc['TEF 76485192-7 AREMKO HOTEL SP']
        self.assertEqual((aremko['clase'], aremko['propio']), ('traspaso', True))
        self.assertEqual(aremko['abono'], 1000000)
        self.assertEqual(
            por_desc['TEF 11744727-8 ALDA ANGELICA T']['clase'], 'personal')
        self.assertEqual(
            por_desc['820407_PAGO TARJ.CRED. POR SWE']['categoria'],
            'tarjeta_alda')
        # Scotiabank llama «EMISION DE VIGENTE» al vale vista: no se adivina
        # su destino, queda por clasificar y visible.
        self.assertEqual(
            por_desc['EMISION DE VIGENTE.     298494']['categoria'],
            'vale_vista_alda')
        self.assertEqual(
            por_desc['EMISION DE VIGENTE.     298494']['cargo'], 20000)

    def test_reglas_de_clasificacion(self):
        from .services import clasificar_fila_alda
        casos = [
            ('TEF 11744727-8 alda Toloza BCI', 'traslado_alda'),
            ('PAGO AUTOMATICO LINEA CREDITO', 'banco_alda'),
            ('PAGO INTERES LINEA DE CREDITO', 'banco_alda'),
            ('ABONO A L.CREDITO POR SGO', 'banco_alda'),
            ('CARGO SEG.Fraude', 'banco_alda'),
            ('820407_PAGO TARJ.CRED. POR SWE', 'tarjeta_alda'),
            ('REDCOMPRA ARTESANIAS WILMA', 'personales_alda_pc'),
            ('eCOMMERCE DL TEMUCOM', 'personales_alda_pc'),
        ]
        for desc, esperada in casos:
            clase, sentido, cat, propio = clasificar_fila_alda(desc, 1000, 0)
            self.assertEqual((clase, sentido, cat, propio),
                             ('gasto', 'sale', esperada, False), desc)
        # Los abonos de su línea de crédito no son ingreso de nadie.
        self.assertEqual(
            clasificar_fila_alda('TRANSFERENCIA DE LINEA CREDITO', 0, 5000)[0],
            'personal')

    def test_registro_convierte_el_retiro_y_es_idempotente(self):
        import io

        from .services import (estado_fila_cartola, parsear_cartola_alda,
                               registrar_filas_alda)
        retiro = self._retiro_aremko()
        datos = parsear_cartola_alda(io.BytesIO(BSA_ALDA))

        # El abono desde Aremko se ofrece como NUEVO (hay retiro que calza).
        aremko = next(f for f in datos['filas']
                      if 'AREMKO' in f['descripcion'])
        self.assertEqual(estado_fila_cartola('scotiabank_alda', aremko),
                         'nuevo')

        creados, saltados, convertidos = registrar_filas_alda(
            datos['filas'], datos['cierres_mes'])
        # 3 cargos de julio + el abono personal (que también mueve su saldo);
        # solo junio queda fuera por cobertura.
        self.assertEqual((creados, convertidos, saltados), (4, 1, 1))
        # El abono que no viene de Aremko entra como ingreso personal: hace
        # que el saldo de la cuenta puente cuadre con el banco, sin tocar el
        # resultado del negocio.
        personal = MovimientoFinanciero.objects.get(monto=400000)
        self.assertEqual(personal.clase, 'ingreso')
        self.assertEqual(personal.categoria.clave, 'abonos_personales')

        retiro.refresh_from_db()
        self.assertEqual(retiro.clase, 'traspaso')
        self.assertIsNone(retiro.categoria)
        entra = retiro.traspaso_par
        self.assertEqual(entra.cuenta.clave, 'scotiabank_alda')
        self.assertEqual((entra.sentido, int(entra.monto)),
                         ('entra', 1000000))
        # La suma global de traspasos sigue en cero.
        tras = (MovimientoFinanciero.objects.filter(clase='traspaso')
                .values('sentido').annotate(t=Sum('monto')))
        por_sentido = {r['sentido']: int(r['t']) for r in tras}
        self.assertEqual(por_sentido.get('entra'), por_sentido.get('sale'))

        # El vale vista quedó como gasto de la EMPRESA (grupo impuestos).
        vale = MovimientoFinanciero.objects.get(monto=20000,
                                                cuenta__clave='scotiabank_alda')
        self.assertEqual(vale.categoria.clave, 'vale_vista_alda')
        self.assertEqual(vale.categoria.grupo, 'otros')
        # La tarjeta quedó personal.
        tarjeta = MovimientoFinanciero.objects.get(
            monto=1000000, cuenta__clave='scotiabank_alda', clase='gasto')
        self.assertEqual(tarjeta.categoria.grupo, 'personales_alda')

        # Re-registrar: nada nuevo, nada se convierte dos veces.
        creados2, saltados2, convertidos2 = registrar_filas_alda(
            datos['filas'], datos['cierres_mes'])
        self.assertEqual((creados2, convertidos2), (0, 0))

    def test_vista_detecta_bsa_y_marca_personales(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        self._retiro_aremko()
        User.objects.create_superuser('jefe', password='x')
        self.client.login(username='jefe', password='x')

        r = self.client.post(reverse('finanzas:cargar_cartola'),
                             {'archivo': SimpleUploadedFile('BSA.dat',
                                                            BSA_ALDA)})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Scotiabank Alda (personal)')
        self.assertContains(r, 'abono personal')
        self.assertContains(r, 'fuera de cobertura')     # la fila de junio

        # Un BSA de otra cuenta se rechaza con mensaje claro.
        otro = BSA_ALDA.replace(b'99-00138-96', b'11-22222-33')
        r2 = self.client.post(reverse('finanzas:cargar_cartola'),
                              {'archivo': SimpleUploadedFile('BSA.dat', otro)})
        self.assertContains(r2, '11-22222-33')


class MovimientosPegadosTest(TestCase):
    """F7b: la CuentaRUT de Jorge se carga pegando lo que muestra la app."""

    def setUp(self):
        call_command('sembrar_finanzas')
        call_command('aplicar_plan_cuentas', '--aplicar')
        User.objects.create_superuser('jefe', password='x')
        self.client.login(username='jefe', password='x')

    PEGADO = """
14-07-2026 ; Pago Google Ads Google ; -180.000
14-07-2026 ; Tef De Aremko Hotel Spa ; 200.000
06-07-2026 ; Pago Fs Dataforseo ; -56.560
06-07-2026 ; Comision Transaccion Internacional ; -1.075
06-07-2026 ; Tef A Martin Aguilera Toloza 777021 ; -20.000
18-07-2026 ; Abono Convenio Afp Plan Vital ; 162.154
20-06-2026 ; Pago viejo antes del corte ; -9.999
"""

    def test_parseo_y_clasificacion(self):
        from .services import preparar_filas_manual
        filas, errores = preparar_filas_manual(self.PEGADO, 'cuentarut_jorge')
        self.assertEqual(errores, [])
        self.assertEqual(len(filas), 7)
        por_desc = {f['descripcion']: f for f in filas}
        self.assertEqual(por_desc['Pago Google Ads Google']['categoria'],
                         'publicidad')
        self.assertEqual(por_desc['Pago Google Ads Google']['cargo'], 180000)
        self.assertEqual(por_desc['Pago Fs Dataforseo']['categoria'],
                         'infraestructura')
        self.assertEqual(por_desc['Comision Transaccion Internacional']
                         ['categoria'], 'comisiones')
        self.assertEqual(por_desc['Tef A Martin Aguilera Toloza 777021']
                         ['categoria'], 'personales_martin')
        aremko = por_desc['Tef De Aremko Hotel Spa']
        self.assertEqual((aremko['clase'], aremko['abono']), ('traspaso', 200000))
        self.assertEqual(por_desc['Abono Convenio Afp Plan Vital']['clase'],
                         'personal')

    def test_lineas_malas_se_reportan(self):
        from .services import preparar_filas_manual
        filas, errores = preparar_filas_manual(
            "14-07-2026 ; Buena ; -1.000\n"
            "sin fecha ; Mala ; -2.000\n"
            "15-07-2026 ; Sin monto ; abc\n"
            "15-07-2026 solo dos campos\n", 'cuentarut_jorge')
        self.assertEqual(len(filas), 1)
        self.assertEqual(len(errores), 3)
        self.assertEqual([n for n, _, _ in errores], [2, 3, 4])

    def test_carga_e2e_por_la_pagina_es_idempotente(self):
        url = reverse('finanzas:cargar_movimientos')
        r = self.client.post(url, {'cuenta': 'cuentarut_jorge',
                                   'texto': self.PEGADO})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'CuentaRUT Jorge')
        self.assertContains(r, 'publicidad')
        # El abono de AFP entra como personal; el de junio queda fuera.
        self.assertContains(r, 'abono personal')
        self.assertContains(r, 'fuera de cobertura')
        # El abono desde Aremko entra igual (queda por calzar): descartarlo
        # dejaba el saldo de la cuenta corto.
        self.assertContains(r, 'traspaso desde Aremko')
        payload = r.context['payload']
        self.assertEqual(r.context['n_nuevas'], 6)   # 4 cargos + AFP + Aremko

        r2 = self.client.post(url, {'payload': payload, 'confirmar': '1'})
        self.assertEqual(r2.context['resultado']['creados'], 6)
        google = MovimientoFinanciero.objects.get(monto=180000)
        self.assertEqual(google.cuenta.clave, 'cuentarut_jorge')
        self.assertEqual(google.categoria.grupo, 'marketing')
        self.assertEqual(google.fuente, 'captura')

        # Re-pegar el mismo bloque: todo "ya está", nada se duplica.
        r3 = self.client.post(url, {'cuenta': 'cuentarut_jorge',
                                    'texto': self.PEGADO})
        self.assertEqual(r3.context['n_nuevas'], 0)
        self.assertEqual(MovimientoFinanciero.objects.filter(
            cuenta__clave='cuentarut_jorge').count(), 6)

    def test_retiro_a_jorge_se_convierte_en_traspaso(self):
        from .services import registrar_filas_puente, preparar_filas_manual
        cat = CategoriaFinanciera.objects.get(clave='personales_jorge')
        retiro = MovimientoFinanciero.objects.create(
            fecha=date(2026, 7, 14),
            cuenta=CuentaFinanciera.objects.get(clave='mercado_pago'),
            clase='gasto', sentido='sale', monto=200000, categoria=cat,
            descripcion='Transferencia a Jorge', fuente='correo',
            referencia='correo:mp:test-jorge')
        filas, _ = preparar_filas_manual(
            "14-07-2026 ; Tef De Aremko Hotel Spa ; 200.000",
            'cuentarut_jorge')
        creados, saltados, convertidos = registrar_filas_puente(
            filas, 'cuentarut_jorge')
        self.assertEqual((creados, convertidos), (0, 1))
        retiro.refresh_from_db()
        self.assertEqual(retiro.clase, 'traspaso')
        self.assertEqual(retiro.traspaso_par.cuenta.clave, 'cuentarut_jorge')

    def test_dos_movimientos_identicos_el_mismo_dia_entran_los_dos(self):
        from .services import preparar_filas_manual
        filas, _ = preparar_filas_manual(
            "20-07-2026 ; Pago Farmacias Del Doc ; -3.090\n"
            "20-07-2026 ; Pago Farmacias Del Doc ; -3.090\n",
            'cuentarut_jorge')
        self.assertEqual(len({f['referencia'] for f in filas}), 2)


class CalzarRetirosTest(TestCase):
    """El calce a mano cuando los montos no coinciden exactos (Jorge
    2026-08-09): $499.001 en Aremko contra $500.000 en la cuenta puente."""

    def setUp(self):
        call_command('sembrar_finanzas')
        call_command('aplicar_plan_cuentas', '--aplicar')
        User.objects.create_superuser('duenio', 'x@x.cl', 'x')
        self.client.login(username='duenio', password='x')
        self.rut = CuentaFinanciera.objects.get(clave='cuentarut_jorge')
        self.mp = CuentaFinanciera.objects.get(clave='mercado_pago')

    def _pendiente(self, monto=500000, fecha=date(2026, 8, 5)):
        from .services import CLAVE_POR_CALZAR
        cat, _ = CategoriaFinanciera.objects.get_or_create(
            clave=CLAVE_POR_CALZAR,
            defaults={'nombre': 'Abono desde Aremko por calzar',
                      'clase': 'ingreso', 'grupo': 'ingresos'})
        return MovimientoFinanciero.objects.create(
            fecha=fecha, cuenta=self.rut, clase='ingreso', sentido='entra',
            monto=monto, categoria=cat, fuente='captura',
            referencia=f'man:test:{monto}',
            descripcion='Cartola cuentarut jorge: Tef De Aremko Hotel Spa')

    def _retiro(self, monto=499001, fecha=date(2026, 8, 5)):
        return MovimientoFinanciero.objects.create(
            fecha=fecha, cuenta=self.mp, clase='gasto', sentido='sale',
            monto=monto, fuente='correo', referencia=f'correo:mp:{monto}',
            categoria=CategoriaFinanciera.objects.get(
                clave='personales_jorge'),
            descripcion='Transferencia a Jorge Aguilera')

    def test_abono_sin_calce_queda_pendiente_y_no_se_pierde(self):
        """Antes se descartaba y el saldo de la cuenta quedaba corto."""
        from .services import preparar_filas_manual, registrar_filas_puente
        filas, _ = preparar_filas_manual(
            "05-08-2026 ; Tef De Aremko Hotel Spa ; 500.000",
            'cuentarut_jorge')
        creados, saltados, convertidos = registrar_filas_puente(
            filas, 'cuentarut_jorge')
        self.assertEqual((creados, convertidos), (1, 0))
        mov = MovimientoFinanciero.objects.get(cuenta=self.rut)
        self.assertEqual(mov.clase, 'ingreso')
        self.assertEqual(mov.categoria.clave, 'abono_aremko_por_calzar')

    def test_pagina_ofrece_candidatos_ordenados_por_parecido(self):
        abono = self._pendiente()
        self._retiro(monto=120000)              # lejano en monto
        cerca = self._retiro(monto=499001)      # el correcto
        r = self.client.get(reverse('finanzas:calzar_retiros'))
        self.assertEqual(r.status_code, 200)
        p = r.context['pendientes'][0]
        self.assertEqual(p['obj'].id, abono.id)
        self.assertEqual(p['candidatos'][0]['id'], cerca.id)
        self.assertContains(r, '$499.001')

    def test_calce_con_monto_distinto_deja_el_resto_visible(self):
        abono = self._pendiente(monto=500000)
        retiro = self._retiro(monto=499001)
        r = self.client.post(reverse('finanzas:calzar_retiros'),
                             {'abono': abono.id, 'retiro': retiro.id})
        self.assertEqual(r.context['resultado']['comun'], '$499.001')
        self.assertEqual(r.context['resultado']['resto_abono'], '$999')

        abono.refresh_from_db()
        retiro.refresh_from_db()
        self.assertEqual((abono.clase, int(abono.monto)), ('traspaso', 499001))
        self.assertEqual((retiro.clase, int(retiro.monto)), ('traspaso', 499001))
        self.assertEqual(abono.traspaso_par_id, retiro.id)
        self.assertIsNone(retiro.categoria)

        # Los traspasos siguen sumando cero: la regla que sostiene todo.
        agg = {x['sentido']: int(x['t']) for x in
               MovimientoFinanciero.objects.filter(clase='traspaso')
               .values('sentido').annotate(t=Sum('monto'))}
        self.assertEqual(agg['entra'], agg['sale'])

        # El resto del abono queda pendiente, no se inventa nada.
        resto = MovimientoFinanciero.objects.get(monto=999)
        self.assertEqual(resto.categoria.clave, 'abono_aremko_por_calzar')
        self.assertEqual(r.context['pendientes'][0]['obj'].id, resto.id)

    def test_calce_exacto_no_deja_restos(self):
        abono = self._pendiente(monto=230000)
        retiro = self._retiro(monto=230000)
        r = self.client.post(reverse('finanzas:calzar_retiros'),
                             {'abono': abono.id, 'retiro': retiro.id})
        self.assertEqual(r.context['resultado']['resto_abono'], '')
        self.assertEqual(r.context['resultado']['resto_retiro'], '')
        self.assertEqual(r.context['pendientes'], [])

    def test_par_inexistente_no_explota(self):
        r = self.client.post(reverse('finanzas:calzar_retiros'),
                             {'abono': 999999, 'retiro': 999998})
        self.assertEqual(r.status_code, 200)
        self.assertIn('error', r.context)

    def test_staff_comun_no_entra(self):
        User.objects.create_user('deborah', password='x', is_staff=True)
        self.client.login(username='deborah', password='x')
        self.assertEqual(self.client.get(
            reverse('finanzas:calzar_retiros')).status_code, 302)


class EnlacesReportesTest(TestCase):
    """Cada cifra del reporte abre sus movimientos ya filtrados — el camino
    para reasignar un gasto mal clasificado (Jorge 2026-08-09)."""

    def setUp(self):
        call_command('sembrar_finanzas')
        call_command('aplicar_plan_cuentas', '--aplicar')
        User.objects.create_superuser('duenio', 'x@x.cl', 'x')
        self.client.login(username='duenio', password='x')
        self.infra = CategoriaFinanciera.objects.get(clave='infraestructura')
        self.be = CuentaFinanciera.objects.get(clave='bancoestado')
        MovimientoFinanciero.objects.create(
            fecha=date(2026, 7, 10), cuenta=self.be, clase='gasto',
            sentido='sale', monto=954118, categoria=self.infra,
            fuente='captura', referencia='enl:1', descripcion='Render y otros')

    def test_url_movimientos_arma_los_filtros(self):
        from .views import url_movimientos
        url = url_movimientos(grupo='infra_web', ano=2026, mes=7)
        self.assertIn('/admin/finanzas/movimientofinanciero/', url)
        self.assertIn('clase__exact=gasto', url)
        self.assertIn('categoria__grupo__exact=infra_web', url)
        self.assertIn('fecha__year=2026', url)
        self.assertIn('fecha__month=7', url)
        # Con categoría concreta manda la categoría, no el grupo.
        url2 = url_movimientos(grupo='infra_web', cat_id=7, ano=2026)
        self.assertIn('categoria__id__exact=7', url2)
        self.assertNotIn('categoria__grupo', url2)

    def test_celdas_del_ano_enlazan_al_mes_correcto(self):
        r = self.client.get(reverse('finanzas:gastos_ano'), {'ano': 2026})
        fila = [f for f in r.context['filas']
                if f['nombre'] == 'Infraestructura web e IA'][0]
        julio = fila['celdas'][0]
        self.assertEqual(julio['txt'], '$954.118')
        self.assertIn('categoria__grupo__exact=infra_web', julio['url'])
        self.assertIn('fecha__month=7', julio['url'])
        # Un mes sin gasto no lleva enlace: no hay nada que abrir.
        self.assertEqual(fila['celdas'][1]['txt'], '')
        self.assertEqual(fila['celdas'][1]['url'], '')
        # Y el enlace vive en el HTML.
        self.assertContains(r, 'categoria__grupo__exact=infra_web')

    def test_celdas_del_mes_enlazan_tambien_por_cuenta(self):
        r = self.client.get(reverse('finanzas:gastos_mes'),
                            {'ano': 2026, 'mes': 7})
        fila = [f for f in r.context['filas']
                if f['nombre'] == 'Infraestructura web e IA'][0]
        celda = fila['celdas'][0]
        self.assertIn(f'cuenta__id__exact={self.be.id}', celda['url'])
        # El total de la fila NO filtra por cuenta: son todas.
        self.assertNotIn('cuenta__id__exact', fila['total']['url'])


class SaludFuentesTest(TestCase):
    """La página que audita a la máquina (Jorge 2026-08-10): ¿las cartolas
    cubren el período? ¿siguen llegando MP y SumUp?"""

    def setUp(self):
        call_command('sembrar_finanzas')
        call_command('aplicar_plan_cuentas', '--aplicar')
        User.objects.create_superuser('duenio', 'x@x.cl', 'x')
        self.client.login(username='duenio', password='x')
        self.be = CuentaFinanciera.objects.get(clave='bancoestado')
        self.pc = CategoriaFinanciera.objects.get(clave='por_clasificar')

    def _mov(self, dia, monto=1000, ref=None, cuenta=None):
        return MovimientoFinanciero.objects.create(
            fecha=dia, cuenta=cuenta or self.be, clase='gasto', sentido='sale',
            monto=monto, categoria=self.pc, fuente='captura',
            referencia=ref or f'sal:{dia}:{monto}')

    def test_tramos_vacios_detecta_el_hueco_largo_y_no_un_dia_suelto(self):
        from .views import _tramos_vacios
        fechas = {date(2026, 7, 1), date(2026, 7, 3), date(2026, 7, 15)}
        tramos = _tramos_vacios(fechas, date(2026, 7, 1), date(2026, 7, 15))
        # El 2 de julio solo (1 día) no es salto; del 4 al 14 sí (11 días).
        self.assertEqual(len(tramos), 1)
        self.assertEqual((tramos[0][0], tramos[0][1], tramos[0][2]),
                         (date(2026, 7, 4), date(2026, 7, 14), 11))

    def test_pagina_reporta_cobertura_atraso_y_saltos(self):
        self._mov(date(2026, 7, 1))
        self._mov(date(2026, 7, 2))
        self._mov(date(2026, 7, 20))        # deja un salto largo en medio
        SaldoMensual.objects.create(cuenta=self.be, periodo=date(2026, 7, 1),
                                    saldo_cierre=100, fuente='cartola')

        r = self.client.get(reverse('finanzas:salud_fuentes'))
        self.assertEqual(r.status_code, 200)
        fila = [c for c in r.context['cartolas']
                if 'BancoEstado' in c['nombre']][0]
        self.assertEqual(fila['primera'], date(2026, 7, 1))
        self.assertEqual(fila['ultima'], date(2026, 7, 20))
        self.assertEqual(fila['n_movs'], 3)
        self.assertTrue(fila['atrasada'])       # nada nuevo hace semanas
        self.assertTrue(fila['ancla'])
        self.assertEqual(len(fila['tramos']), 1)
        self.assertContains(r, 'salto')

        # Una cuenta sin nada cargado se declara vacía, no se esconde.
        vacia = [c for c in r.context['cartolas']
                 if c['nombre'] == 'Scotiabank'][0]
        self.assertTrue(vacia['vacia'])

    def test_avisa_cuando_la_cartola_no_cubre_el_inicio(self):
        """El hueco que más fácil se pasa por alto: la cartola empieza
        despues del 1 de julio (BancoEstado partia el 14-07 en prod)."""
        self._mov(date(2026, 7, 14))
        self._mov(date(2026, 7, 15))
        r = self.client.get(reverse('finanzas:salud_fuentes'))
        fila = [c for c in r.context['cartolas']
                if 'BancoEstado' in c['nombre']][0]
        self.assertEqual(fila['falta_inicio'], 13)
        self.assertContains(r, 'no cubre el comienzo del')

        # Si parte el dia 1, no hay nada que avisar.
        self._mov(date(2026, 7, 1))
        r2 = self.client.get(reverse('finanzas:salud_fuentes'))
        fila2 = [c for c in r2.context['cartolas']
                 if 'BancoEstado' in c['nombre']][0]
        self.assertEqual(fila2['falta_inicio'], 0)

    def test_fuentes_automaticas_marcan_alerta_cuando_dejan_de_llegar(self):
        r = self.client.get(reverse('finanzas:salud_fuentes'))
        por_nombre = {f['nombre']: f for f in r.context['fuentes']}
        sumup = por_nombre['SumUp · comisiones (API)']
        self.assertIsNone(sumup['ultimo'])
        self.assertTrue(sumup['alerta'])

        # Con un dato de hoy deja de alertar.
        self._mov(date.today(), ref='sumup:fee:1')
        r2 = self.client.get(reverse('finanzas:salud_fuentes'))
        sumup2 = {f['nombre']: f for f in
                  r2.context['fuentes']}['SumUp · comisiones (API)']
        self.assertEqual(sumup2['dias'], 0)
        self.assertFalse(sumup2['alerta'])

    def test_controles_de_traspasos_y_pendientes(self):
        # Un traspaso por un solo lado: el descalce se declara.
        MovimientoFinanciero.objects.create(
            fecha=date(2026, 8, 1), cuenta=self.be, clase='traspaso',
            sentido='sale', monto=50000, fuente='manual', referencia='t:1')
        r = self.client.get(reverse('finanzas:salud_fuentes'))
        self.assertFalse(r.context['traspasos_cuadran'])
        self.assertEqual(r.context['descalce'], '$50.000')
        # Y los gastos por clasificar se cuentan.
        self._mov(date(2026, 8, 2), monto=7000)
        r2 = self.client.get(reverse('finanzas:salud_fuentes'))
        self.assertEqual(r2.context['n_sin_clasificar'], 1)
        self.assertEqual(r2.context['monto_sin_clasificar'], '$7.000')

    def test_staff_comun_no_entra(self):
        User.objects.create_user('deborah', password='x', is_staff=True)
        self.client.login(username='deborah', password='x')
        self.assertEqual(self.client.get(
            reverse('finanzas:salud_fuentes')).status_code, 302)


class CartolaBancoEstadoEnLineaTest(TestCase):
    """Segunda variante real de BancoEstado (2026-08-10): «Cartola en Línea»,
    hoja Registros, sin fechas de período y con las filas DESORDENADAS."""

    RESUMEN = [
        ('Rut Empresa', '', '', '', '76.485.192-7'),
        ('Chequera Electrónica', '', '', '', '82370351925'),
        ('Saldo', '', '', '', ''),
        ('Inicial', '', '', '', '$ 16.127.957'),
        ('Saldo Contable', '', '', '', '$ 16.203.323'),
        ('Total Abonos', '', '', '', '$ 575.366'),
        ('Total Cargos', '', '', '', '$ 500.000'),
    ]
    # Tal como las entrega el banco: fuera de orden.
    REGISTROS = [
        ('Fecha', 'Sucursal', 'N° Operación', 'Descripción', 'Cargos',
         'Abonos', 'Saldo'),
        ('06/08/2026', 'STGO', '1', 'TEF A TOLOZA POBLETE ALDA ANGELICA',
         '$ 500.000', '', '$ 16.049.397'),
        ('05/08/2026', 'STGO', '2', 'TEF DE FLOW PAGOS CHILE SPA', '',
         '$ 202.029', '$ 16.491.675'),
        ('04/08/2026', 'STGO', '3', 'TEF DE FLOW S A', '', '$ 48.102',
         '$ 16.176.059'),
        ('05/08/2026', 'STGO', '4', 'TEF DE FLOW S A', '', '$ 57.722',
         '$ 16.549.397'),
        ('07/08/2026', 'STGO', '5', 'TEF DE FLOW S A', '', '$ 96.204',
         '$ 16.203.323'),
        ('04/08/2026', 'STGO', '6', 'TEF DE SUMUP CHILE PAYMENTS S A', '',
         '$ 113.587', '$ 16.289.646'),
        ('06/08/2026', 'STGO', '7', 'TEF DE FLOW S A', '', '$ 57.722',
         '$ 16.107.119'),
    ]

    def _archivo(self, registros=None):
        import io

        import openpyxl
        wb = openpyxl.Workbook()
        hoja = wb.active
        hoja.title = 'Resumen'
        for fila in self.RESUMEN:
            hoja.append(list(fila))
        reg = wb.create_sheet('Registros')
        for fila in (registros or self.REGISTROS):
            reg.append(list(fila))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_ordena_por_la_cadena_de_saldos_no_por_fecha(self):
        from .services import parsear_cartola_bancoestado
        d = parsear_cartola_bancoestado(self._archivo())
        self.assertEqual(d['cuenta_numero'], '82370351925')
        self.assertEqual(d['cadena_rota'], 0)
        self.assertTrue(d['cuadra'])
        self.assertEqual(d['saldo_inicial'], 16127957)
        self.assertEqual(d['saldo_final_calculado'], 16203323)
        self.assertEqual(d['total_abonos'], 575366)
        self.assertEqual(d['total_cargos'], 500000)
        # El orden lo dicta el saldo: dentro del 04-08, FLOW va antes que
        # SumUp aunque en el archivo venga al revés.
        desc = [f['descripcion'][:20] for f in d['filas']]
        self.assertEqual(desc[0], 'TEF DE FLOW S A')
        self.assertEqual(desc[1], 'TEF DE SUMUP CHILE P')
        # Y la cadena queda coherente de punta a punta.
        saldo = d['saldo_inicial']
        for f in d['filas']:
            saldo += f['abono'] - f['cargo']
            self.assertEqual(saldo, f['saldo'], f['descripcion'])
        self.assertEqual(saldo, 16203323)

    def test_clasifica_con_el_plan_de_cuentas(self):
        from .services import parsear_cartola_bancoestado
        d = parsear_cartola_bancoestado(self._archivo())
        por_desc = {f['descripcion']: f for f in d['filas']}
        self.assertEqual(por_desc['TEF DE SUMUP CHILE PAYMENTS S A']['categoria'],
                         'liquidacion_sumup')
        self.assertEqual(por_desc['TEF DE FLOW S A']['categoria'],
                         'liquidacion_flow')
        # El retiro a Alda se reconoce solo: si no, no aparecía de candidato
        # en Calzar retiros (visto 2026-08-10).
        self.assertEqual(
            por_desc['TEF A TOLOZA POBLETE ALDA ANGELICA']['categoria'],
            'personales_alda')

    def test_una_fila_que_no_encaja_se_declara_cadena_rota(self):
        from .services import parsear_cartola_bancoestado
        rotas = list(self.REGISTROS)
        rotas[3] = ('04/08/2026', 'STGO', '3', 'TEF DE FLOW S A', '',
                    '$ 48.102', '$ 99.999.999')      # saldo imposible
        d = parsear_cartola_bancoestado(self._archivo(rotas))
        self.assertGreater(d['cadena_rota'], 0)
        self.assertFalse(d['cuadra'])
        # Igual se muestran todas las filas: nada se esconde.
        self.assertEqual(len(d['filas']), 7)

    def test_se_registra_y_es_idempotente(self):
        from .services import (parsear_cartola_bancoestado,
                               registrar_filas_cartola)
        call_command('sembrar_finanzas')
        call_command('aplicar_plan_cuentas', '--aplicar')
        d = parsear_cartola_bancoestado(self._archivo())
        creados, _ = registrar_filas_cartola(d['filas'], d['cierres_mes'],
                                             cuenta_clave='bancoestado')
        self.assertEqual(creados, 7)
        # La transferencia a Alda es TRASPASO (dos piernas), no retiro:
        # esa plata cambió de bolsillo, no salió del negocio.
        piernas = MovimientoFinanciero.objects.filter(monto=500000)
        self.assertEqual(piernas.count(), 2)
        self.assertEqual({p.clase for p in piernas}, {'traspaso'})
        self.assertEqual({p.cuenta.clave for p in piernas},
                         {'bancoestado', 'scotiabank_alda'})
        creados2, _ = registrar_filas_cartola(d['filas'], d['cierres_mes'],
                                              cuenta_clave='bancoestado')
        self.assertEqual(creados2, 0)


class TraspasoACuentaPuenteTest(TestCase):
    """Corrección de Jorge (2026-08-10): la plata que Aremko manda a la
    cuenta de Alda NO es retiro, es traspaso — cambia de bolsillo. El retiro
    ocurre después, cuando ella gasta en algo personal."""

    def setUp(self):
        call_command('sembrar_finanzas')
        call_command('aplicar_plan_cuentas', '--aplicar')
        self.be = CuentaFinanciera.objects.get(clave='bancoestado')
        self.alda = CuentaFinanciera.objects.get(clave='scotiabank_alda')

    def _fila(self, desc, monto=500000, dia=date(2026, 8, 6), ref='be:t1'):
        return {'fecha': dia.isoformat(), 'descripcion': desc,
                'cargo': monto, 'abono': 0, 'saldo': 0, 'clase': 'gasto',
                'sentido': 'sale', 'categoria': 'personales_alda',
                'referencia': ref}

    def test_a_cuenta_seguida_es_traspaso_con_dos_piernas(self):
        from .services import registrar_filas_cartola
        creados, _ = registrar_filas_cartola(
            [self._fila('TEF A TOLOZA POBLETE ALDA ANGELICA')],
            cuenta_clave='bancoestado')
        self.assertEqual(creados, 1)

        sale = MovimientoFinanciero.objects.get(cuenta=self.be)
        entra = MovimientoFinanciero.objects.get(cuenta=self.alda)
        self.assertEqual(sale.clase, 'traspaso')
        self.assertIsNone(sale.categoria)
        self.assertEqual((entra.clase, entra.sentido), ('traspaso', 'entra'))
        self.assertEqual(sale.traspaso_par_id, entra.id)
        self.assertEqual(entra.traspaso_par_id, sale.id)
        # Y NO cuenta como retiro de la familia.
        self.assertEqual(MovimientoFinanciero.objects.filter(
            clase='gasto', categoria__grupo='personales_alda').count(), 0)

    def test_a_una_cuenta_que_no_seguimos_sigue_siendo_retiro(self):
        from .services import registrar_filas_cartola
        registrar_filas_cartola(
            [self._fila('TEF A ALDA BCI', ref='be:t2')],
            cuenta_clave='bancoestado')
        mov = MovimientoFinanciero.objects.get(cuenta=self.be)
        self.assertEqual(mov.clase, 'gasto')
        self.assertEqual(mov.categoria.grupo, 'personales_alda')
        # Nada entró a la cuenta puente: esa plata salió de nuestra vista.
        self.assertEqual(
            MovimientoFinanciero.objects.filter(cuenta=self.alda).count(), 0)

    def test_usa_el_abono_que_ya_esperaba_en_vez_de_duplicar(self):
        """Si su cartola se cargó primero, ese abono ES la pierna que entra."""
        from .services import CLAVE_POR_CALZAR, registrar_filas_cartola
        cat, _ = CategoriaFinanciera.objects.get_or_create(
            clave=CLAVE_POR_CALZAR,
            defaults={'nombre': 'Abono desde Aremko por calzar',
                      'clase': 'ingreso', 'grupo': 'ingresos'})
        pendiente = MovimientoFinanciero.objects.create(
            fecha=date(2026, 8, 5), cuenta=self.alda, clase='ingreso',
            sentido='entra', monto=500000, categoria=cat, fuente='captura',
            referencia='alda:x', descripcion='TEF AREMKO HOTEL SPA')

        registrar_filas_cartola(
            [self._fila('TEF A TOLOZA POBLETE ALDA ANGELICA')],
            cuenta_clave='bancoestado')

        pendiente.refresh_from_db()
        self.assertEqual(pendiente.clase, 'traspaso')
        self.assertIsNone(pendiente.categoria)
        # Una sola pierna que entra: la plata no se duplicó en su cuenta.
        self.assertEqual(MovimientoFinanciero.objects.filter(
            cuenta=self.alda, sentido='entra').count(), 1)
        agg = {r['sentido']: int(r['t']) for r in
               MovimientoFinanciero.objects.filter(clase='traspaso')
               .values('sentido').annotate(t=Sum('monto'))}
        self.assertEqual(agg['entra'], agg['sale'])

    def test_su_cartola_despues_no_vuelve_a_meter_la_plata(self):
        """Orden inverso: primero Aremko, después la cartola de ella."""
        from .services import registrar_filas_cartola, registrar_filas_puente
        registrar_filas_cartola(
            [self._fila('TEF A TOLOZA POBLETE ALDA ANGELICA')],
            cuenta_clave='bancoestado')
        antes = MovimientoFinanciero.objects.filter(cuenta=self.alda).count()

        fila_de_ella = {'fecha': '2026-08-05',
                        'descripcion': 'TEF 76485192-7 AREMKO',
                        'cargo': 0, 'abono': 500000, 'saldo': 0,
                        'clase': 'traspaso', 'sentido': 'entra',
                        'categoria': '', 'propio': True,
                        'referencia': 'alda:zz'}
        creados, saltados, convertidos = registrar_filas_puente(
            [fila_de_ella], 'scotiabank_alda')
        self.assertEqual((creados, convertidos), (0, 0))
        self.assertEqual(saltados, 1)
        self.assertEqual(
            MovimientoFinanciero.objects.filter(cuenta=self.alda).count(),
            antes)

    def test_traslado_a_si_mismo_no_es_traspaso(self):
        """Desde la CuentaRUT a otra cuenta de Jorge: no hay a dónde ir."""
        from .services import destino_puente
        self.assertIsNone(destino_puente('TEF A JORGE ANTONIO AGUILERA',
                                         'cuentarut_jorge'))
        self.assertEqual(destino_puente('TEF A JORGE ANTONIO AGUILERA',
                                        'bancoestado'), 'cuentarut_jorge')


class VerificacionMPDevolucionesTest(TestCase):
    """Caso real de Jorge (2026-08-10): un cliente anula, se le devuelve la
    plata y vuelve a tomar la reserva. La API de MP no reporta devoluciones,
    así que restarlas del lado sistema las hacía aparecer como descuadre."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._mover_senales(desconectar=True)

    @classmethod
    def tearDownClass(cls):
        cls._mover_senales(desconectar=False)
        super().tearDownClass()

    @staticmethod
    def _mover_senales(desconectar):
        """Las señales CRM/Meta piden tablas del drift AR-033/034 que el shim
        de tests no migra — mismo patrón que tests_checkout_agenda."""
        from django.db.models.signals import post_save

        from control_gestion.signals import react_to_reserva_change
        from ventas.models import Pago, VentaReserva
        from ventas.signals.main_signals import actualizar_tramo_y_premios_on_pago

        mover = post_save.disconnect if desconectar else post_save.connect
        for receptor in (actualizar_tramo_y_premios_on_pago,
                         react_to_reserva_change):
            for emisor in (VentaReserva, Pago):
                mover(receptor, sender=emisor)

    def setUp(self):
        # Limpiar ANTES también: otra suite pudo dejar en el thread-local un
        # usuario que ya no existe (su transacción se revirtió), y el Pago que
        # creamos acá heredaría esa FK muerta. Pasaba solo en conjunto.
        self._limpiar_thread_local()
        call_command('sembrar_finanzas')
        User.objects.create_superuser('duenio', 'x@x.cl', 'x')
        self.client.login(username='duenio', password='x')

    def tearDown(self):
        self._limpiar_thread_local()
        super().tearDown()

    @staticmethod
    def _limpiar_thread_local():
        """ThreadLocalMiddleware guarda el usuario del request y nadie lo
        limpia — mismo remedio que tests_comandas_cocina."""
        from ventas import middleware
        middleware._thread_locals.user = None

    def _pago(self, monto, cuando):
        from ventas.models import Cliente, Pago, VentaReserva
        cliente, _ = Cliente.objects.get_or_create(
            telefono='+56900000123', defaults={'nombre': 'Cliente Prueba'})
        reserva, _ = VentaReserva.objects.get_or_create(
            cliente=cliente, defaults={'fecha_reserva': timezone.now()})
        return Pago.objects.create(venta_reserva=reserva, monto=monto,
                                   metodo_pago='mercadopago_link',
                                   fecha_pago=cuando)

    def test_la_devolucion_no_se_cuenta_como_descuadre(self):
        cuando = timezone.now() - timedelta(days=1)
        dia = timezone.localtime(cuando).date()
        # Cobró 200.000, se devolvió, volvió a tomar por 200.000.
        self._pago(200000, cuando)
        self._pago(-200000, cuando)
        self._pago(200000, cuando)
        # MP solo ve los DOS cobros: no reporta la devolución.
        for i, monto in enumerate((200000, 200000)):
            MovimientoMP.objects.create(mp_payment_id=f'dev{i}', monto=monto,
                                        fecha=cuando)

        r = self.client.get(reverse('finanzas:tablero'))
        fila = [f for f in r.context['verif_mp'] if f['dia'] == dia][0]
        self.assertEqual(fila['sistema'], '$400.000')   # cobros, sin netear
        self.assertEqual(fila['dev'], '$200.000')       # la devolución, aparte
        self.assertEqual(fila['mp'], '$400.000')
        self.assertEqual(fila['dif'], '')               # ya no hay descuadre
        self.assertFalse(fila['dif_alerta'])
        self.assertTrue(r.context['verif_cuadra'])
        self.assertEqual(r.context['verif_dev'], '$200.000')

    def test_una_diferencia_de_verdad_sigue_saltando(self):
        """El arreglo no puede tapar la plata que sí falta registrar."""
        cuando = timezone.now() - timedelta(days=1)
        self._pago(100000, cuando)
        MovimientoMP.objects.create(mp_payment_id='real1', monto=100000,
                                    fecha=cuando)
        MovimientoMP.objects.create(mp_payment_id='real2', monto=50000,
                                    fecha=cuando)
        r = self.client.get(reverse('finanzas:tablero'))
        self.assertFalse(r.context['verif_cuadra'])
        self.assertEqual(r.context['verif_dif'], '+$50.000')


class ComisionDeMPTest(TestCase):
    """El cálculo que responde «¿subió la comisión o subieron las ventas?»."""

    def test_solo_cuenta_lo_que_paga_el_vendedor(self):
        from finanzas.management.commands.comparar_comisiones_mp import comision_de
        pago = {'fee_details': [
            {'type': 'mercadopago_fee', 'amount': 3500, 'fee_payer': 'collector'},
            {'type': 'financing_fee', 'amount': 8000, 'fee_payer': 'payer'},
            {'type': 'application_fee', 'amount': 500},   # sin fee_payer
        ]}
        # 3.500 del vendedor + 500 del que no declara pagador (default
        # collector); los 8.000 que paga el cliente NO son costo nuestro.
        self.assertEqual(comision_de(pago), 4000)

    def test_sin_comisiones_devuelve_cero(self):
        from finanzas.management.commands.comparar_comisiones_mp import comision_de
        self.assertEqual(comision_de({}), 0)
        self.assertEqual(comision_de({'fee_details': []}), 0)


TEXTO_TARJETA = """10-08-26, 1:35 p. m.
FFFFEEEECCCCHHHHAAAA DDDDEEEESSSSCCCCRRRRIIIIPPPPCCCCIIIIÓÓÓÓNNNN CCCCIIIIUUUUDDDDAAAADDDD MMMMOOOONNNNTTTTOOOO
02/07/2026 PAGO EN EFECTIVO $-1.000.000
23/06/2026 CASA TRONCO PUERTO LLANQUIHUE $298.000
VARAS
28/06/2026 JUMBO SUPER DONA EMA PUERTO $71.137
VARAS
06/07/2026 COPEC ASISTIDO LAS CONDES $79.465
14/07/2026 COMISION COMPRA $745
INTERNACIONAL
18/05/2025 MARKETING Y EVENTOS SANTIAGO ORGAN TASA INT. $5.353
about:blank Página 1 de 5"""


class TarjetaAldaTest(TestCase):
    """F7c: el PDF de la tarjeta de Alda. Las COMPRAS son el gasto real; el
    pago de la tarjeta es traspaso desde su cuenta corriente, no gasto."""

    def setUp(self):
        call_command('sembrar_finanzas')
        call_command('aplicar_plan_cuentas', '--aplicar')
        self.tarjeta = CuentaFinanciera.objects.get(clave='tarjeta_alda_1')
        self.corriente = CuentaFinanciera.objects.get(clave='scotiabank_alda')

    def _filas(self):
        from .services import parsear_lineas_tarjeta
        return parsear_lineas_tarjeta(TEXTO_TARJETA, 'tarjeta_alda_1')

    def test_lee_las_lineas_y_pega_las_continuaciones(self):
        filas = self._filas()
        self.assertEqual(len(filas), 6)
        por_monto = {f['cargo'] or f['abono']: f for f in filas}
        # La ciudad partida en dos renglones se pega a su compra.
        self.assertIn('VARAS', por_monto[298000]['descripcion'])
        # El pago viene negativo y se marca como tal.
        pago = por_monto[1000000]
        self.assertTrue(pago['es_pago'])
        self.assertEqual(pago['abono'], 1000000)
        self.assertEqual(pago['cargo'], 0)
        # Las líneas de ruido del navegador no entran.
        self.assertFalse(any('about:blank' in f['descripcion'] for f in filas))
        self.assertFalse(any('FFFF' in f['descripcion'] for f in filas))

    def test_clasificacion_default_es_por_clasificar(self):
        """Es tarjeta mixta: el sistema no adivina qué es de Aremko."""
        filas = {f['descripcion'][:12]: f for f in self._filas()}
        self.assertEqual(filas['CASA TRONCO ']['categoria'], 'por_clasificar')
        # Las comisiones del banco sí se reconocen.
        self.assertEqual(filas['COMISION COM']['categoria'], 'banco_alda')

    def test_repetir_el_archivo_no_duplica(self):
        """La compra que hoy está «por facturar» y mañana «facturada» tiene
        la misma fecha, glosa y monto: se reconoce sola."""
        from .services import registrar_filas_tarjeta
        # De las 6 líneas: 2 compras de julio entran; 3 son de junio o de
        # 2025 (fuera del corte) y 1 es el pago, que espera su cargo.
        creados, saltados, _, sin_calce = registrar_filas_tarjeta(
            self._filas(), 'tarjeta_alda_1')
        self.assertEqual((creados, saltados, sin_calce), (2, 3, 1))

        creados2, saltados2, _, _ = registrar_filas_tarjeta(
            self._filas(), 'tarjeta_alda_1')
        self.assertEqual(creados2, 0)
        self.assertEqual(saltados2, 5)   # las 2 ya escritas + las 3 de fuera

    def test_el_pago_calza_con_el_cargo_de_la_cuenta_corriente(self):
        from .services import registrar_filas_tarjeta
        cat, _ = CategoriaFinanciera.objects.get_or_create(
            clave='tarjeta_alda',
            defaults={'nombre': 'Tarjeta de crédito (Alda)', 'clase': 'gasto',
                      'grupo': 'personales_alda'})
        cargo = MovimientoFinanciero.objects.create(
            fecha=date(2026, 7, 2), cuenta=self.corriente, clase='gasto',
            sentido='sale', monto=1000000, fuente='captura',
            referencia='alda:pagotarj', categoria=cat,
            descripcion='Cartola alda: 820407_PAGO TARJ.CRED. POR SWE')

        creados, _, calzados, sin_calce = registrar_filas_tarjeta(
            self._filas(), 'tarjeta_alda_1')
        self.assertEqual((calzados, sin_calce), (1, 0))

        cargo.refresh_from_db()
        # Deja de ser gasto: era plata que cambiaba de bolsillo.
        self.assertEqual(cargo.clase, 'traspaso')
        self.assertIsNone(cargo.categoria)
        self.assertEqual(cargo.traspaso_par.cuenta.clave, 'tarjeta_alda_1')
        agg = {r['sentido']: int(r['t']) for r in
               MovimientoFinanciero.objects.filter(clase='traspaso')
               .values('sentido').annotate(t=Sum('monto'))}
        self.assertEqual(agg['entra'], agg['sale'])

    def test_pago_sin_su_cargo_se_informa_y_no_se_inventa(self):
        """Sin la cartola de la cuenta corriente cargada, el pago espera."""
        from .services import registrar_filas_tarjeta
        creados, _, calzados, sin_calce = registrar_filas_tarjeta(
            self._filas(), 'tarjeta_alda_1')
        self.assertEqual((calzados, sin_calce), (0, 1))
        # No quedó una pierna suelta que rompa la suma cero.
        self.assertEqual(MovimientoFinanciero.objects.filter(
            clase='traspaso').count(), 0)

    def test_las_cuotas_viejas_quedan_fuera_del_corte(self):
        """Las compras en cuotas de meses anteriores (dicen «TASA INT.»)
        traen la fecha de la compra ORIGINAL, así que caen antes de julio
        2026 y no entran. Se ven en la propuesta como fuera de cobertura."""
        from .services import registrar_filas_tarjeta
        registrar_filas_tarjeta(self._filas(), 'tarjeta_alda_1')
        self.assertFalse(MovimientoFinanciero.objects.filter(
            monto=5353).exists())


class DetectarRecurrentesTest(TestCase):
    """Detección de gastos recurrentes y sus fechas (Jorge 2026-08-10)."""

    def test_normaliza_el_mismo_comercio_escrito_de_varias_formas(self):
        from finanzas.management.commands.detectar_recurrentes import (
            nombre_comercio)
        # El banco escribe Meta de cinco maneras: sin alias serían cinco
        # comercios distintos y la recurrencia no se vería.
        for glosa in ('FACEBK *AB12CD', 'Cartola alda: PAGO Facebook Ireland',
                      'META PLATFORMS IRELAND', 'Tarjeta: FACEBOOK ADS'):
            self.assertEqual(nombre_comercio(glosa), 'Meta Ads', glosa)
        self.assertEqual(nombre_comercio('Pago Google Ads Google'),
                         'Google Ads')
        self.assertEqual(nombre_comercio('Pago Render.com'), 'Render')

    def test_limpia_identificadores_cuando_no_hay_alias(self):
        from finanzas.management.commands.detectar_recurrentes import (
            nombre_comercio)
        # Los ids de transacción cambian en cada cobro: si no se sacan, el
        # mismo comercio nunca se agrupa consigo mismo.
        a = nombre_comercio('Cartola alda: MERPAGO*S 760803626 LAS CONDES')
        b = nombre_comercio('Cartola alda: MERPAGO*S 998877665 LAS CONDES')
        self.assertEqual(a, b)
        self.assertEqual(nombre_comercio(''), 'Sin descripción')

    def test_reconoce_los_tres_patrones_que_le_importan_a_jorge(self):
        from finanzas.management.commands.detectar_recurrentes import patron_de
        # Mensual, como el arriendo de un servicio.
        mensual = [date(2026, 5, 10), date(2026, 6, 10), date(2026, 7, 10)]
        etiqueta, dias = patron_de(mensual)
        self.assertEqual(etiqueta, 'mensual')
        self.assertIn(dias, (30, 31))
        # Facebook: varias veces en el mismo mes.
        seguido = [date(2026, 7, 2), date(2026, 7, 9), date(2026, 7, 16),
                   date(2026, 7, 23), date(2026, 8, 1)]
        etiqueta2, _ = patron_de(seguido)
        self.assertIn('varias veces al mes', etiqueta2)
        # Un cobro solo no es un patrón.
        self.assertEqual(patron_de([date(2026, 7, 1)]), ('una sola vez', None))

    def test_el_comando_corre_y_no_escribe_nada(self):
        from io import StringIO
        call_command('sembrar_finanzas')
        call_command('aplicar_plan_cuentas', '--aplicar')
        cuenta = CuentaFinanciera.objects.get(clave='bancoestado')
        cat = CategoriaFinanciera.objects.get(clave='publicidad')
        for i, dia in enumerate((5, 12, 19, 26)):
            MovimientoFinanciero.objects.create(
                fecha=date(2026, 7, dia), cuenta=cuenta, clase='gasto',
                sentido='sale', monto=50000 + i, categoria=cat,
                fuente='captura', referencia=f'rec:{i}',
                descripcion=f'FACEBK *X{i}00{i}')
        antes = MovimientoFinanciero.objects.count()

        salida = StringIO()
        call_command('detectar_recurrentes', '--desde', '2026-07-01',
                     stdout=salida)
        texto = salida.getvalue()
        self.assertIn('Meta Ads', texto)
        self.assertIn('varias veces al mes', texto)
        self.assertIn('días del mes: 5, 12, 19, 26', texto)
        self.assertIn('próximo cobro estimado', texto)
        self.assertEqual(MovimientoFinanciero.objects.count(), antes)


class DuplicadosTest(TestCase):
    """Doble carga y atribución doble (Jorge 2026-08-10)."""

    def setUp(self):
        call_command('sembrar_finanzas')
        call_command('aplicar_plan_cuentas', '--aplicar')
        self.sc = CuentaFinanciera.objects.get(clave='scotiabank')
        self.visa = CuentaFinanciera.objects.get(clave='visa_2936')
        self.rut = CuentaFinanciera.objects.get(clave='cuentarut_jorge')
        self.cat = CategoriaFinanciera.objects.get(clave='por_clasificar')

    def _fila(self, desc, monto, saldo, ref):
        return {'fecha': '2026-07-20', 'descripcion': desc, 'cargo': monto,
                'abono': 0, 'saldo': saldo, 'clase': 'gasto',
                'sentido': 'sale', 'categoria': 'por_clasificar',
                'referencia': ref}

    def test_dos_exports_del_mismo_periodo_no_duplican(self):
        """El saldo cambia entre exports y por eso la referencia difería;
        ahora se compara por fecha, monto y glosa."""
        from .services import registrar_filas_cartola
        creados, _ = registrar_filas_cartola(
            [self._fila('REDCOMPRA EXPRESS PUERTO', 157433, 900000, 'sc:a')],
            cuenta_clave='scotiabank')
        self.assertEqual(creados, 1)
        # Mismo movimiento, otro export: otro saldo → otra referencia.
        creados2, saltados2 = registrar_filas_cartola(
            [self._fila('REDCOMPRA EXPRESS PUERTO', 157433, 111111, 'sc:b')],
            cuenta_clave='scotiabank')
        self.assertEqual((creados2, saltados2), (0, 1))
        self.assertEqual(MovimientoFinanciero.objects.filter(
            cuenta=self.sc).count(), 1)

    def test_dos_compras_iguales_de_verdad_entran_las_dos(self):
        """Dos cobros idénticos el mismo día existen (dos cafés seguidos):
        el guardián permite tantos como traiga el archivo."""
        from .services import registrar_filas_cartola
        creados, _ = registrar_filas_cartola(
            [self._fila('FARMACIA', 3090, 900000, 'sc:1'),
             self._fila('FARMACIA', 3090, 896910, 'sc:2')],
            cuenta_clave='scotiabank')
        self.assertEqual(creados, 2)

    def test_no_toca_los_de_fecha_estimada(self):
        """El histórico cargó al día 1 los movimientos cuyo día no venía en
        el correo. Ahí «mismo día y mismo monto» no prueba nada: pueden ser
        dos pagos distintos del mismo monto. Borrarlos destruía pagos reales
        (pasó en prod el 2026-08-10)."""
        from io import StringIO
        M = MovimientoFinanciero
        for i in (1, 2):
            M.objects.create(
                fecha=date(2026, 7, 1), cuenta=self.sc, clase='gasto',
                sentido='sale', monto=54240, categoria=self.cat,
                fuente='correo', referencia=f'hist:{i}', fecha_estimada=True,
                descripcion='Transferencia saliente Scotiabank (día no capturado)')
        call_command('revisar_duplicados', '--desde', '2026-07-01',
                     '--eliminar-de', 'scotiabank', stdout=StringIO())
        self.assertEqual(M.objects.filter(monto=54240).count(), 2)

    def test_dos_comisiones_de_ventas_distintas_no_son_duplicado(self):
        """Las comisiones de MP y SumUp llevan el id de la venta en la glosa.
        Agrupar por comercio normalizado se lo comía y dos comisiones de
        ventas distintas con el mismo monto parecían la misma."""
        from io import StringIO
        M = MovimientoFinanciero
        mp = CuentaFinanciera.objects.get(clave='mercado_pago')
        for ident in ('170402030994', '169395340059'):
            M.objects.create(
                fecha=date(2026, 7, 24), cuenta=mp, clase='gasto',
                sentido='sale', monto=7392, categoria=self.cat, fuente='api',
                referencia=f'mp:fee:{ident}',
                descripcion=f'Comisión MP del cobro {ident} (Reserva #1)')
        salida = StringIO()
        call_command('revisar_duplicados', '--desde', '2026-07-01',
                     '--eliminar-de', 'mercado_pago', stdout=salida)
        self.assertEqual(M.objects.filter(monto=7392).count(), 2)
        self.assertIn('MISMA CUENTA (doble carga): 0 casos', salida.getvalue())

    def test_dos_pagos_iguales_de_la_misma_carga_no_son_duplicado(self):
        """Dos transferencias de $20.000 a Martín el mismo día, con la misma
        glosa porque el pegado no lleva la hora. Entraron en la MISMA subida,
        así que el archivo traía las dos — no es doble carga (visto en prod
        2026-08-10)."""
        from io import StringIO
        M = MovimientoFinanciero
        for i in (1, 2):
            M.objects.create(
                fecha=date(2026, 8, 4), cuenta=self.rut, clase='gasto',
                sentido='sale', monto=20000, categoria=self.cat,
                fuente='captura', referencia=f'man:mar:{i}',
                descripcion='Cartola cuentarut jorge: Tef A Martin Aguilera')
        salida = StringIO()
        call_command('revisar_duplicados', '--desde', '2026-07-01',
                     '--eliminar-de', 'cuentarut_jorge', stdout=salida)
        self.assertEqual(M.objects.filter(monto=20000).count(), 2)
        self.assertIn('MISMA CUENTA (doble carga): 0 casos', salida.getvalue())

    def test_el_comando_separa_los_dos_tipos_de_duplicado(self):
        from io import StringIO
        M = MovimientoFinanciero
        # Doble carga: misma cuenta, y en DOS subidas distintas (por eso se
        # separa el creado_en: una doble carga nunca ocurre en el mismo
        # segundo).
        for i in (1, 2):
            m = M.objects.create(fecha=date(2026, 7, 20), cuenta=self.sc,
                                 clase='gasto', sentido='sale', monto=157433,
                                 categoria=self.cat, fuente='captura',
                                 referencia=f'dup:sc:{i}',
                                 descripcion='Cartola scotiabank: REDCOMPRA EXPRESS PUERTO')
            M.objects.filter(pk=m.pk).update(
                creado_en=timezone.now() - timedelta(hours=i))
        # Atribución: mismo cobro en dos cuentas distintas.
        for cuenta, i in ((self.rut, 3), (self.visa, 4)):
            M.objects.create(fecha=date(2026, 7, 14), cuenta=cuenta,
                             clase='gasto', sentido='sale', monto=180000,
                             categoria=self.cat, fuente='captura',
                             referencia=f'dup:g:{i}',
                             descripcion='Pago Google Ads Google')

        salida = StringIO()
        call_command('revisar_duplicados', '--desde', '2026-07-01',
                     stdout=salida)
        texto = salida.getvalue()
        self.assertIn('MISMA CUENTA (doble carga): 1 casos', texto)
        self.assertIn('CUENTAS DISTINTAS (atribución): 1 casos', texto)
        self.assertIn('Google Ads', texto)
        self.assertIn('Nada se borró', texto)
        self.assertEqual(M.objects.count(), 4)   # solo miró

    def test_eliminar_de_borra_solo_esa_cuenta_y_nunca_deja_vacio(self):
        from io import StringIO
        M = MovimientoFinanciero
        for cuenta, i in ((self.rut, 1), (self.visa, 2)):
            M.objects.create(fecha=date(2026, 7, 14), cuenta=cuenta,
                             clase='gasto', sentido='sale', monto=180000,
                             categoria=self.cat, fuente='captura',
                             referencia=f'el:{i}',
                             descripcion='Pago Google Ads Google')
        call_command('revisar_duplicados', '--desde', '2026-07-01',
                     '--eliminar-de', 'visa_2936', stdout=StringIO())
        # Se fue el de la Visa; queda el de la CuentaRUT, que es el bueno.
        self.assertEqual(M.objects.count(), 1)
        self.assertEqual(M.objects.first().cuenta.clave, 'cuentarut_jorge')

        # Si TODOS fueran de la cuenta objetivo, se conserva uno.
        for i in (5, 6):
            m = M.objects.create(fecha=date(2026, 7, 15), cuenta=self.visa,
                                 clase='gasto', sentido='sale', monto=9990,
                                 categoria=self.cat, fuente='captura',
                                 referencia=f'el:{i}', descripcion='ALGO')
            M.objects.filter(pk=m.pk).update(
                creado_en=timezone.now() - timedelta(hours=i))
        call_command('revisar_duplicados', '--desde', '2026-07-01',
                     '--eliminar-de', 'visa_2936', stdout=StringIO())
        self.assertEqual(M.objects.filter(monto=9990).count(), 1)
