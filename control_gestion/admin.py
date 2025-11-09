"""
Admin para Control de Gestión

Admin completo con inlines, acciones y validaciones para el sistema de tareas.
"""

from django.contrib import admin, messages
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.utils import timezone
import csv
from datetime import datetime
from .models import (
    Task, ChecklistItem, TaskLog, CustomerSegment, DailyReport,
    TaskState, Priority, TaskTemplate, EmpleadoDisponibilidad
)
from .forms import TaskAdminForm


class ChecklistInline(admin.TabularInline):
    """Inline para items de checklist"""
    model = ChecklistItem
    extra = 0
    fields = ('text', 'done')


class TaskLogInline(admin.TabularInline):
    """Inline para logs de la tarea (solo lectura)"""
    model = TaskLog
    extra = 0
    readonly_fields = ('when', 'actor', 'action', 'note')
    can_delete = False
    
    def has_add_permission(self, request, obj=None):
        return False


# ===== ACCIONES ADMIN =====

@admin.action(description="⬆️ Mover arriba en la cola")
def move_up(modeladmin, request, queryset):
    """Mueve las tareas seleccionadas hacia arriba en la cola"""
    for task in queryset:
        if task.queue_position > 1:
            task.queue_position -= 1
            task.save()
            TaskLog.objects.create(
                task=task,
                actor=request.user,
                action="REORDERED",
                note=f"Movida arriba en cola (posición {task.queue_position})"
            )
    messages.success(request, f"{queryset.count()} tarea(s) movida(s) arriba en la cola.")


@admin.action(description="⬇️ Mover abajo en la cola")
def move_down(modeladmin, request, queryset):
    """Mueve las tareas seleccionadas hacia abajo en la cola"""
    for task in queryset:
        task.queue_position += 1
        task.save()
        TaskLog.objects.create(
            task=task,
            actor=request.user,
            action="REORDERED",
            note=f"Movida abajo en cola (posición {task.queue_position})"
        )
    messages.success(request, f"{queryset.count()} tarea(s) movida(s) abajo en la cola.")


@admin.action(description="▶️ Marcar EN CURSO (respeta WIP=1)")
def mark_in_progress(modeladmin, request, queryset):
    """Marca tareas como en curso, respetando WIP=1"""
    errors = []
    success_count = 0
    
    for task in queryset:
        try:
            task.state = TaskState.IN_PROGRESS
            task.save()
            TaskLog.objects.create(
                task=task,
                actor=request.user,
                action="STARTED",
                note="Tarea iniciada"
            )
            success_count += 1
        except ValidationError as e:
            errors.append(f"{task.title}: {str(e)}")
    
    if success_count > 0:
        messages.success(request, f"✅ {success_count} tarea(s) marcada(s) en curso.")
    
    if errors:
        for error in errors:
            messages.error(request, error)


@admin.action(description="✅ Marcar HECHA")
def mark_done(modeladmin, request, queryset):
    """Marca tareas como completadas"""
    count = queryset.update(state=TaskState.DONE)
    
    for task in queryset:
        TaskLog.objects.create(
            task=task,
            actor=request.user,
            action="DONE",
            note="Tarea completada"
        )
    
    messages.success(request, f"✅ {count} tarea(s) completada(s).")


@admin.action(description="🚫 Marcar BLOQUEADA")
def mark_blocked(modeladmin, request, queryset):
    """Marca tareas como bloqueadas"""
    count = queryset.update(state=TaskState.BLOCKED)
    
    for task in queryset:
        TaskLog.objects.create(
            task=task,
            actor=request.user,
            action="BLOCKED",
            note="Tarea bloqueada"
        )
    
    messages.warning(request, f"🚫 {count} tarea(s) bloqueada(s).")


@admin.action(description="🤖 Generar checklist IA")
def ai_generate_checklist_action(modeladmin, request, queryset):
    """Genera checklist automático usando IA"""
    try:
        from . import ai
        
        count = 0
        for task in queryset:
            # Preparar contexto para IA
            ctx = {
                "swimlane": task.swimlane,
                "servicio": task.service_type,
                "ubicacion": task.location_ref,
                "titulo": task.title,
                "descripcion": task.description[:500]
            }
            
            try:
                items = ai.generate_checklist(ctx)
                
                for item_text in items:
                    ChecklistItem.objects.create(task=task, text=item_text)
                
                TaskLog.objects.create(
                    task=task,
                    actor=request.user,
                    action="COMMENT",
                    note=f"Checklist IA generado ({len(items)} items)"
                )
                count += 1
            except Exception as e:
                messages.warning(request, f"Error generando checklist para '{task.title}': {str(e)}")
        
        if count > 0:
            messages.success(request, f"🤖 Checklist IA generado para {count} tarea(s).")
    
    except ImportError:
        messages.error(request, "⚠️ Módulo de IA no disponible aún (se implementará en Etapa 2)")


@admin.action(description="📥 Exportar a CSV")
def export_to_csv(modeladmin, request, queryset):
    """Exporta las tareas seleccionadas a CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="tareas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Encabezados
    writer.writerow([
        'ID', 'Título', 'Área', 'Responsable', 'Estado', 'Prioridad',
        'Posición Cola', 'Fecha Creación', 'Fecha Actualización',
        'Promesa Entrega', 'Reserva ID', 'Teléfono Cliente',
        'Segmento', 'Ubicación', 'Tipo Servicio', 'Origen',
        'Checklist Items', 'Logs Count'
    ])
    
    # Datos
    for task in queryset:
        checklist_count = task.checklist.count()
        logs_count = task.logs.count()
        
        writer.writerow([
            task.id,
            task.title,
            task.get_swimlane_display(),
            task.owner.username if task.owner else '',
            task.get_state_display(),
            task.get_priority_display(),
            task.queue_position,
            task.created_at.strftime('%Y-%m-%d %H:%M:%S') if task.created_at else '',
            task.updated_at.strftime('%Y-%m-%d %H:%M:%S') if task.updated_at else '',
            task.promise_due_at.strftime('%Y-%m-%d %H:%M:%S') if task.promise_due_at else '',
            task.reservation_id,
            task.customer_phone_last9,
            task.segment_tag,
            task.get_location_ref_display() if task.location_ref else '',
            task.service_type,
            task.get_source_display(),
            checklist_count,
            logs_count
        ])
    
    return response


@admin.action(description="📊 Exportar a Excel")
def export_to_excel(modeladmin, request, queryset):
    """Exporta las tareas seleccionadas a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, "⚠️ openpyxl no está instalado. Instala con: pip install openpyxl")
        return
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tareas"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Encabezados
    headers = [
        'ID', 'Título', 'Área', 'Responsable', 'Estado', 'Prioridad',
        'Posición Cola', 'Fecha Creación', 'Fecha Actualización',
        'Promesa Entrega', 'Reserva ID', 'Teléfono Cliente',
        'Segmento', 'Ubicación', 'Tipo Servicio', 'Origen',
        'Checklist Items', 'Logs Count'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for row_num, task in enumerate(queryset, 2):
        checklist_count = task.checklist.count()
        logs_count = task.logs.count()
        
        ws.cell(row=row_num, column=1, value=task.id)
        ws.cell(row=row_num, column=2, value=task.title)
        ws.cell(row=row_num, column=3, value=task.get_swimlane_display())
        ws.cell(row=row_num, column=4, value=task.owner.username if task.owner else '')
        ws.cell(row=row_num, column=5, value=task.get_state_display())
        ws.cell(row=row_num, column=6, value=task.get_priority_display())
        ws.cell(row=row_num, column=7, value=task.queue_position)
        ws.cell(row=row_num, column=8, value=task.created_at.strftime('%Y-%m-%d %H:%M:%S') if task.created_at else '')
        ws.cell(row=row_num, column=9, value=task.updated_at.strftime('%Y-%m-%d %H:%M:%S') if task.updated_at else '')
        ws.cell(row=row_num, column=10, value=task.promise_due_at.strftime('%Y-%m-%d %H:%M:%S') if task.promise_due_at else '')
        ws.cell(row=row_num, column=11, value=task.reservation_id)
        ws.cell(row=row_num, column=12, value=task.customer_phone_last9)
        ws.cell(row=row_num, column=13, value=task.segment_tag)
        ws.cell(row=row_num, column=14, value=task.get_location_ref_display() if task.location_ref else '')
        ws.cell(row=row_num, column=15, value=task.service_type)
        ws.cell(row=row_num, column=16, value=task.get_source_display())
        ws.cell(row=row_num, column=17, value=checklist_count)
        ws.cell(row=row_num, column=18, value=logs_count)
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="tareas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


# ===== ADMIN DE TASK =====

@admin.register(Task)
class TaskAdmin(admin.ModelAdmin):
    form = TaskAdminForm  # Usar form personalizado con validación WIP=1
    
    list_display = (
        'title',
        'swimlane',
        'owner',
        'state',
        'priority',
        'queue_position',
        'promise_due_at',
        'updated_at',
        'reservation_id'
    )
    list_filter = (
        'swimlane',
        'state',
        'priority',
        'owner',
        'source',
        'created_at'
    )
    search_fields = (
        'title',
        'description',
        'reservation_id',
        'customer_phone_last9',
        'segment_tag'
    )
    
    inlines = [ChecklistInline, TaskLogInline]
    
    actions = [
        move_up,
        move_down,
        mark_in_progress,
        mark_done,
        mark_blocked,
        ai_generate_checklist_action,
        export_to_csv,
        export_to_excel
    ]
    
    readonly_fields = ('created_at', 'updated_at')
    
    list_per_page = 50
    
    fieldsets = (
        ('Información Básica', {
            'fields': ('title', 'description', 'swimlane', 'owner', 'created_by')
        }),
        ('Estado y Prioridad', {
            'fields': ('state', 'priority', 'queue_position', 'promise_due_at', 'source')
        }),
        ('Contexto de Reserva/Cliente', {
            'fields': ('reservation_id', 'customer_phone_last9', 'segment_tag'),
            'classes': ('collapse',)
        }),
        ('Ubicación y Servicio', {
            'fields': ('location_ref', 'service_type'),
            'classes': ('collapse',)
        }),
        ('Evidencia', {
            'fields': ('media',),
            'classes': ('collapse',)
        }),
        ('Fechas', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def save_model(self, request, obj, form, change):
        """Asignar created_by si es nuevo"""
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    # ===== PERMISOS POR GRUPO =====
    
    def has_view_permission(self, request, obj=None):
        """
        Controla quién puede ver tareas
        
        - ADMIN/SUPERUSER: Ve todas
        - SUPERVISION: Ve todas
        - Otros grupos: Solo sus propias tareas
        """
        if request.user.is_superuser:
            return True
        
        # Grupo SUPERVISION puede ver todas
        if request.user.groups.filter(name='SUPERVISION').exists():
            return True
        
        # Si hay objeto específico, verificar si es el owner
        if obj is not None:
            return obj.owner == request.user
        
        # Para listado, se filtra en get_queryset
        return True
    
    def has_change_permission(self, request, obj=None):
        """
        Controla quién puede modificar tareas
        
        - ADMIN/SUPERUSER: Puede modificar todas
        - SUPERVISION: Puede modificar todas
        - Owner: Solo puede modificar sus propias tareas
        """
        if request.user.is_superuser:
            return True
        
        # Grupo SUPERVISION puede modificar todas
        if request.user.groups.filter(name='SUPERVISION').exists():
            return True
        
        # Solo el owner puede modificar su tarea
        if obj is not None:
            return obj.owner == request.user
        
        # Para acciones masivas, permitir si tiene permisos básicos
        return True
    
    def has_delete_permission(self, request, obj=None):
        """
        Controla quién puede eliminar tareas
        
        - ADMIN/SUPERUSER: Puede eliminar todas
        - SUPERVISION: Puede eliminar todas
        - Owner: Solo puede eliminar sus propias tareas
        """
        if request.user.is_superuser:
            return True
        
        # Grupo SUPERVISION puede eliminar todas
        if request.user.groups.filter(name='SUPERVISION').exists():
            return True
        
        # Solo el owner puede eliminar su tarea
        if obj is not None:
            return obj.owner == request.user
        
        return True
    
    def get_queryset(self, request):
        """
        Filtra el queryset según los permisos del usuario
        
        - ADMIN/SUPERUSER: Ve todas
        - SUPERVISION: Ve todas
        - Otros: Solo sus propias tareas
        """
        qs = super().get_queryset(request)
        
        if request.user.is_superuser:
            return qs
        
        # Grupo SUPERVISION ve todas
        if request.user.groups.filter(name='SUPERVISION').exists():
            return qs
        
        # Otros usuarios solo ven sus propias tareas
        return qs.filter(owner=request.user)


# ===== ADMIN DE CUSTOMER SEGMENT =====

@admin.register(CustomerSegment)
class CustomerSegmentAdmin(admin.ModelAdmin):
    list_display = ('name', 'min_spend', 'max_spend', 'benefit')
    list_filter = ('name',)
    search_fields = ('name', 'benefit')
    ordering = ['min_spend']


# ===== ADMIN DE DAILY REPORT =====

@admin.register(DailyReport)
class DailyReportAdmin(admin.ModelAdmin):
    list_display = ('date', 'generated_at')
    list_filter = ('date',)
    search_fields = ('summary',)
    readonly_fields = ('generated_at',)
    ordering = ['-date']


# ===== ADMIN DE TASK TEMPLATE =====

@admin.register(TaskTemplate)
class TaskTemplateAdmin(admin.ModelAdmin):
    list_display = (
        'title_template',
        'swimlane',
        'get_dias_display',
        'asignar_a_grupo',
        'asignar_a_usuario',
        'activa',
        'solo_martes'
    )
    list_filter = ('swimlane', 'activa', 'solo_martes', 'asignar_a_grupo')
    search_fields = ('title_template', 'description')
    
    fieldsets = (
        ('Información de la Tarea', {
            'fields': ('title_template', 'description')
        }),
        ('Configuración', {
            'fields': ('swimlane', 'priority', 'queue_position')
        }),
        ('Recurrencia', {
            'fields': ('dias_activa', 'solo_martes', 'activa'),
            'description': 'Configurar qué días se genera esta tarea'
        }),
        ('Asignación', {
            'fields': ('asignar_a_grupo', 'asignar_a_usuario'),
            'description': 'A quién se asigna (grupo o usuario específico)'
        }),
    )
    
    def get_dias_display(self, obj):
        return obj.get_dias_str()
    get_dias_display.short_description = 'Días'
    
    actions = ['generar_tareas_ahora']
    
    @admin.action(description="🔄 Generar tareas AHORA de estas plantillas")
    def generar_tareas_ahora(self, request, queryset):
        """Genera tareas inmediatamente de las plantillas seleccionadas"""
        count = 0
        for template in queryset:
            if template.aplica_hoy():
                task = template.generar_tarea()
                if task:
                    count += 1
        
        if count > 0:
            messages.success(request, f"✅ {count} tarea(s) generada(s) desde plantillas")
        else:
            messages.warning(request, "⚠️ Ninguna plantilla aplica para hoy o ya fueron generadas")


# ===== ADMIN DE EMPLEADO DISPONIBILIDAD =====

@admin.register(EmpleadoDisponibilidad)
class EmpleadoDisponibilidadAdmin(admin.ModelAdmin):
    list_display = ('empleado', 'get_dias_display', 'trabaja_hoy_display', 'notas')
    list_filter = ('empleado__groups',)
    search_fields = ('empleado__username', 'empleado__first_name', 'notas')
    
    fieldsets = (
        ('Empleado', {
            'fields': ('empleado',)
        }),
        ('Disponibilidad', {
            'fields': ('dias_trabajo', 'notas'),
            'description': 'Configurar qué días trabaja este empleado'
        }),
    )
    
    def get_dias_display(self, obj):
        return obj.get_dias_str()
    get_dias_display.short_description = 'Días de trabajo'
    
    def trabaja_hoy_display(self, obj):
        trabaja = obj.trabaja_hoy()
        return "✅ Sí" if trabaja else "❌ No"
    trabaja_hoy_display.short_description = 'Trabaja hoy'

